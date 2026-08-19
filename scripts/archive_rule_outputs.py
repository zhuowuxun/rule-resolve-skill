#!/usr/bin/env python3
"""Archive rule workflow outputs into the project result tree.

Outputs should not stay in ~/Downloads. This helper moves generated workbooks
and reports into:

  ~/Documents/tag管理系统/结果文档/YYYY/MM/<category>/

For xlsx files it rewrites the workbook with keep_links=False to remove stale
external workbook links that can make Excel repair or refuse to open files.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Iterable


PROJECT_RESULT_ROOT = Path("/Users/carmenz/Documents/tag管理系统/结果文档")


def infer_year_month(paths: Iterable[Path]) -> tuple[str, str]:
    for path in paths:
        match = re.search(r"(20\d{2})[-_\.]?(0[1-9]|1[0-2])[-_\.]?([0-3]\d)?", path.name)
        if match:
            return match.group(1), match.group(2)
    newest = max((path.stat().st_mtime for path in paths if path.exists()), default=datetime.now().timestamp())
    dt = datetime.fromtimestamp(newest)
    return f"{dt.year:04d}", f"{dt.month:02d}"


def infer_category(paths: Iterable[Path], explicit: str | None = None) -> str:
    if explicit:
        return explicit
    names = " ".join(path.name.lower() for path in paths)
    if "mit_" in names or "mitigation" in names:
        return "mitigation"
    if "detection_rule" in names or "detection" in names:
        return "detection"
    if re.search(r"(?:^|[-_])t(?:_|[-])|validation|_cn-en", names):
        return "validation"
    if "tag" in names or "mastertable" in names:
        return "tag"
    return "process"


def unique_destination(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    index = 1
    while True:
        candidate = path.with_name(f"{stem}__dup{index}{suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def rewrite_xlsx_without_external_links(path: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to archive xlsx outputs") from exc

    wb = load_workbook(path, data_only=False, keep_links=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                _ = cell.value
    wb.save(path)

    with zipfile.ZipFile(path) as archive:
        external_parts = [name for name in archive.namelist() if "externalLinks" in name or "connections" in name]
    if external_parts:
        raise RuntimeError(f"External links remain after archive rewrite: {external_parts}")


def update_json_paths(path: Path, moved: dict[str, str]) -> None:
    if path.suffix.lower() != ".json":
        return
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return
    changed = False
    for key in ("output", "workbook", "xlsx"):
        value = data.get(key)
        if isinstance(value, str) and value in moved:
            data[key] = moved[value]
            changed = True
    if isinstance(data.get("source_consistency"), dict):
        value = data["source_consistency"].get("output")
        if isinstance(value, str) and value in moved:
            data["source_consistency"]["output"] = moved[value]
            changed = True
    if changed:
        data.setdefault("archived_from", moved)
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("files", nargs="+", type=Path, help="Generated output files to archive.")
    parser.add_argument("--source", type=Path, help="Original source workbook; used for year/month inference.")
    parser.add_argument("--category", help="Override output category folder.")
    parser.add_argument("--root", type=Path, default=PROJECT_RESULT_ROOT, help="Project result root.")
    args = parser.parse_args()

    files = [path.expanduser().resolve() for path in args.files]
    existing = [path for path in files if path.exists()]
    if not existing:
        raise FileNotFoundError("No output files exist to archive")

    infer_paths = existing + ([args.source.expanduser().resolve()] if args.source else [])
    year, month = infer_year_month(infer_paths)
    category = infer_category(infer_paths, args.category)
    dest_dir = args.root.expanduser() / year / month / category
    dest_dir.mkdir(parents=True, exist_ok=True)

    moved: dict[str, str] = {}
    archived: list[dict[str, str]] = []
    for source in existing:
        if source.suffix.lower() == ".xlsx":
            rewrite_xlsx_without_external_links(source)
        dest = unique_destination(dest_dir / source.name)
        shutil.move(str(source), str(dest))
        moved[str(source)] = str(dest)
        archived.append({"from": str(source), "to": str(dest)})

    for item in archived:
        dest = Path(item["to"])
        update_json_paths(dest, moved)

    print(json.dumps({"destination": str(dest_dir), "archived": archived}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
