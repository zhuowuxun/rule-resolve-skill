#!/usr/bin/env python3
import argparse
import json
import os
import re
import shlex
import shutil
import subprocess
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook


DEFAULT_REPO_ROOT = Path.home() / "Documents/翻译软件"
DEFAULT_API_BASE = "http://192.168.10.89"
DEFAULT_TRANSLATION_DICTS = ["专业名称翻译", "software翻译"]
DEFAULT_REPLACEMENT_DICTS = ["基础字符校对", "detection校对"]
DEFAULT_SOURCE_HEADERS = ["name.1", "desc", "notes"]
CN_RE = re.compile(r"[\u4e00-\u9fff]")
URL_RE = re.compile(r"https?://[^\s]+|www\.[^\s]+", re.IGNORECASE)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Create and run a detection translation project through AI Translation Studio."
    )
    parser.add_argument("--input", required=True, help="Source detection workbook (.xlsx)")
    parser.add_argument("--project-name", help="Platform project name. Defaults to the workbook stem.")
    parser.add_argument("--output", help="Exported bilingual workbook path.")
    parser.add_argument("--report", help="JSON summary report path.")
    parser.add_argument("--api-base", default=DEFAULT_API_BASE, help="AI Translation Studio backend URL.")
    parser.add_argument("--repo-root", default=str(DEFAULT_REPO_ROOT), help="Repo root path.")
    parser.add_argument(
        "--platform-ssh",
        default=os.environ.get("AI_TRANSLATION_PLATFORM_SSH", ""),
        help="SSH target for non-local platform DB operations, for example dx@192.168.10.89.",
    )
    parser.add_argument(
        "--platform-root",
        default=os.environ.get("AI_TRANSLATION_PLATFORM_ROOT", "/opt/Aitrans"),
        help="Remote AI Translation Studio root used with --platform-ssh.",
    )
    parser.add_argument(
        "--ssh-command",
        default=os.environ.get("AI_TRANSLATION_SSH_COMMAND", "ssh"),
        help="SSH command prefix. Example: 'sshpass -e ssh -o StrictHostKeyChecking=no'.",
    )
    parser.add_argument(
        "--source-headers",
        nargs="+",
        default=DEFAULT_SOURCE_HEADERS,
        help="Workbook headers to translate. Defaults to detection fields.",
    )
    parser.add_argument(
        "--translation-dicts",
        nargs="+",
        default=DEFAULT_TRANSLATION_DICTS,
        help="Translation dictionary names.",
    )
    parser.add_argument(
        "--replacement-dicts",
        nargs="+",
        default=DEFAULT_REPLACEMENT_DICTS,
        help="Replacement dictionary names.",
    )
    parser.add_argument("--skip-export", action="store_true", help="Skip bilingual Excel export.")
    parser.add_argument(
        "--keep-failed-project",
        action="store_true",
        help="Keep the platform project if translation fails before replacement/export.",
    )
    parser.add_argument(
        "--activate-google",
        action="store_true",
        help="Explicitly activate the Google Translate config. This changes platform-global model settings.",
    )
    parser.add_argument(
        "--manual-review-limit",
        type=int,
        default=15,
        help="Number of warnings to keep in the report.",
    )
    return parser.parse_args()


def ensure_ok(resp, context):
    try:
        resp.raise_for_status()
    except Exception as exc:
        body = resp.text[:1000] if resp is not None else ""
        raise RuntimeError(f"{context} failed: {exc}\n{body}") from exc
    return resp


def perform_request(session, method, url, context, **kwargs):
    try:
        resp = session.request(method, url, **kwargs)
    except requests.RequestException as exc:
        raise RuntimeError(
            f"{context} failed: could not reach {url}. Confirm the AI Translation Studio API base URL first."
        ) from exc
    return ensure_ok(resp, context)


def is_local_api(api_base):
    host = (urlparse(api_base).hostname or "").lower()
    return host in {"127.0.0.1", "localhost", "::1", "0.0.0.0"}


def run_remote(platform_ssh, ssh_command, remote_command, context):
    if not platform_ssh:
        raise RuntimeError(
            f"{context} requires remote platform access because api-base is not local. "
            "Pass --platform-ssh, for example --platform-ssh dx@192.168.10.89. "
            "Refusing to run local DB tools against a remote platform project."
        )
    cmd = shlex.split(ssh_command) + [platform_ssh, remote_command]
    return subprocess.run(cmd, capture_output=True, text=True, check=True)


def sanitize_project_name(name):
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", name.strip())
    cleaned = re.sub(r"-{2,}", "-", cleaned).strip("-")
    return cleaned or "detection-translate"


def detect_source_columns(xlsx_path, headers_needed):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    found = {header: set() for header in headers_needed}
    sheets = {}

    for ws in wb.worksheets:
        first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        headers = []
        if first_row:
            for cell in first_row:
                headers.append(str(cell.value).strip() if cell.value is not None else "")
        sheets[ws.title] = headers
        for idx, header in enumerate(headers):
            if header in found:
                found[header].add(idx)

    wb.close()

    missing = [header for header, indexes in found.items() if not indexes]
    if missing:
        raise RuntimeError(f"Missing required headers: {', '.join(missing)}")

    result = []
    for header in headers_needed:
        indexes = sorted(found[header])
        if len(indexes) != 1:
            raise RuntimeError(
                f"Header '{header}' appears in multiple different columns: {indexes}. "
                "Pass a workbook with stable detection columns."
            )
        result.append(indexes[0])

    return result, sheets


def get_json(session, url, context):
    resp = perform_request(session, "GET", url, context, timeout=60)
    return resp.json()


def post_json(session, url, payload, context):
    resp = perform_request(session, "POST", url, context, json=payload, timeout=300)
    return resp.json()


def put_json(session, url, payload, context):
    resp = perform_request(session, "PUT", url, context, json=payload, timeout=120)
    return resp.json()


def delete_project(session, api_base, project_id):
    try:
        perform_request(
            session,
            "DELETE",
            f"{api_base}/api/project/{project_id}",
            "Delete failed project",
            timeout=60,
        )
        return True
    except Exception:
        return False


def count_project_chunks(project):
    chunks = project.get("chunks")
    if isinstance(chunks, list):
        return len(chunks)
    try:
        return int(project.get("chunk_count") or 0)
    except (TypeError, ValueError):
        return 0


def ensure_translation_completed(result, expected_count, context):
    errors = result.get("errors") or []
    translated = int(result.get("translated") or 0)
    if expected_count <= 0:
        raise RuntimeError(f"{context} did not create any chunks. Check source headers/source_col before translating.")
    if errors or translated < expected_count:
        error_preview = json.dumps(errors[:5], ensure_ascii=False) if isinstance(errors, list) else str(errors)
        raise RuntimeError(
            f"{context} incomplete: translated {translated}/{expected_count}; errors={error_preview}. "
            "Stop here and check AI Translation Studio / Google Translate server configuration before replacement/export."
        )


def resolve_google_translate_config(session, api_base, activate=False):
    settings = get_json(session, f"{api_base}/api/settings/model", "Fetch model configs")
    configs = settings.get("configs", [])
    google_configs = [cfg for cfg in configs if cfg.get("provider") == "google_translate"]
    if not google_configs:
        raise RuntimeError("No Google Translate model config found. Configure one first in the platform.")

    active_id = settings.get("active_id")
    active = next((cfg for cfg in configs if cfg.get("id") == active_id), None)
    if active and active.get("provider") == "google_translate":
        return active

    selected = google_configs[0]
    if not activate:
        active_desc = f"{active.get('name')} ({active.get('provider')})" if active else "none"
        raise RuntimeError(
            "Active platform model is not Google Translate "
            f"(current: {active_desc}). Refusing to change global model settings silently. "
            "Activate Google Translate in the platform first, or rerun with --activate-google after user confirmation."
        )

    post_json(
        session,
        f"{api_base}/api/settings/model/{selected['id']}/activate",
        {},
        "Activate Google Translate config",
    )
    return selected


def resolve_dictionary_ids(session, api_base, names):
    dictionaries = get_json(session, f"{api_base}/api/dict/", "Fetch dictionaries")
    by_name = {item["name"]: item["id"] for item in dictionaries}
    missing = [name for name in names if name not in by_name]
    if missing:
        raise RuntimeError(f"Missing dictionaries: {', '.join(missing)}")
    return [by_name[name] for name in names]


def create_project(session, api_base, input_path, project_name, source_cols, translation_ids, replacement_ids):
    data = {
        "name": project_name,
        "target_lang": "EN",
        "source_type": "xlsx",
        "workflow": "translate",
        "translation_dict_ids": ",".join(str(i) for i in translation_ids),
        "replacement_dict_ids": ",".join(str(i) for i in replacement_ids),
        "source_col": ",".join(str(i) for i in source_cols),
    }
    with input_path.open("rb") as f:
        files = {
            "file": (
                input_path.name,
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        resp = perform_request(
            session,
            "POST",
            f"{api_base}/api/project/",
            "Create project",
            data=data,
            files=files,
            timeout=300,
        )
    return resp.json()


def backup_database(repo_root, project_name, api_base, platform_ssh, platform_root, ssh_command):
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_name = f"translator_before_{project_name}_detection_translate_{stamp}.db"

    if not is_local_api(api_base):
        root = shlex.quote(platform_root.rstrip("/"))
        backup_name_q = shlex.quote(backup_name)
        remote_command = (
            f"set -e; cd {root}; mkdir -p output/env-backup; "
            f"cp backend/instance/translator.db output/env-backup/{backup_name_q}; "
            f"printf '%s' output/env-backup/{backup_name_q}"
        )
        result = run_remote(platform_ssh, ssh_command, remote_command, "Remote database backup")
        return f"{platform_ssh}:{platform_root.rstrip('/')}/{result.stdout.strip()}"

    db_path = repo_root / "backend/instance/translator.db"
    if not db_path.exists():
        raise RuntimeError(f"Translator DB not found: {db_path}")
    backup_dir = repo_root / "output/env-backup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / backup_name
    shutil.copy2(db_path, backup_path)
    return backup_path


def run_detection_proofread(repo_root, project_id, api_base, platform_ssh, platform_root, ssh_command):
    if not is_local_api(api_base):
        root = shlex.quote(platform_root.rstrip("/"))
        project_id_q = shlex.quote(str(project_id))
        remote_py = "./backend/venv/bin/python3"
        remote_script = "tools/detection/check_and_fix.py"
        repair = run_remote(
            platform_ssh,
            ssh_command,
            f"set -e; cd {root}; {remote_py} {remote_script} {project_id_q} --repair",
            "Remote detection proofreading repair",
        )
        verify = run_remote(
            platform_ssh,
            ssh_command,
            f"set -e; cd {root}; {remote_py} {remote_script} {project_id_q}",
            "Remote detection proofreading verify",
        )
        return {
            "repair_stdout": repair.stdout,
            "verify_stdout": verify.stdout,
            "verify_clean": "没有发现任何问题" in verify.stdout,
            "target": f"{platform_ssh}:{platform_root.rstrip('/')}",
        }

    proofread_py = repo_root / "backend/venv/bin/python"
    proofread_script = repo_root / "tools/detection/check_and_fix.py"
    if not proofread_py.exists():
        raise RuntimeError(f"Backend venv Python not found: {proofread_py}")

    repair = subprocess.run(
        [str(proofread_py), str(proofread_script), str(project_id), "--repair"],
        cwd=repo_root,
        capture_output=True,
        text=True,
        check=True,
    )
    verify = subprocess.run(
        [str(proofread_py), str(proofread_script), str(project_id)],
        cwd=repo_root,
        capture_output=True,
        text=True,
        check=True,
    )
    return {
        "repair_stdout": repair.stdout,
        "verify_stdout": verify.stdout,
        "verify_clean": "没有发现任何问题" in verify.stdout,
        "target": str(repo_root),
    }


def audit_project(project, manual_review_limit):
    warnings = []
    rows = defaultdict(dict)

    for chunk in project.get("chunks", []):
        fmt = chunk.get("format_data")
        if not fmt:
            continue
        try:
            meta = json.loads(fmt)
        except Exception:
            continue
        row = meta.get("row")
        header = meta.get("header")
        if row and header:
            rows[row][header] = chunk.get("translated_text", "")

    for row, cells in sorted(rows.items()):
        for header in ("name.1", "desc", "notes"):
            text = cells.get(header, "")
            if not text:
                continue
            if CN_RE.search(text):
                warnings.append({"row": row, "header": header, "issue": "contains_chinese", "text": text})
            if re.search(r",(?=[A-Za-z/])", text):
                warnings.append({"row": row, "header": header, "issue": "comma_spacing", "text": text})
            if re.search(r",,|\.\.(?!/)", text):
                warnings.append({"row": row, "header": header, "issue": "double_punctuation", "text": text})
            if header == "name.1" and re.search(r"\bvulnerability\b", text):
                warnings.append({"row": row, "header": header, "issue": "lowercase_vulnerability", "text": text})

    return warnings[:manual_review_limit]


def export_bilingual(session, api_base, project_id, output_path):
    resp = perform_request(
        session,
        "GET",
        f"{api_base}/api/export/{project_id}?format=xlsx&bilingual=true",
        "Export bilingual workbook",
        timeout=600,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(resp.content)


def replace_text(text, old, new):
    if not isinstance(text, str) or old not in text:
        return text, False
    return text.replace(old, new), True


def regex_replace_text(text, pattern, repl):
    if not isinstance(text, str):
        return text, False
    new_text, count = re.subn(pattern, repl, text)
    return new_text, count > 0


def postprocess_exported_workbook(output_path):
    """Apply deterministic output-only fixes that the platform dictionaries miss.

    These fixes are intentionally narrow and evidence-backed by source columns or
    vendor URLs preserved in the exported workbook. They do not overwrite source
    Chinese columns.
    """
    wb = load_workbook(output_path)
    fixes = []

    for ws in wb.worksheets:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}
        english_headers = [h for h in ("name_en", "desc_en", "notes_en") if h in header_index]

        for row in range(2, ws.max_row + 1):
            source_text = " ".join(
                str(ws.cell(row, header_index[h]).value or "")
                for h in ("name.1", "desc", "notes")
                if h in header_index
            )
            source_lower = source_text.lower()
            urls = " ".join(URL_RE.findall(source_text)).lower()
            replacements = []

            if "fangmail.net" in urls:
                replacements.extend(
                    [
                        ("Directional Mail Gateway", "FangMail Email Security Gateway"),
                        ("Direction Mail Gateway", "FangMail Email Security Gateway"),
                        ("FangMail email gateway", "FangMail Email Security Gateway"),
                        ("FangMail Email Gateway", "FangMail Email Security Gateway"),
                    ]
                )

            if "macrowing.com" in urls:
                replacements.extend(
                    [
                        (
                            "Hongyi Software Electronic Document Management System",
                            "Macrowing Electronic Document Management System",
                        ),
                        (
                            "Hongyi Electronic Document Management System",
                            "Macrowing Electronic Document Management System",
                        ),
                        ("Hongyi Software", "Macrowing"),
                    ]
                )

            if "crawl4ai" in source_lower:
                replacements.append(("Crawl4ai", "Crawl4AI"))

            if "pythonreplcomponent" in source_lower:
                replacements.extend(
                    [
                        ("Python REPLComponent", "PythonREPLComponent"),
                        ("Python REPL Component", "PythonREPLComponent"),
                    ]
                )

            if "security-constraint" in source_lower:
                replacements.extend(
                    [
                        ("Apache Tomcat security constraint", "Apache Tomcat security-constraint"),
                        ("security constraint, Authentication", "security-constraint, Authentication"),
                        (
                            "When a single<security-constraint> Multiple extension patterns that share the same URL are defined.<web-resource-collection> When a block is selected,",
                            "When a single `<security-constraint>` defines multiple `<web-resource-collection>` blocks that share the same URL extension pattern,",
                        ),
                    ]
                )

            replacements.append(("Privilege Escalation漏洞", "Privilege Escalation Vulnerability"))

            for header in english_headers:
                cell = ws.cell(row, header_index[header])
                original = cell.value
                value = original
                for old, new in replacements:
                    value, changed = replace_text(value, old, new)
                    if changed:
                        fixes.append({"sheet": ws.title, "row": row, "header": header, "from": old, "to": new})

                if header == "name_en":
                    value, changed = regex_replace_text(value, r"\bstored XSS Vulnerability\b", "Stored XSS Vulnerability")
                    if changed:
                        fixes.append(
                            {
                                "sheet": ws.title,
                                "row": row,
                                "header": header,
                                "from": "stored XSS Vulnerability",
                                "to": "Stored XSS Vulnerability",
                            }
                        )

                value, changed = regex_replace_text(value, r"(^|[\s(])(/[^`\s]+)`(?=[,.;])", r"\1\2")
                if changed:
                    fixes.append(
                        {
                            "sheet": ws.title,
                            "row": row,
                            "header": header,
                            "from": "unopened stray backtick after path",
                            "to": "path without stray backtick",
                        }
                    )

                value, changed = replace_text(
                    value,
                    "causing Python `exec()` function to automatically injection the complete",
                    "causing the Python `exec()` function to automatically inject the complete",
                )
                if changed:
                    fixes.append(
                        {
                            "sheet": ws.title,
                            "row": row,
                            "header": header,
                            "from": "automatically injection",
                            "to": "automatically inject",
                        }
                    )

                if value != original:
                    cell.value = value

    if fixes:
        wb.save(output_path)
    return fixes


def audit_exported_workbook(output_path, manual_review_limit):
    warnings = []
    wb = load_workbook(output_path, read_only=True, data_only=True)

    for ws in wb.worksheets:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}
        english_headers = [h for h in ("name_en", "desc_en", "notes_en") if h in header_index]

        for row in range(2, ws.max_row + 1):
            source_text = " ".join(
                str(ws.cell(row, header_index[h]).value or "")
                for h in ("name.1", "desc", "notes")
                if h in header_index
            )
            urls = " ".join(URL_RE.findall(source_text)).lower()

            for header in english_headers:
                text = ws.cell(row, header_index[header]).value or ""
                if not isinstance(text, str) or not text:
                    continue
                if CN_RE.search(text):
                    warnings.append({"row": row, "header": header, "issue": "contains_chinese", "text": text})
                if re.search(r",(?=[A-Za-z/])", text):
                    warnings.append({"row": row, "header": header, "issue": "comma_spacing", "text": text})
                if re.search(r",,|\.\.(?!/)", text):
                    warnings.append({"row": row, "header": header, "issue": "double_punctuation", "text": text})
                if header == "name_en" and re.search(r"\bvulnerability\b", text):
                    warnings.append({"row": row, "header": header, "issue": "lowercase_vulnerability", "text": text})
                if "fangmail.net" in urls and "Directional Mail Gateway" in text:
                    warnings.append({"row": row, "header": header, "issue": "vendor_link_product_name", "text": text})
                if "macrowing.com" in urls and "Hongyi" in text:
                    warnings.append({"row": row, "header": header, "issue": "vendor_link_product_name", "text": text})
                if "crawl4ai" in source_text.lower() and "Crawl4ai" in text:
                    warnings.append({"row": row, "header": header, "issue": "product_case", "text": text})

    wb.close()
    return warnings[:manual_review_limit]


def main():
    args = parse_args()
    repo_root = Path(args.repo_root).expanduser().resolve()
    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    project_name = sanitize_project_name(args.project_name or input_path.stem)
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else input_path.with_name(f"{input_path.stem}_translated.xlsx")
    )
    report_path = (
        Path(args.report).expanduser().resolve()
        if args.report
        else output_path.with_name(f"{output_path.stem}_report.json")
    )

    source_cols, sheet_headers = detect_source_columns(input_path, args.source_headers)

    session = requests.Session()
    health = get_json(session, f"{args.api_base}/api/health", "Check backend health")
    if health.get("status") != "ok":
        raise RuntimeError(f"Backend is not healthy: {health}")
    if not is_local_api(args.api_base) and not args.platform_ssh:
        raise RuntimeError(
            "Remote platform API selected but --platform-ssh is missing. "
            "Pass --platform-ssh so DB backup and check_and_fix.py run on the same platform, "
            "or use an explicitly confirmed local API base."
        )

    google_cfg = resolve_google_translate_config(session, args.api_base, activate=args.activate_google)
    translation_ids = resolve_dictionary_ids(session, args.api_base, args.translation_dicts)
    replacement_ids = resolve_dictionary_ids(session, args.api_base, args.replacement_dicts)

    created = create_project(
        session,
        args.api_base,
        input_path,
        project_name,
        source_cols,
        translation_ids,
        replacement_ids,
    )
    project_id = created["id"]
    initial_project = get_json(session, f"{args.api_base}/api/project/{project_id}", "Fetch created project")
    expected_chunks = count_project_chunks(initial_project)

    try:
        translate_result = post_json(
            session,
            f"{args.api_base}/api/translate/{project_id}/translate-all",
            {},
            "Translate all chunks",
        )
        ensure_translation_completed(translate_result, expected_chunks, "Translate all chunks")
    except Exception:
        if not args.keep_failed_project:
            delete_project(session, args.api_base, project_id)
        raise
    replace_result = post_json(
        session,
        f"{args.api_base}/api/proofread/{project_id}/batch-replace-all",
        {},
        "Run batch replacement",
    )

    backup_path = backup_database(
        repo_root,
        project_name,
        args.api_base,
        args.platform_ssh,
        args.platform_root,
        args.ssh_command,
    )
    proofread_result = run_detection_proofread(
        repo_root,
        project_id,
        args.api_base,
        args.platform_ssh,
        args.platform_root,
        args.ssh_command,
    )
    put_json(
        session,
        f"{args.api_base}/api/project/{project_id}/status",
        {"status": "done"},
        "Set project status",
    )

    project = get_json(session, f"{args.api_base}/api/project/{project_id}", "Fetch final project")
    project_warnings = audit_project(project, args.manual_review_limit)
    export_fixes = []
    warnings = project_warnings

    if not args.skip_export:
        export_bilingual(session, args.api_base, project_id, output_path)
        export_fixes = postprocess_exported_workbook(output_path)
        warnings = audit_exported_workbook(output_path, args.manual_review_limit)

    report = {
        "project_id": project_id,
        "project_name": project_name,
        "input": str(input_path),
        "output": None if args.skip_export else str(output_path),
        "repo_root": str(repo_root),
        "api_base": args.api_base,
        "db_tools_target": proofread_result.get("target"),
        "google_config": {
            "id": google_cfg.get("id"),
            "name": google_cfg.get("name"),
            "provider": google_cfg.get("provider"),
            "model_name": google_cfg.get("model_name"),
        },
        "translation_dicts": list(zip(args.translation_dicts, translation_ids)),
        "replacement_dicts": list(zip(args.replacement_dicts, replacement_ids)),
        "source_headers": args.source_headers,
        "source_cols": source_cols,
        "sheet_headers": sheet_headers,
        "chunk_count": project.get("chunk_count"),
        "translate_result": {
            "translated": translate_result.get("translated"),
            "errors": translate_result.get("errors", []),
        },
        "replace_result": replace_result,
        "backup": str(backup_path),
        "proofread": {
            "verify_clean": proofread_result["verify_clean"],
            "repair_stdout_tail": proofread_result["repair_stdout"][-4000:],
            "verify_stdout": proofread_result["verify_stdout"],
        },
        "project_manual_review_warnings": project_warnings,
        "export_postprocess_fixes": export_fixes,
        "manual_review_warnings": warnings,
        "status": project.get("status"),
    }

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2))

    print(json.dumps({
        "project_id": project_id,
        "project_name": project_name,
        "output": None if args.skip_export else str(output_path),
        "report": str(report_path),
        "warnings": len(warnings),
        "verify_clean": proofread_result["verify_clean"],
    }, ensure_ascii=False, indent=2))

    if not proofread_result["verify_clean"]:
        raise SystemExit("Detection proofreading did not return a clean result.")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
