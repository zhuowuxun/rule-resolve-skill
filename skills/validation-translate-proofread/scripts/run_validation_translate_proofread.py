#!/usr/bin/env python3
import argparse
import json
import os
import re
import shlex
import shutil
import subprocess
import sys
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook


DEFAULT_REPO_ROOT = Path.home() / "Documents/翻译软件"
DEFAULT_API_BASE = "http://192.168.10.89"
DEFAULT_TRANSLATION_DICTS = ["专业名称翻译", "software翻译"]
DEFAULT_MAIN_REPLACEMENT_DICTS = ["基础字符校对", "validation校对"]
DEFAULT_NOTE_REPLACEMENT_DICTS = ["基础字符校对", "validation校对", "validation note replacement"]
DEFAULT_MAIN_SOURCE_HEADERS = ["cn_name", "cn_desc"]
DEFAULT_NOTE_SOURCE_HEADER = "cn_notes"
CN_RE = re.compile(r"[\u4e00-\u9fff]")
URL_RE = re.compile(r"https?://\S+")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Create and run a validation translate+proofread project through AI Translation Studio."
    )
    parser.add_argument("--input", required=True, help="Source validation workbook (.xlsx)")
    parser.add_argument("--project-name", help="Platform project name. Defaults to the workbook stem.")
    parser.add_argument("--output", help="Exported bilingual workbook path.")
    parser.add_argument("--report", help="JSON summary report path.")
    parser.add_argument("--api-base", default=DEFAULT_API_BASE, help="AI Translation Studio backend URL.")
    parser.add_argument("--repo-root", default=str(DEFAULT_REPO_ROOT), help="Repo root path.")
    parser.add_argument(
        "--platform-ssh",
        default=os.environ.get("AI_TRANSLATION_PLATFORM_SSH", ""),
        help="SSH target for non-local platform DB operations, for example dx@192.168.10.89.",
    )
    parser.add_argument(
        "--platform-root",
        default=os.environ.get("AI_TRANSLATION_PLATFORM_ROOT", "/opt/Aitrans"),
        help="Remote AI Translation Studio root used with --platform-ssh.",
    )
    parser.add_argument(
        "--ssh-command",
        default=os.environ.get("AI_TRANSLATION_SSH_COMMAND", "ssh"),
        help="SSH command prefix. Example: 'sshpass -e ssh -o StrictHostKeyChecking=no'.",
    )
    parser.add_argument(
        "--translation-dicts",
        nargs="+",
        default=DEFAULT_TRANSLATION_DICTS,
        help="Validation translation dictionary names.",
    )
    parser.add_argument(
        "--main-replacement-dicts",
        nargs="+",
        default=DEFAULT_MAIN_REPLACEMENT_DICTS,
        help="Replacement dictionaries for the main validation project.",
    )
    parser.add_argument(
        "--note-replacement-dicts",
        nargs="+",
        default=DEFAULT_NOTE_REPLACEMENT_DICTS,
        help="Replacement dictionaries for the notes-only validation project.",
    )
    parser.add_argument("--skip-export", action="store_true", help="Skip bilingual Excel export.")
    parser.add_argument(
        "--keep-failed-project",
        action="store_true",
        help="Keep the platform project if translation fails before replacement/export.",
    )
    parser.add_argument(
        "--activate-google",
        action="store_true",
        help="Explicitly activate the Google Translate config. This changes platform-global model settings.",
    )
    parser.add_argument(
        "--manual-review-limit",
        type=int,
        default=20,
        help="Number of QA warnings to keep in the report.",
    )
    return parser.parse_args()


def ensure_ok(resp, context):
    try:
        resp.raise_for_status()
    except Exception as exc:
        body = resp.text[:1000] if resp is not None else ""
        raise RuntimeError(f"{context} failed: {exc}\n{body}") from exc
    return resp


def perform_request(session, method, url, context, **kwargs):
    try:
        resp = session.request(method, url, **kwargs)
    except requests.RequestException as exc:
        raise RuntimeError(
            f"{context} failed: could not reach {url}. Confirm the AI Translation Studio API base URL first."
        ) from exc
    return ensure_ok(resp, context)


def is_local_api(api_base):
    host = (urlparse(api_base).hostname or "").lower()
    return host in {"127.0.0.1", "localhost", "::1", "0.0.0.0"}


def run_remote(platform_ssh, ssh_command, remote_command, context):
    if not platform_ssh:
        raise RuntimeError(
            f"{context} requires remote platform access because api-base is not local. "
            "Pass --platform-ssh, for example --platform-ssh dx@192.168.10.89. "
            "Refusing to run local DB tools against a remote platform project."
        )
    cmd = shlex.split(ssh_command) + [platform_ssh, remote_command]
    return subprocess.run(cmd, capture_output=True, text=True, check=True)


def get_json(session, url, context):
    return perform_request(session, "GET", url, context, timeout=120).json()


def post_json(session, url, payload, context):
    return perform_request(session, "POST", url, context, json=payload, timeout=1800).json()


def put_json(session, url, payload, context):
    return perform_request(session, "PUT", url, context, json=payload, timeout=1800).json()


def delete_project(session, api_base, project_id):
    try:
        perform_request(
            session,
            "DELETE",
            f"{api_base}/api/project/{project_id}",
            "Delete failed project",
            timeout=60,
        )
        return True
    except Exception:
        return False


def count_project_chunks(project):
    chunks = project.get("chunks")
    if isinstance(chunks, list):
        return len(chunks)
    try:
        return int(project.get("chunk_count") or 0)
    except (TypeError, ValueError):
        return 0


def ensure_translation_completed(result, expected_count, context):
    errors = result.get("errors") or []
    translated = int(result.get("translated") or 0)
    if expected_count <= 0:
        raise RuntimeError(f"{context} did not create any chunks. Check source headers/source_col before translating.")
    if errors or translated < expected_count:
        error_preview = json.dumps(errors[:5], ensure_ascii=False) if isinstance(errors, list) else str(errors)
        raise RuntimeError(
            f"{context} incomplete: translated {translated}/{expected_count}; errors={error_preview}. "
            "Stop here and check AI Translation Studio / Google Translate server configuration before replacement/export."
        )


def translate_all_with_retries(session, api_base, project_id, expected_chunks, max_attempts=4):
    last_result = None
    for attempt in range(1, max_attempts + 1):
        last_result = post_json(
            session,
            f"{api_base}/api/translate/{project_id}/translate-all",
            {},
            f"Translate all chunks attempt {attempt}",
        )
        translated = int(last_result.get("translated") or 0)
        errors = last_result.get("errors") or []
        if translated >= expected_chunks and not errors:
            return last_result

        # The platform skips chunks that already have translated_text, so a
        # second call safely retries only the chunks that failed transiently.
        if attempt < max_attempts:
            continue

    ensure_translation_completed(last_result or {}, expected_chunks, "Translate all chunks")
    return last_result


def verify_note_replacement_scope(repo_root, replacement_dict_names, api_base, platform_ssh, platform_root, ssh_command):
    note_dict_requested = any(name.strip().lower() == "validation note replacement" for name in replacement_dict_names)
    if not note_dict_requested:
        return

    if not is_local_api(api_base):
        root = shlex.quote(platform_root.rstrip("/"))
        remote_command = (
            f"set -e; cd {root}; "
            "grep -q 'validation note replacement' backend/services/check_flow.py; "
            "grep -q 'NOTE_ONLY_DICTIONARY_NAMES' backend/services/check_flow.py; "
            "grep -q 'cn_notes' backend/services/check_flow.py"
        )
        run_remote(platform_ssh, ssh_command, remote_command, "Remote validation note replacement scope check")
        return

    check_flow_path = repo_root / "backend/services/check_flow.py"
    if not check_flow_path.exists():
        raise RuntimeError(
            f"Cannot verify note-only replacement scoping because check_flow.py is missing: {check_flow_path}"
        )

    source = check_flow_path.read_text(encoding="utf-8", errors="replace")
    has_scope_guard = (
        "validation note replacement" in source
        and "NOTE_ONLY_DICTIONARY_NAMES" in source
        and "cn_notes" in source
        and "continue" in source
    )
    if not has_scope_guard:
        raise RuntimeError(
            "AI Translation Studio backend does not appear to enforce validation note replacement as cn_notes-only. "
            "Update backend/services/check_flow.py before running validation translation; do not split the workbook into multiple platform projects."
        )


def sanitize_project_name(name):
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", name.strip())
    cleaned = re.sub(r"-{2,}", "-", cleaned).strip("-")
    return cleaned or "validation-translate-proofread"


def collect_sheet_headers(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_headers = {}
    for ws in wb.worksheets:
        first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        headers = []
        if first_row:
            for cell in first_row:
                headers.append(str(cell.value).strip() if cell.value is not None else "")
        sheet_headers[ws.title] = headers
    wb.close()
    return sheet_headers


def prepare_platform_input_workbook(input_path, source_headers, source_cols):
    wb = load_workbook(input_path)
    selected_cols = set(source_cols)
    source_header_set = set(source_headers)
    blanked_cells = 0

    for ws in wb.worksheets:
        headers = [ws.cell(1, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        for col_idx_zero in selected_cols:
            col_idx = col_idx_zero + 1
            header = headers[col_idx_zero] if col_idx_zero < len(headers) else ""
            if header in source_header_set:
                continue
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.value not in (None, ""):
                    cell.value = None
                    blanked_cells += 1

    if not blanked_cells:
        return input_path, None, 0

    tmpdir = tempfile.TemporaryDirectory(prefix="validation-upload-")
    upload_path = Path(tmpdir.name) / input_path.name
    wb.save(upload_path)
    return upload_path, tmpdir, blanked_cells


def detect_stable_column(sheet_headers, header_name, limit_sheets=None):
    indexes = set()
    found_sheets = []
    for sheet_name, headers in sheet_headers.items():
        if limit_sheets and sheet_name not in limit_sheets:
            continue
        if header_name in headers:
            indexes.add(headers.index(header_name))
            found_sheets.append(sheet_name)
    if not found_sheets:
        return None, []
    if len(indexes) != 1:
        raise RuntimeError(f"Header '{header_name}' appears in different columns across sheets: {sorted(indexes)}")
    return next(iter(indexes)), found_sheets


def resolve_google_translate_config(session, api_base, activate=False):
    settings = get_json(session, f"{api_base}/api/settings/model", "Fetch model configs")
    configs = settings.get("configs", [])
    google_configs = [cfg for cfg in configs if cfg.get("provider") == "google_translate"]
    if not google_configs:
        raise RuntimeError("No Google Translate model config found. Configure one first in the platform.")

    active_id = settings.get("active_id")
    active = next((cfg for cfg in configs if cfg.get("id") == active_id), None)
    if active and active.get("provider") == "google_translate":
        return active

    selected = google_configs[0]
    if not activate:
        active_desc = f"{active.get('name')} ({active.get('provider')})" if active else "none"
        raise RuntimeError(
            "Active platform model is not Google Translate "
            f"(current: {active_desc}). Refusing to change global model settings silently. "
            "Activate Google Translate in the platform first, or rerun with --activate-google after user confirmation."
        )

    post_json(
        session,
        f"{api_base}/api/settings/model/{selected['id']}/activate",
        {},
        "Activate Google Translate config",
    )
    return selected


def resolve_dictionary_ids(session, api_base, names):
    dictionaries = get_json(session, f"{api_base}/api/dict/", "Fetch dictionaries")
    by_name = {item["name"]: item["id"] for item in dictionaries}
    missing = [name for name in names if name not in by_name]
    if missing:
        raise RuntimeError(f"Missing dictionaries: {', '.join(missing)}")
    return [by_name[name] for name in names]


def create_project(session, api_base, input_path, project_name, source_cols, translation_ids, replacement_ids):
    data = {
        "name": project_name,
        "target_lang": "EN",
        "source_type": "xlsx",
        "workflow": "translate",
        "translation_dict_ids": ",".join(str(i) for i in translation_ids),
        "replacement_dict_ids": ",".join(str(i) for i in replacement_ids),
        "source_col": ",".join(str(i) for i in source_cols),
    }
    with input_path.open("rb") as f:
        files = {
            "file": (
                input_path.name,
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        resp = perform_request(
            session,
            "POST",
            f"{api_base}/api/project/",
            "Create project",
            data=data,
            files=files,
            timeout=1800,
        )
    return resp.json()


def backup_database(repo_root, project_name, api_base, platform_ssh, platform_root, ssh_command):
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_name = f"translator_before_{project_name}_translate_validation_{stamp}.db"

    if not is_local_api(api_base):
        root = shlex.quote(platform_root.rstrip("/"))
        backup_name_q = shlex.quote(backup_name)
        remote_command = (
            f"set -e; cd {root}; mkdir -p output/env-backup; "
            f"cp backend/instance/translator.db output/env-backup/{backup_name_q}; "
            f"printf '%s' output/env-backup/{backup_name_q}"
        )
        result = run_remote(platform_ssh, ssh_command, remote_command, "Remote database backup")
        return f"{platform_ssh}:{platform_root.rstrip('/')}/{result.stdout.strip()}"

    db_path = repo_root / "backend/instance/translator.db"
    if not db_path.exists():
        raise RuntimeError(f"Translator DB not found: {db_path}")
    backup_dir = repo_root / "output/env-backup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / backup_name
    shutil.copy2(db_path, backup_path)
    return backup_path


def run_validation_proofread(repo_root, project_id, api_base, platform_ssh, platform_root, ssh_command):
    if not is_local_api(api_base):
        root = shlex.quote(platform_root.rstrip("/"))
        project_id_q = shlex.quote(str(project_id))
        remote_py = "./backend/venv/bin/python3"
        remote_script = "tools/validation/check_and_fix.py"
        repair = run_remote(
            platform_ssh,
            ssh_command,
            f"set -e; cd {root}; {remote_py} {remote_script} {project_id_q} --repair",
            "Remote validation proofreading repair",
        )
        verify = run_remote(
            platform_ssh,
            ssh_command,
            f"set -e; cd {root}; {remote_py} {remote_script} {project_id_q}",
            "Remote validation proofreading verify",
        )
        return {
            "repair_stdout": repair.stdout,
            "verify_stdout": verify.stdout,
            "verify_clean": "没有发现任何问题" in verify.stdout,
            "target": f"{platform_ssh}:{platform_root.rstrip('/')}",
        }

    proofread_py = repo_root / "backend/venv/bin/python"
    proofread_script = repo_root / "tools/validation/check_and_fix.py"
    if not proofread_py.exists():
        raise RuntimeError(f"Backend venv Python not found: {proofread_py}")

    repair = subprocess.run(
        [str(proofread_py), str(proofread_script), str(project_id), "--repair"],
        cwd=repo_root,
        capture_output=True,
        text=True,
        check=True,
    )
    verify = subprocess.run(
        [str(proofread_py), str(proofread_script), str(project_id)],
        cwd=repo_root,
        capture_output=True,
        text=True,
        check=True,
    )
    return {
        "repair_stdout": repair.stdout,
        "verify_stdout": verify.stdout,
        "verify_clean": "没有发现任何问题" in verify.stdout,
        "target": str(repo_root),
    }


def target_header_for_source(header):
    if isinstance(header, str) and header.startswith("cn_"):
        return "en_" + header[3:]
    return None


def export_bilingual(session, api_base, project_id, input_path, output_path):
    project = get_json(session, f"{api_base}/api/project/{project_id}", "Fetch project chunks for export")
    chunks = project.get("chunks") or []
    wb = load_workbook(input_path)

    header_maps = {}
    for ws in wb.worksheets:
        headers = [ws.cell(1, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        header_maps[ws.title] = {header: col_idx + 1 for col_idx, header in enumerate(headers) if header}

    applied = 0
    skipped = []
    for chunk in chunks:
        translated = chunk.get("translated_text") or ""
        fmt_raw = chunk.get("format_data") or ""
        try:
            meta = json.loads(fmt_raw) if isinstance(fmt_raw, str) else fmt_raw
        except Exception:
            skipped.append({"chunk": chunk.get("id"), "reason": "bad_format_data"})
            continue
        if not isinstance(meta, dict) or meta.get("type") != "xlsx_cell":
            skipped.append({"chunk": chunk.get("id"), "reason": "not_xlsx_cell"})
            continue
        sheet = meta.get("sheet")
        row = meta.get("row")
        source_header = meta.get("header")
        target_header = target_header_for_source(source_header)
        if sheet not in wb.sheetnames or not target_header:
            skipped.append({"chunk": chunk.get("id"), "reason": "no_target", "sheet": sheet, "header": source_header})
            continue
        target_col = header_maps.get(sheet, {}).get(target_header)
        if not target_col:
            skipped.append({"chunk": chunk.get("id"), "reason": "missing_target_col", "sheet": sheet, "target": target_header})
            continue
        wb[sheet].cell(int(row), target_col).value = translated
        applied += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {"applied_chunks": applied, "skipped_chunks": skipped}


def _split_note_reference_block(text, markers):
    if not isinstance(text, str) or not text.strip():
        return None

    marker_positions = [text.find(marker) for marker in markers if marker in text]
    marker_positions = [pos for pos in marker_positions if pos != -1]
    if not marker_positions:
        return None

    start = min(marker_positions)
    note_body = text[:start].rstrip()
    tail = text[start:]
    urls = URL_RE.findall(tail)
    if not urls:
        return None

    skeleton = tail
    for url in urls:
        skeleton = skeleton.replace(url, "", 1)
    for marker in markers:
        skeleton = skeleton.replace(marker, "")
    skeleton = re.sub(r"[:：，,\s]+", "", skeleton)
    if skeleton:
        return None

    return note_body, urls


def _normalize_desc_reference_block(text, markers, preferred_marker):
    if not isinstance(text, str) or not text.strip():
        return text

    marker_positions = [text.rfind(marker) for marker in markers if marker in text]
    marker_positions = [pos for pos in marker_positions if pos != -1]
    if not marker_positions:
        return text

    start = max(marker_positions)
    prefix = text[:start].rstrip()
    tail = text[start:]
    urls = URL_RE.findall(tail)
    if not urls:
        return text

    skeleton = tail
    for url in urls:
        skeleton = skeleton.replace(url, "", 1)
    for marker in markers:
        skeleton = skeleton.replace(marker, "")
    skeleton = re.sub(r"[:：，,\s]+", "", skeleton)
    if skeleton:
        return text

    ref_block = f"{preferred_marker}\n\n" + "\n".join(urls)
    return f"{prefix}\n\n{ref_block}" if prefix else ref_block


def relocate_reference_links_from_notes(output_path):
    wb = load_workbook(output_path)
    relocated_rows = 0
    normalized_rows = 0

    for ws in wb.worksheets:
        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            continue
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]
        required = {"cn_desc", "cn_notes", "en_desc", "en_notes"}
        if not required.issubset(set(headers)):
            continue

        cn_desc_col = headers.index("cn_desc") + 1
        cn_notes_col = headers.index("cn_notes") + 1
        en_desc_col = headers.index("en_desc") + 1
        en_notes_col = headers.index("en_notes") + 1

        for row_idx in range(2, ws.max_row + 1):
            cn_desc = ws.cell(row_idx, cn_desc_col).value or ""
            cn_notes = ws.cell(row_idx, cn_notes_col).value or ""
            en_desc = ws.cell(row_idx, en_desc_col).value or ""
            en_notes = ws.cell(row_idx, en_notes_col).value or ""

            cn_split = _split_note_reference_block(cn_notes, ["参考链接", "参考链接：", "参考链接:"])
            en_split = _split_note_reference_block(
                en_notes,
                ["Please refer to", "Please refer to:", "Reference link", "Reference link:", "Reference links", "Reference links:"],
            )
            if not cn_split and not en_split:
                continue

            cn_note_body, cn_urls = cn_split if cn_split else (cn_notes.rstrip(), [])
            en_note_body, en_urls = en_split if en_split else (en_notes.rstrip(), [])
            urls = cn_urls or en_urls
            if not urls:
                continue

            if not URL_RE.search(str(cn_desc)):
                cn_desc = str(cn_desc).rstrip()
                cn_desc = f"{cn_desc}\n\n参考链接：\n\n" + "\n".join(urls) if cn_desc else "参考链接：\n\n" + "\n".join(urls)
                ws.cell(row_idx, cn_desc_col).value = cn_desc
            if not URL_RE.search(str(en_desc)):
                en_desc = str(en_desc).rstrip()
                en_desc = f"{en_desc}\n\nPlease refer to:\n\n" + "\n".join(urls) if en_desc else "Please refer to:\n\n" + "\n".join(urls)
                ws.cell(row_idx, en_desc_col).value = en_desc

            ws.cell(row_idx, cn_notes_col).value = cn_note_body
            ws.cell(row_idx, en_notes_col).value = en_note_body
            relocated_rows += 1

        for row_idx in range(2, ws.max_row + 1):
            cn_desc = ws.cell(row_idx, cn_desc_col).value or ""
            en_desc = ws.cell(row_idx, en_desc_col).value or ""
            normalized_cn_desc = _normalize_desc_reference_block(cn_desc, ["参考链接", "参考链接：", "参考链接:"], "参考链接：")
            normalized_en_desc = _normalize_desc_reference_block(
                en_desc,
                ["Please refer to", "Please refer to:", "Reference link", "Reference link:", "Reference links", "Reference links:"],
                "Please refer to:",
            )
            if normalized_cn_desc != cn_desc:
                ws.cell(row_idx, cn_desc_col).value = normalized_cn_desc
                normalized_rows += 1
            if normalized_en_desc != en_desc:
                ws.cell(row_idx, en_desc_col).value = normalized_en_desc
                normalized_rows += 1

    wb.save(output_path)
    wb.close()
    return {"relocated_rows": relocated_rows, "normalized_rows": normalized_rows}


def audit_workbook(path, manual_review_limit):
    wb = load_workbook(path, read_only=True, data_only=True)
    warnings = []
    reference_note_rows = []
    title_headers = {"en_name", "en_subject", "name_en"}
    rows = defaultdict(dict)

    for ws in wb.worksheets:
        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        headers = []
        if header_row:
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_map = {}
            for idx, header in enumerate(headers):
                hk = header.lower()
                row_map[hk] = row[idx] if idx < len(row) else None
                if not (hk.startswith("en_") or hk.endswith("_en")):
                    continue
                text = row[idx]
                if text in (None, ""):
                    continue
                rows[(ws.title, row_idx)][hk] = str(text)
            cn_notes_val = row_map.get("cn_notes")
            en_notes_val = row_map.get("en_notes")
            if (
                isinstance(cn_notes_val, str)
                and "参考链接" in cn_notes_val
                or isinstance(en_notes_val, str)
                and ("Reference link" in en_notes_val or "Please refer to" in en_notes_val)
            ):
                reference_note_rows.append({"sheet": ws.title, "row": row_idx})

    wb.close()

    for (sheet_name, row_num), cells in sorted(rows.items()):
        for header, text in sorted(cells.items()):
            if CN_RE.search(text):
                warnings.append({"sheet": sheet_name, "row": row_num, "header": header, "issue": "contains_chinese", "text": text})
            if re.search(r",(?=[A-Za-z/])", text):
                warnings.append({"sheet": sheet_name, "row": row_num, "header": header, "issue": "comma_spacing", "text": text})
            if re.search(r",,|\.\.(?!/)", text):
                warnings.append({"sheet": sheet_name, "row": row_num, "header": header, "issue": "double_punctuation", "text": text})
            if header in title_headers and re.search(r"\bvulnerability\b", text):
                warnings.append({"sheet": sheet_name, "row": row_num, "header": header, "issue": "lowercase_vulnerability", "text": text})
            if header in title_headers and text.rstrip().endswith("."):
                warnings.append({"sheet": sheet_name, "row": row_num, "header": header, "issue": "title_trailing_period", "text": text})
            if header in title_headers and re.search(r"\b(a|an|the)\b", text):
                warnings.append({"sheet": sheet_name, "row": row_num, "header": header, "issue": "title_article", "text": text})
            if header in {"en_desc", "en_notes"} and re.search(r"Please refer to:\nhttps?://", text):
                warnings.append({"sheet": sheet_name, "row": row_num, "header": header, "issue": "reference_link_missing_blank_line_after_marker", "text": text})

    for item in reference_note_rows:
        warnings.append({"sheet": item["sheet"], "row": item["row"], "header": "notes", "issue": "reference_link_left_in_notes"})

    return warnings[:manual_review_limit]


def main():
    args = parse_args()
    repo_root = Path(args.repo_root).expanduser().resolve()
    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    project_name = sanitize_project_name(args.project_name or input_path.stem)
    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else input_path.with_name(f"{input_path.stem}_translated.xlsx")
    )
    report_path = (
        Path(args.report).expanduser().resolve()
        if args.report
        else output_path.with_name(f"{output_path.stem}_report.json")
    )

    sheet_headers = collect_sheet_headers(input_path)
    main_cols = []
    for header in DEFAULT_MAIN_SOURCE_HEADERS:
        col_idx, found_sheets = detect_stable_column(sheet_headers, header)
        if col_idx is None:
            raise RuntimeError(f"Missing required validation header: {header}")
        if not found_sheets:
            raise RuntimeError(f"Header '{header}' was not found in any sheet")
        main_cols.append(col_idx)

    notes_col, note_sheets = detect_stable_column(sheet_headers, DEFAULT_NOTE_SOURCE_HEADER)

    session = requests.Session()
    health = get_json(session, f"{args.api_base}/api/health", "Check backend health")
    if health.get("status") != "ok":
        raise RuntimeError(f"Backend is not healthy: {health}")
    if not is_local_api(args.api_base) and not args.platform_ssh:
        raise RuntimeError(
            "Remote platform API selected but --platform-ssh is missing. "
            "Pass --platform-ssh so DB backup and check_and_fix.py run on the same platform, "
            "or use an explicitly confirmed local API base."
        )

    google_cfg = resolve_google_translate_config(session, args.api_base, activate=args.activate_google)
    translation_ids = resolve_dictionary_ids(session, args.api_base, args.translation_dicts)
    replacement_dict_names = list(dict.fromkeys(args.main_replacement_dicts + args.note_replacement_dicts))
    replacement_ids = resolve_dictionary_ids(session, args.api_base, replacement_dict_names)
    verify_note_replacement_scope(
        repo_root,
        replacement_dict_names,
        args.api_base,
        args.platform_ssh,
        args.platform_root,
        args.ssh_command,
    )

    source_headers = list(DEFAULT_MAIN_SOURCE_HEADERS)
    source_cols = list(main_cols)
    if notes_col is not None and note_sheets:
        source_headers.append(DEFAULT_NOTE_SOURCE_HEADER)
        source_cols.append(notes_col)

    platform_input_path, upload_tmpdir, blanked_upload_cells = prepare_platform_input_workbook(
        input_path,
        source_headers,
        source_cols,
    )
    try:
        created = create_project(
            session,
            args.api_base,
            platform_input_path,
            project_name,
            source_cols,
            translation_ids,
            replacement_ids,
        )
    finally:
        if upload_tmpdir is not None:
            upload_tmpdir.cleanup()
    project_id = created["id"]
    initial_project = get_json(
        session,
        f"{args.api_base}/api/project/{project_id}",
        "Fetch created project",
    )
    expected_chunks = count_project_chunks(initial_project)

    try:
        translate_result = translate_all_with_retries(session, args.api_base, project_id, expected_chunks)
    except Exception:
        if not args.keep_failed_project:
            delete_project(session, args.api_base, project_id)
        raise
    replace_result = post_json(
        session,
        f"{args.api_base}/api/proofread/{project_id}/batch-replace-all",
        {},
        "Run batch replacement",
    )

    backup_path = backup_database(
        repo_root,
        project_name,
        args.api_base,
        args.platform_ssh,
        args.platform_root,
        args.ssh_command,
    )
    proofread_result = run_validation_proofread(
        repo_root,
        project_id,
        args.api_base,
        args.platform_ssh,
        args.platform_root,
        args.ssh_command,
    )
    if not proofread_result["verify_clean"]:
        raise SystemExit("Validation proofreading did not return a clean result.")

    put_json(
        session,
        f"{args.api_base}/api/project/{project_id}/status",
        {"status": "done"},
        "Set project status",
    )

    project = get_json(session, f"{args.api_base}/api/project/{project_id}", "Fetch final project")

    warnings = []
    export_stats = {"applied_chunks": 0, "skipped_chunks": []}
    if not args.skip_export:
        export_stats = export_bilingual(session, args.api_base, project_id, input_path, output_path)
        relocate_stats = relocate_reference_links_from_notes(output_path)
        warnings = audit_workbook(output_path, args.manual_review_limit)
    else:
        relocate_stats = {"relocated_rows": 0, "normalized_rows": 0}

    report = {
        "project_name": project_name,
        "input": str(input_path),
        "output": None if args.skip_export else str(output_path),
        "repo_root": str(repo_root),
        "api_base": args.api_base,
        "db_tools_target": proofread_result.get("target"),
        "google_config": {
            "id": google_cfg.get("id"),
            "name": google_cfg.get("name"),
            "provider": google_cfg.get("provider"),
            "model_name": google_cfg.get("model_name"),
        },
        "translation_dicts": list(zip(args.translation_dicts, translation_ids)),
        "replacement_dicts": list(zip(replacement_dict_names, replacement_ids)),
        "sheet_headers": sheet_headers,
        "project": {
            "id": project_id,
            "source_headers": source_headers,
            "source_cols": source_cols,
            "notes_sheets": note_sheets,
            "blanked_upload_cells": blanked_upload_cells,
            "translate_result": {
                "translated": translate_result.get("translated"),
                "errors": translate_result.get("errors", []),
            },
            "replace_result": replace_result,
            "chunk_count": project.get("chunk_count"),
            "verify_clean": proofread_result["verify_clean"],
            "verify_stdout": proofread_result["verify_stdout"],
            "export_stats": export_stats,
        },
        "backup": str(backup_path),
        "manual_review_warnings": warnings,
        "reference_rows_relocated": relocate_stats["relocated_rows"],
        "reference_rows_normalized": relocate_stats["normalized_rows"],
    }

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2))

    print(
        json.dumps(
            {
                "project_name": project_name,
                "project_id": project_id,
                "output": None if args.skip_export else str(output_path),
                "report": str(report_path),
                "warnings": len(warnings),
                "reference_rows_relocated": relocate_stats["relocated_rows"],
                "reference_rows_normalized": relocate_stats["normalized_rows"],
                "verify_clean": proofread_result["verify_clean"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
