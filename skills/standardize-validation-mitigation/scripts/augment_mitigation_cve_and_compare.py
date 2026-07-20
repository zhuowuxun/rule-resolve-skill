#!/usr/bin/env python3
"""Append CVE description/reference to mit_1 and compare with mit_2."""

from __future__ import annotations

import argparse
import json
import re
import time
import urllib.parse
import urllib.request
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from collections import Counter
from typing import Dict, List, Optional, Tuple


NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "p": "http://schemas.openxmlformats.org/package/2006/relationships",
}
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
SHEET_MAIN_NS = NS["a"]


def qname(local: str) -> str:
    return f"{{{SHEET_MAIN_NS}}}{local}"


def col_letters_to_index(letters: str) -> int:
    result = 0
    for char in letters:
        result = result * 26 + (ord(char) - 64)
    return result


def col_index_to_letters(index: int) -> str:
    letters: List[str] = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def split_ref(ref: str) -> Tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", ref)
    if not match:
        raise ValueError(f"Unsupported cell ref: {ref}")
    return col_letters_to_index(match.group(1)), int(match.group(2))


class XlsxArchive:
    def __init__(self, path: Path) -> None:
        self.path = path
        with zipfile.ZipFile(path) as archive:
            self.files: Dict[str, bytes] = {name: archive.read(name) for name in archive.namelist()}

    def write(self, output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, data in self.files.items():
                archive.writestr(name, data)


def load_shared_strings(files: Dict[str, bytes]) -> List[str]:
    if "xl/sharedStrings.xml" not in files:
        return []
    root = ET.fromstring(files["xl/sharedStrings.xml"])
    values: List[str] = []
    for item in root.findall("a:si", NS):
        values.append("".join(text.text or "" for text in item.iterfind(".//a:t", NS)))
    return values


def workbook_sheet_targets(files: Dict[str, bytes]) -> List[Tuple[str, str]]:
    workbook_root = ET.fromstring(files["xl/workbook.xml"])
    rel_root = ET.fromstring(files["xl/_rels/workbook.xml.rels"])
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rel_root.findall("p:Relationship", NS)}
    targets: List[Tuple[str, str]] = []
    for sheet in workbook_root.find("a:sheets", NS):
        name = sheet.attrib["name"]
        rel_id = sheet.attrib[f"{{{REL_NS}}}id"]
        target = rel_map[rel_id]
        if target.startswith("/"):
            target = target.lstrip("/")
        elif not target.startswith("xl/"):
            target = "xl/" + target
        targets.append((name, target))
    return targets


def read_cell_value(cell: Optional[ET.Element], shared_strings: List[str]) -> str:
    if cell is None:
        return ""
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.iterfind(".//a:t", NS))
    value_node = cell.find("a:v", NS)
    if value_node is None:
        return ""
    raw_value = value_node.text or ""
    if cell_type == "s" and raw_value.isdigit():
        idx = int(raw_value)
        if idx < len(shared_strings):
            return shared_strings[idx]
    return raw_value


def row_cells_by_col(row: ET.Element) -> Dict[int, ET.Element]:
    mapping: Dict[int, ET.Element] = {}
    for cell in row.findall("a:c", NS):
        ref = cell.attrib.get("r", "")
        if not ref:
            continue
        col_idx, _ = split_ref(ref)
        mapping[col_idx] = cell
    return mapping


def ensure_cell(row: ET.Element, col_idx: int, row_idx: int) -> ET.Element:
    existing = row_cells_by_col(row)
    if col_idx in existing:
        return existing[col_idx]

    new_cell = ET.Element(qname("c"))
    new_cell.attrib["r"] = f"{col_index_to_letters(col_idx)}{row_idx}"

    inserted = False
    cells = row.findall("a:c", NS)
    for idx, cell in enumerate(cells):
        current_col, _ = split_ref(cell.attrib["r"])
        if current_col > col_idx:
            row.insert(idx, new_cell)
            inserted = True
            break
    if not inserted:
        row.append(new_cell)
    return new_cell


def set_inline_string(cell: ET.Element, value: str) -> None:
    for child in list(cell):
        cell.remove(child)
    cell.attrib["t"] = "inlineStr"
    is_node = ET.SubElement(cell, qname("is"))
    t_node = ET.SubElement(is_node, qname("t"))
    if value.startswith(" ") or value.endswith(" ") or "\n" in value:
        t_node.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"
    t_node.text = value


def load_rows(path: Path) -> Dict[str, Dict[str, List[str]]]:
    archive = XlsxArchive(path)
    shared = load_shared_strings(archive.files)
    result: Dict[str, Dict[str, List[str]]] = {}
    for sheet_name, target in workbook_sheet_targets(archive.files):
        root = ET.fromstring(archive.files[target])
        for row in root.findall(".//a:sheetData/a:row", NS):
            values = [read_cell_value(cell, shared) for cell in row.findall("a:c", NS)]
            if not values:
                continue
            result[values[0]] = {"sheet": sheet_name, "values": values}
    return result


def extract_manual_cn_appendix(text: str) -> Tuple[str, List[str]]:
    normalized = text.replace("\r\n", "\n")
    if "\n\n" not in normalized:
        return "", []
    appendix = normalized.split("\n\n", 1)[1].strip()
    if not appendix:
        return "", []
    if "\n\n请参考：" in appendix:
        desc, refs_block = appendix.split("\n\n请参考：", 1)
        refs = [line.strip() for line in refs_block.splitlines() if line.strip()]
        return desc.strip(), refs
    return appendix.strip(), []


def split_english_sentences(text: str) -> List[str]:
    text = text.strip()
    if not text:
        return []
    parts = re.split(r"(?<=[.!?])\s+", text)
    return [part.strip() for part in parts if part.strip()]


def translate_chunk(text: str) -> str:
    url = (
        "https://translate.googleapis.com/translate_a/single"
        "?client=gtx&sl=en&tl=zh-CN&dt=t&q=" + urllib.parse.quote(text)
    )
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 Codex Validation Skill"})
    last_error: Exception | None = None
    data = None
    for attempt in range(3):
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                data = json.loads(response.read().decode("utf-8", errors="ignore"))
            break
        except Exception as exc:
            last_error = exc
            if attempt == 2:
                raise RuntimeError(f"Failed to translate chunk: {exc}") from exc
            time.sleep(1.5 * (attempt + 1))
    assert data is not None
    parts = []
    for item in data[0]:
        if item and item[0]:
            parts.append(item[0])
    return "".join(parts).strip()


def protect_product(text: str, product: str) -> Tuple[str, str]:
    if not product or not re.search(r"[A-Za-z0-9]", product):
        return text, product
    marker = "__PRODUCT_TOKEN__"
    protected = re.sub(re.escape(product), marker, text, flags=re.IGNORECASE)
    return protected, marker


def translate_version_range_sentence(sentence: str) -> Optional[str]:
    match = re.match(
        r"^A\s+(?:flaw|vulnerability|weakness|bug)\s+has\s+been\s+found\s+in\s+(.+?)\s+up\s+to\s+([0-9][A-Za-z0-9._-]*)\.$",
        sentence.strip(),
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    affected_product = match.group(1).strip()
    version = match.group(2).strip()
    return f"{affected_product} {version} 及之前版本存在漏洞。"


def translate_en_to_zh(text: str, product: str) -> str:
    sentences = split_english_sentences(text)
    translated_parts: List[str] = []
    for sentence in sentences:
        version_range_translation = translate_version_range_sentence(sentence)
        if version_range_translation:
            translated_parts.append(version_range_translation)
            continue
        protected, marker = protect_product(sentence, product)
        translated = translate_chunk(protected)
        if marker in translated:
            translated = translated.replace(marker, product)
        translated = re.sub(r"__\s*product\s*_?\s*token\s*__", product, translated, flags=re.IGNORECASE)
        translated_parts.append(translated)
    merged = "".join(translated_parts).strip()
    if product and re.search(r"[A-Za-z0-9]", product) and product not in merged and product.lower() in text.lower():
        # Last-resort anchor: restore the expected product token at the start.
        merged = re.sub(r"^[A-Za-z0-9 ._/-]+", product, merged, count=1)
        if product not in merged:
            merged = f"{product}：{merged}"
    return merged


def build_en_appendix(description: str, references: List[str]) -> str:
    description = description.strip()
    parts = [description] if description else []
    if references:
        parts.append("Please refer to:\n" + "\n".join(references[:5]))
    return "\n\n".join(parts).strip()


def build_cn_appendix(translated_desc: str, references: List[str]) -> str:
    desc = translated_desc.strip()
    parts = [desc] if desc else []
    cn_references = [ref for ref in references if "nist.gov" not in ref.lower()]
    if not cn_references:
        for ref in references:
            match = re.search(r"CVE-\d{4}-\d+", ref, flags=re.IGNORECASE)
            if match:
                cn_references.append(f"https://www.tenable.com/cve/{match.group(0).upper()}")
                break
    if cn_references:
        parts.append("请参考：\n" + "\n".join(cn_references[:5]))
    return "\n\n".join(parts).strip()


def append_block(base: str, appendix: str) -> str:
    base = base.strip()
    appendix = appendix.strip()
    if not appendix:
        return base
    if not base:
        return appendix
    return base + "\n\n" + appendix


def mitigation_tokens(tag: str) -> set[str]:
    return set(re.findall(r"M\d{4}", tag or ""))


def lookup_dictionary(dictionary: Dict[str, Tuple[str, str]], tag_cn: str) -> Tuple[str, str, str]:
    if tag_cn in dictionary:
        cn, en = dictionary[tag_cn]
        return cn, en, "字典精确回填"
    target_tokens = mitigation_tokens(tag_cn)
    if not target_tokens:
        return "", "", "字典未命中待补"

    best_key = ""
    best_score: Tuple[int, int, int, str] | None = None
    for key in dictionary:
        candidate_tokens = mitigation_tokens(key)
        if not candidate_tokens:
            continue
        overlap = len(target_tokens & candidate_tokens)
        if not overlap:
            continue
        symmetric_diff = len(target_tokens ^ candidate_tokens)
        score = (symmetric_diff, -overlap, len(candidate_tokens), key)
        if best_score is None or score < best_score:
            best_score = score
            best_key = key

    if best_key:
        cn, en = dictionary[best_key]
        return cn, en, "字典近似回填"
    return "", "", "字典未命中待补"


def add_update(
    updates: Dict[Tuple[str, int, int], Dict[str, object]],
    sheet: str,
    row: int,
    col: int,
    field: str,
    reason: str,
    old_value: str,
    new_value: str,
) -> None:
    key = (sheet, row, col)
    entry = updates.setdefault(
        key,
        {
            "sheet": sheet,
            "row": row,
            "column": col_index_to_letters(col),
            "field": field,
            "reasons": [],
            "old_value": old_value,
            "new_value": new_value,
        },
    )
    reasons = entry["reasons"]
    assert isinstance(reasons, list)
    if reason not in reasons:
        reasons.append(reason)
    entry["new_value"] = new_value


def apply_yellow_fills_and_audit_sheet(output_path: Path, updates: Dict[Tuple[str, int, int], Dict[str, object]]) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise RuntimeError("openpyxl is required to normalize mitigation workbooks for Excel") from exc

    wb = load_workbook(output_path)
    if not updates:
        # Re-save through openpyxl so Excel opens workbooks generated by XML-level edits reliably.
        wb.save(output_path)
        return

    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for (sheet_name, row_idx, col_idx), _entry in updates.items():
        if sheet_name not in wb.sheetnames:
            continue
        wb[sheet_name].cell(row=row_idx, column=col_idx).fill = yellow_fill

    audit_sheet_name = "自动回填记录"
    if audit_sheet_name in wb.sheetnames:
        del wb[audit_sheet_name]
    ws = wb.create_sheet(audit_sheet_name)
    headers = ["sheet", "row", "column", "field", "reason", "old_value", "new_value"]
    ws.append(headers)
    for entry in sorted(updates.values(), key=lambda item: (str(item["sheet"]), int(item["row"]), str(item["column"]))):
        ws.append(
            [
                entry["sheet"],
                entry["row"],
                entry["column"],
                entry["field"],
                "；".join(entry["reasons"]),
                entry["old_value"],
                entry["new_value"],
            ]
        )
    ws.freeze_panes = "A2"
    widths = {"A": 18, "B": 8, "C": 10, "D": 12, "E": 30, "F": 60, "G": 60}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    wb.save(output_path)


def audit_output_against_source(
    source_path: Path,
    output_path: Path,
    allowed_updates: Dict[Tuple[str, int, int], Dict[str, object]],
) -> Dict[str, object]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise RuntimeError("openpyxl is required to audit mitigation workbooks against source") from exc

    source_wb = load_workbook(source_path, data_only=False)
    output_wb = load_workbook(output_path, data_only=False)
    audit_sheet_name = "自动回填记录"
    allowed_cells = set(allowed_updates)
    unexpected: List[Dict[str, object]] = []
    allowed_changed = 0

    for sheet_name in source_wb.sheetnames:
        if sheet_name not in output_wb.sheetnames:
            unexpected.append({"sheet": sheet_name, "issue": "missing_sheet_in_output"})
            continue
        src_ws = source_wb[sheet_name]
        out_ws = output_wb[sheet_name]
        max_row = max(src_ws.max_row, out_ws.max_row)
        max_col = max(src_ws.max_column, out_ws.max_column)
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                src_val = src_ws.cell(row=row_idx, column=col_idx).value
                out_val = out_ws.cell(row=row_idx, column=col_idx).value
                if src_val == out_val:
                    continue
                key = (sheet_name, row_idx, col_idx)
                if key in allowed_cells:
                    allowed_changed += 1
                    continue
                unexpected.append(
                    {
                        "sheet": sheet_name,
                        "row": row_idx,
                        "column": col_index_to_letters(col_idx),
                        "source_value": "" if src_val is None else str(src_val)[:300],
                        "output_value": "" if out_val is None else str(out_val)[:300],
                    }
                )

    extra_sheets = [name for name in output_wb.sheetnames if name not in source_wb.sheetnames and name != audit_sheet_name]
    for sheet_name in extra_sheets:
        unexpected.append({"sheet": sheet_name, "issue": "unexpected_extra_sheet"})

    return {
        "source": str(source_path),
        "output": str(output_path),
        "allowed_changed_cells": allowed_changed,
        "allowed_recorded_cells": len(allowed_cells),
        "unexpected_diff_count": len(unexpected),
        "unexpected_diffs": unexpected[:100],
    }


def keep_remediation_only(text: str) -> str:
    text = text.strip()
    if not text:
        return ""
    return text.split("\n\n", 1)[0].strip()


def is_missing_note(text: str) -> bool:
    return not text or text.strip() == "#N/A"


def load_history_remediations(path: Optional[Path]) -> Tuple[Dict[str, Dict[str, str]], Dict[str, Dict[str, str]]]:
    if not path:
        return {}, {}
    history_rows = load_rows(path)
    by_uuid: Dict[str, Dict[str, str]] = {}
    by_name: Dict[str, Dict[str, str]] = {}
    for _row_key, info in history_rows.items():
        values = info["values"]
        if len(values) < 4:
            continue
        uuid = values[0].strip()
        name = values[1].strip()
        if uuid == "uuid" or name == "name":
            continue
        cn = keep_remediation_only(values[2])
        en = keep_remediation_only(values[3])
        if not cn and not en:
            continue
        record = {"uuid": uuid, "name": name, "cn": cn, "en": en}
        if uuid:
            by_uuid[uuid] = record
        if name and name not in by_name:
            by_name[name] = record
    return by_uuid, by_name


def append_reason(base_reason: str, extra_reason: str) -> str:
    if not base_reason:
        return extra_reason
    if extra_reason in base_reason.split("；"):
        return base_reason
    return f"{base_reason}；{extra_reason}"


def select_dictionary_by_cn_phrase(
    dictionary: Dict[str, Tuple[str, str]],
    tag_cn: str,
    phrase: str,
    preferred_key: str = "",
) -> Tuple[str, str]:
    if preferred_key and preferred_key in dictionary and phrase in dictionary[preferred_key][0]:
        return dictionary[preferred_key]

    target_tokens = mitigation_tokens(tag_cn)
    best_key = ""
    best_score: Tuple[int, int, int, str] | None = None
    for key, (cn, _en) in dictionary.items():
        if phrase not in cn:
            continue
        candidate_tokens = mitigation_tokens(key)
        overlap = len(target_tokens & candidate_tokens)
        symmetric_diff = len(target_tokens ^ candidate_tokens) if target_tokens else 0
        score = (symmetric_diff, -overlap, len(candidate_tokens), key)
        if best_score is None or score < best_score:
            best_score = score
            best_key = key
    return dictionary.get(best_key, ("", ""))


def apply_fill_context(
    dictionary: Dict[str, Tuple[str, str]],
    tag_cn: str,
    name: str,
    rule_type: str,
    os_scope: str,
    base_cn: str,
    base_en: str,
    fill_reason: str,
) -> Tuple[str, str, str]:
    _ = (name, os_scope)
    if rule_type == "命令与控制" and "C&C服务器" not in base_cn:
        c2_cn, c2_en = select_dictionary_by_cn_phrase(dictionary, tag_cn, "C&C服务器", "M1021M1031")
        if c2_cn and c2_en:
            return c2_cn, c2_en, append_reason(fill_reason, "规则类型二次判断：命令与控制改用C&C模板")

    if rule_type == "恶意文件传输" and "基于传输的恶意文件的Hash" not in base_cn:
        transfer_cn, transfer_en = dictionary.get("M1031", ("", ""))
        if transfer_cn and transfer_en:
            return transfer_cn, transfer_en, append_reason(fill_reason, "规则类型二次判断：恶意文件传输改用文件传输模板")

    return base_cn, base_en, fill_reason


def infer_os_scope(name: str, os_scope: str) -> str:
    scope = (os_scope or "").strip().lower()
    combined = name.lower()
    if scope in {"windows", "linux", "macos"}:
        return {"windows": "Windows", "linux": "Linux", "macos": "macOS"}[scope]
    if any(token in combined for token in ["(windows)", "windows", ".exe", ".dll", "powershell", "cmdkey", "rundll32", "注册表"]):
        return "Windows"
    if any(token in combined for token in ["(linux)", "linux", "/etc/", "ld_preload", "syscall", "bash", "chmod", "zlib"]):
        return "Linux"
    if any(token in combined for token in ["(macos)", "macos", ".macho", " mach-o"]):
        return "macOS"
    return "ANY"


def cleanup_os_mismatch(cn: str, en: str, inferred_os: str, rule_type: str) -> Tuple[str, str]:
    if inferred_os == "Linux":
        cn = cn.replace("、修改注册表等权限控制", "等权限控制")
        cn = cn.replace("、修改注册表", "")
        cn = cn.replace("修改注册表等权限控制", "权限控制")
        cn = cn.replace("PowerShell", "Shell")
        en = en.replace(", modifying the registry, etc.", ", etc.")
        en = en.replace(", modifying the registry", "")
        en = en.replace("modifying the registry, etc.", "permission control, etc.")
        en = en.replace("PowerShell", "Shell")

    if inferred_os == "Windows" and rule_type == "主机命令行":
        cn = cn.replace("Linux ", "").replace("Linux", "")
        en = en.replace("Linux ", "").replace("Linux", "")

    cn = cn.replace("、等", "等").replace("；。", "。").replace("。。", "。")
    en = en.replace(" ,", ",").replace("..", ".")
    return cn.strip(), en.strip()


def normalize_host_command_product_scope(cn: str, en: str, inferred_os: str, rule_type: str) -> Tuple[str, str]:
    if rule_type != "主机命令行":
        return cn, en

    if inferred_os == "Windows":
        cn = re.sub(r"^如果(?:终端/主机|终端或主机|主机|终端)安全产品对此攻击漏检", "如果终端或主机安全产品对此攻击漏检", cn)
        en = re.sub(
            r"^If the (?:endpoint/server|endpoint or server|server|endpoint) security products? miss(?:es)? this attack",
            "If the endpoint or server security products miss this attack",
            en,
            flags=re.IGNORECASE,
        )
    elif inferred_os == "Linux":
        cn = re.sub(r"^如果(?:终端/主机|终端或主机|主机|终端)安全产品对此攻击漏检", "如果主机安全产品对此攻击漏检", cn)
        en = re.sub(
            r"^If the (?:endpoint/server|endpoint or server|server|endpoint) security products? miss(?:es)? this attack",
            "If the server security products miss this attack",
            en,
            flags=re.IGNORECASE,
        )
    elif inferred_os == "macOS":
        cn = re.sub(r"^如果(?:终端/主机|终端或主机|主机|终端)安全产品对此攻击漏检", "如果终端安全产品对此攻击漏检", cn)
        en = re.sub(
            r"^If the (?:endpoint/server|endpoint or server|server|endpoint) security products? miss(?:es)? this attack",
            "If the endpoint security products miss this attack",
            en,
            flags=re.IGNORECASE,
        )
    return cn, en


def normalize_reference_priority(references: List[str]) -> List[str]:
    def score(url: str) -> Tuple[int, str]:
        lowered = url.lower()
        priority = 50
        if "tenable.com/cve/" in lowered:
            priority = 1
        elif "security/advisories" in lowered or "advisory" in lowered:
            priority = 1
        elif "support." in lowered or "docs." in lowered or "release-notes" in lowered:
            priority = 2
        elif "cisa.gov" in lowered or "apache.org" in lowered or "wordpress.org" in lowered:
            priority = 3
        elif "github.com" in lowered and "/security/" in lowered:
            priority = 4
        elif "github.com" in lowered:
            priority = 8
        elif "vuldb.com/?id." in lowered:
            priority = 20
        elif "vuldb.com/?ctiid." in lowered or "submit." in lowered or "x.com/" in lowered:
            priority = 30
        elif "metasploit" in lowered:
            priority = 40
        return (priority, lowered)

    deduped = []
    seen = set()
    for ref in references:
        if ref not in seen:
            seen.add(ref)
            deduped.append(ref)
    return sorted(deduped, key=score)[:5]


def apply_network_only_product_cleanup(cn: str, en: str) -> Tuple[str, str]:
    cn = cn.replace("如果网络或主机安全产品对此攻击漏检", "如果网络安全产品对此攻击漏检")
    en = en.replace(
        "If the network or server security products miss this attack",
        "If the network security products miss this attack",
    )
    cn = cn.replace("；通过优化WAF产品的检测规则实现防御或评估RSAP产品在贵司的适用性", "")
    cn = cn.replace("；或通过优化WAF产品的检测规则实现防御", "")
    cn = cn.replace("；通过优化WAF产品的检测规则实现防御", "")
    cn = cn.replace("；做好应用程序隔离。", "。")
    cn = cn.replace("做好应用程序隔离。", "")
    en = en.replace(
        "; implement defense through the detection rule of the optimize WAF product or evaluate the applicability of RSAP products in your company",
        "",
    )
    en = en.replace(
        "; or implement defense through the detection rule of the optimize WAF product",
        "",
    )
    en = en.replace(
        "; implement defense through the detection rule of the optimize WAF product",
        "",
    )
    en = en.replace("; and do a good job of application isolation.", ".")
    en = en.replace("and do a good job of application isolation.", "")
    cn = cn.replace("。。", "。").replace("；。", "。")
    en = en.replace("..", ".").replace(";.", ".")
    return cn, en


def sandbox_should_omit_destructive_head(name: str) -> bool:
    """Sandbox cleanup/rollback-difficult rules are not necessarily destructive."""
    normalized = re.sub(r"\s+", " ", name).lower()
    non_destructive_tokens = (
        "持久化",
        "c&c",
        "c2",
        "命令与控制",
        "信标",
        "beacon",
        "连接至",
        "联络",
        "通信",
        "采集日志",
        "日志采集",
        "计划任务",
        "任务计划",
        "scheduled task",
        "注册表",
        "registry",
        "run key",
        "service",
        "自启动",
        "启动项",
        "登录项",
        "驻留",
    )
    return any(token in normalized for token in non_destructive_tokens) and not sandbox_has_explicit_destructive_effect(
        name
    )


def sandbox_has_explicit_destructive_effect(name: str) -> bool:
    normalized = re.sub(r"\s+", " ", name).lower()
    destructive_tokens = (
        "破坏",
        "擦除",
        "销毁",
        "加密",
        "勒索",
        "wiper",
        "终止安全防护",
        "停用数据服务",
        "删除卷影副本",
        "shadow copy",
        "disable security",
        "disable service",
        "terminate security",
        "delete shadow",
        "wipe",
        "encrypt",
        "ransom",
        "destroy",
    )
    return any(token in normalized for token in destructive_tokens)


def remove_destructive_head(cn: str, en: str) -> Tuple[str, str]:
    cn = cn.replace("此攻击手法是具有破坏性的，如果", "如果")
    cn = cn.replace("此攻击手法是具有危险性的，如果", "如果")
    cn = cn.replace("此攻击手法是有危险性的，如果", "如果")
    en = en.replace("This attack method is destructive. If", "If")
    en = en.replace("This attack method is dangerous. If", "If")
    return cn, en


def add_destructive_head(cn: str, en: str) -> Tuple[str, str]:
    if cn.startswith("如果"):
        cn = "此攻击手法是具有破坏性的，" + cn
    if en.startswith("If "):
        en = "This attack method is destructive. " + en
    return cn, en


def is_cloud_system_or_software(product: str, name: str, rule_type: str) -> bool:
    if rule_type not in {"Web应用程序漏洞", "应用程序漏洞"}:
        return False
    product_normalized = product.replace("\xa0", " ").strip()
    name_normalized = name.replace("\xa0", " ").strip()
    lowered = f"{product_normalized} {name_normalized}".lower()
    if "云" in product_normalized:
        return True
    explicit_cloud_products = [
        "yudao-cloud",
        "u8 cloud",
        "u8cloud",
        "nc cloud",
        "nccloud",
        "analyticscloud",
    ]
    return any(item in lowered for item in explicit_cloud_products)


def vendor_contact_candidates() -> Dict[str, Tuple[str, str]]:
    return {
        "泛微": ("泛微", "Weaver"),
        "e-cology": ("泛微", "Weaver"),
        "e-colology": ("泛微", "Weaver"),
        "e-bridge": ("泛微", "Weaver"),
        "用友": ("用友", "UFIDA"),
        "UFIDA": ("用友", "UFIDA"),
        "U8 Cloud": ("用友", "UFIDA"),
        "U8Cloud": ("用友", "UFIDA"),
        "NC Cloud": ("用友", "UFIDA"),
        "NcCloud": ("用友", "UFIDA"),
        "NCCloud": ("用友", "UFIDA"),
        "孚盟云": ("孚盟云", "Fumasoft"),
        "FUMAsoft": ("孚盟云", "Fumasoft"),
        "金蝶云": ("金蝶云", "Kingdee"),
        "金蝶": ("金蝶", "Kingdee"),
        "Kingdee": ("金蝶", "Kingdee"),
        "百易云": ("百易云", "ehub"),
        "智联云": ("智联云", "Dahui Zhilian"),
        "致远互联": ("致远互联", "Seeyon"),
        "Seeyon": ("致远互联", "Seeyon"),
        "红海云": ("红海云", "Honghai Cloud"),
        "全程云": ("全程云", "eqccd.com"),
        "方天云": ("方天云", "Fangtian Cloud"),
        "华望云": ("华望云", "Huawang Cloud"),
        "华望": ("华望", "Huawang"),
        "ServiceNow": ("ServiceNow", "ServiceNow"),
        "Service Now": ("ServiceNow", "ServiceNow"),
        "Nextcloud": ("Nextcloud", "Nextcloud"),
        "ownCloud": ("ownCloud", "ownCloud"),
        "Cloudflare": ("Cloudflare", "Cloudflare"),
        "腾讯云": ("腾讯云", "Tencent Cloud"),
        "阿里云": ("阿里云", "Alibaba Cloud"),
        "CloudPanel": ("CloudPanel", "CloudPanel"),
        "Akamai": ("Akamai", "Akamai"),
        "Adobe": ("Adobe", "Adobe"),
        "微软": ("微软", "Microsoft"),
        "Microsoft": ("微软", "Microsoft"),
        "海康威视": ("海康威视", "Hikvision"),
        "Hikvision": ("海康威视", "Hikvision"),
        "D-Link": ("D-Link", "D-Link"),
        "F5 BIG-IP": ("F5", "F5"),
        "Cisco": ("Cisco", "Cisco"),
        "Pulse Secure": ("Pulse Secure", "Pulse Secure"),
        "Ivanti": ("Ivanti", "Ivanti"),
        "Fortinet": ("Fortinet", "Fortinet"),
        "Buffalo 路由器": ("Buffalo", "Buffalo"),
        "飞鱼星": ("飞鱼星", "Volans"),
        "腾达": ("腾达", "Tenda"),
        "网神SecGate": ("网神SecGate", "SecGate"),
        "WIFISKY": ("WIFISKY", "WIFISKY"),
        "瑞斯康达": ("瑞斯康达", "Raisecom"),
        "HPE OneView": ("HPE", "HPE"),
        "BMC FootPrints": ("BMC", "BMC"),
        "深信服": ("深信服", "Sangfor"),
        "安恒": ("安恒", "DBAPPSecurity"),
    }


def build_vendor_contact_notes(product: str, rule_type: str, name: str) -> Tuple[str, str]:
    if not any(token in rule_type for token in ("Web应用程序漏洞", "应用程序漏洞", "AI应用程序漏洞", "WAF绕过")):
        return "", ""
    haystack = f"{product} {name}".replace("\xa0", " ")
    vendor_map = {
        "孚盟云": ("孚盟云", "Fumasoft"),
        "金蝶云星空": ("金蝶云", "Kingdee"),
        "用友 U8 Cloud": ("用友", "UFIDA"),
        "用友 U8Cloud": ("用友", "UFIDA"),
        "用友U8 Cloud": ("用友", "UFIDA"),
        "用友U8Cloud": ("用友", "UFIDA"),
        "UFIDA U8 cloud": ("用友", "UFIDA"),
        "用友 NcCloud": ("用友", "UFIDA"),
        "用友NcCloud": ("用友", "UFIDA"),
        "用友 NC Cloud": ("用友", "UFIDA"),
        "用友NC Cloud": ("用友", "UFIDA"),
    }
    vendor_map.update(vendor_contact_candidates())
    for key, vendor in vendor_map.items():
        if key.replace("\xa0", " ") in haystack:
            cn_vendor, en_vendor = vendor
            cn = (
                f"如果网络安全产品对此攻击漏检，塞讯验证短期建议：联络{cn_vendor}告知漏洞。"
                "中长期建议：增加特权帐号管理提升身份安全；增加多因素认证；"
                "加强互联网访问的管控与下载文件检测，并加强内网网络流量的威胁文件传输检测；"
                "在内部各个安全区域之间部署NIPS加强检测与防御。"
            )
            en = (
                "If the network security products miss this attack, digiDations' short-term recommendation is: "
                f"contact {en_vendor} to report the vulnerability. Mid- to long-term recommendations: "
                "add privileged account management to improve identity security; add multi-factor authentication; "
                "strengthen Internet access control and download file detection, and strengthen threat file transfer "
                "detection of Malicious Network traffic in the internal; deploy NIPS between internal Security Zones "
                "to strengthen detection and defense."
            )
            return cn, en
    return "", ""


def has_cve_identifier(name: str, cve: str) -> bool:
    if cve and cve.strip() != "#N/A":
        return True
    return bool(re.search(r"CVE-\d{4}-\d+", name or "", flags=re.IGNORECASE))


def transform_base_notes(
    base_cn: str,
    base_en: str,
    rule_type: str,
    product: str,
    name: str,
    os_scope: str,
    cve: str = "",
) -> Tuple[str, str]:
    network_only_products = {
        "深信服运维安全管理系统",
        "深信服下一代防火墙",
        "安恒明御WEB应用防火墙",
        "BMC FootPrints ITSM",
        "F5 BIG-IP",
        "HPE OneView",
        "Cisco ASA 和 Firepower",
        "Cisco vManage",
        "Cisco Smart Licensing Utility",
        "Cisco IOS 和 IOS XE 集群管理协议（CMP）",
        "Pulse Secure SSL VPN",
        "Ivanti Connect Secure VPN",
        "Ivanti Endpoint Manager Mobile",
        "Fortinet FortiMail",
        "D-Link",
        "D-Link Central WiFiManager 软件控制器",
        "Buffalo 路由器",
        "飞鱼星路由器",
        "腾达 FH1201 路由器",
        "网神SecGate 3600防火墙",
        "WIFISKY-7层流控路由器",
        "瑞斯康达 多业务智能网关",
    }

    cn = base_cn
    en = base_en

    if "受保护的沙盘" in rule_type and not sandbox_should_omit_destructive_head(name):
        cn = cn.replace("此攻击手法是具有危险性的，如果", "此攻击手法是具有破坏性的，如果")
        cn = cn.replace("此攻击手法是有危险性的，如果", "此攻击手法是具有破坏性的，如果")
        en = en.replace("This attack method is dangerous. If", "This attack method is destructive. If")
        if sandbox_has_explicit_destructive_effect(name):
            cn, en = add_destructive_head(cn, en)
    else:
        cn, en = remove_destructive_head(cn, en)

    network_only_prefixes = (
        "D-Link",
        "F5 BIG-IP",
        "Cisco ",
        "Pulse Secure",
        "Ivanti Connect Secure",
        "Ivanti Endpoint Manager Mobile",
        "Fortinet ",
        "Buffalo 路由器",
        "飞鱼星路由器",
        "腾达 ",
        "网神SecGate",
        "WIFISKY",
        "瑞斯康达",
        "深信服下一代防火墙",
        "安恒明御WEB应用防火墙",
    )
    if product in network_only_products or product.startswith(network_only_prefixes):
        cn, en = apply_network_only_product_cleanup(cn, en)

    contact_cn, contact_en = build_vendor_contact_notes(product, rule_type, name)
    if contact_cn and contact_en:
        cn, en = contact_cn, contact_en

    if is_cloud_system_or_software(product, name, rule_type):
        if not has_cve_identifier(name, cve) and not contact_cn:
            contact_cn, contact_en = build_vendor_contact_notes(product, rule_type, name)
            if contact_cn and contact_en:
                cn, en = contact_cn, contact_en
        cn, en = apply_network_only_product_cleanup(cn, en)

    inferred_os = infer_os_scope(name, os_scope)
    cn, en = cleanup_os_mismatch(cn, en, inferred_os, rule_type)
    cn, en = normalize_host_command_product_scope(cn, en, inferred_os, rule_type)
    en = en.replace("Digidations", "digiDations")

    return cn.strip(), en.strip()


def compare_notes(our_cn: str, our_en: str, manual_cn: str, manual_en: str) -> Dict[str, bool]:
    return {
        "cn_same": our_cn == manual_cn,
        "en_same": our_en == manual_en,
        "manual_network_only": "如果网络安全产品对此攻击漏检" in manual_cn and "如果网络或主机安全产品对此攻击漏检" not in manual_cn,
        "ours_network_or_host": "如果网络或主机安全产品对此攻击漏检" in our_cn,
        "manual_removed_waf_rsap": "通过优化WAF产品的检测规则实现防御或评估RSAP产品在贵司的适用性" not in manual_cn,
        "ours_has_waf_rsap": "通过优化WAF产品的检测规则实现防御或评估RSAP产品在贵司的适用性" in our_cn,
        "manual_removed_app_isolation": "做好应用程序隔离" not in manual_cn,
        "ours_has_app_isolation": "做好应用程序隔离" in our_cn,
        "refs_same": (
            ("请参考：" in our_cn) == ("请参考：" in manual_cn)
            and ("Please refer to:" in our_en) == ("Please refer to:" in manual_en)
        ),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", required=True, type=Path)
    parser.add_argument("--manual", required=True, type=Path)
    parser.add_argument("--cve-json", required=True, type=Path)
    parser.add_argument("--dictionary", required=True, type=Path)
    parser.add_argument("--history", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    cve_data = {item["cve"]: item for item in json.loads(args.cve_json.read_text(encoding="utf-8"))}
    dictionary_rows = load_rows(args.dictionary)
    dictionary = {}
    for _, info in dictionary_rows.items():
        values = info["values"]
        if len(values) >= 3 and values[0] != "mitigation":
            dictionary[values[0]] = (values[1], values[2] if len(values) > 2 else "")
    manual_rows = load_rows(args.manual)
    history_by_uuid, history_by_name = load_history_remediations(args.history)
    archive = XlsxArchive(args.base)
    shared = load_shared_strings(archive.files)

    summary = {
        "output": str(args.output),
        "cve_rows_updated": 0,
        "compare_rows": 0,
        "diff_rows": 0,
        "highlighted_cells": 0,
        "highlight_reasons": {},
        "auto_updates": [],
        "improvement_hints": {
            "network_or_host_vs_network_only": 0,
            "waf_rsap_phrase_should_be_removed": 0,
            "application_isolation_should_be_removed": 0,
        },
        "details": [],
    }
    auto_updates: Dict[Tuple[str, int, int], Dict[str, object]] = {}

    for sheet_name, target in workbook_sheet_targets(archive.files):
        root = ET.fromstring(archive.files[target])
        changed = False
        for row in root.findall(".//a:sheetData/a:row", NS):
            row_idx = int(row.attrib.get("r", "0"))
            if row_idx <= 1:
                continue
            cell_map = row_cells_by_col(row)
            uuid = read_cell_value(cell_map.get(1), shared).strip()
            tag_cn = read_cell_value(cell_map.get(2), shared).strip()
            cve = read_cell_value(cell_map.get(6), shared).strip()
            if not uuid:
                continue

            base_cn = read_cell_value(cell_map.get(7), shared).strip()
            base_en = read_cell_value(cell_map.get(8), shared).strip()
            original_cn = base_cn
            original_en = base_en
            had_dangerous_old_wording = "危险性" in base_cn or "dangerous" in base_en
            name = read_cell_value(cell_map.get(3), shared).strip()
            rule_type = read_cell_value(cell_map.get(4), shared).strip()
            os_scope = read_cell_value(cell_map.get(5), shared).strip()
            filled_cn_reason: str | None = None
            filled_en_reason: str | None = None
            if is_missing_note(base_cn) or is_missing_note(base_en):
                history = history_by_uuid.get(uuid) or history_by_name.get(name)
                if history:
                    if is_missing_note(base_cn) and history.get("cn"):
                        base_cn = history["cn"]
                        filled_cn_reason = "历史标准化语料精确回填"
                    if is_missing_note(base_en) and history.get("en"):
                        base_en = history["en"]
                        filled_en_reason = "历史标准化语料精确回填"

            if is_missing_note(base_cn) or is_missing_note(base_en):
                dict_cn, dict_en, fill_reason = lookup_dictionary(dictionary, tag_cn)
                if is_missing_note(base_cn):
                    base_cn = dict_cn
                    if dict_cn:
                        filled_cn_reason = fill_reason
                if is_missing_note(base_en):
                    base_en = dict_en
                    if dict_en:
                        filled_en_reason = fill_reason

            product = name.split(" - ", 1)[1].split("，")[0].strip() if " - " in name else name
            fill_reasons = [reason for reason in [filled_cn_reason, filled_en_reason] if reason]
            should_contextualize_dictionary_fill = any(reason.startswith("字典") for reason in fill_reasons)
            if should_contextualize_dictionary_fill and base_cn and base_en:
                fill_context_reason = "；".join(dict.fromkeys(fill_reasons))
                base_cn, base_en, contextual_reason = apply_fill_context(
                    dictionary,
                    tag_cn,
                    name,
                    rule_type,
                    os_scope,
                    base_cn,
                    base_en,
                    fill_context_reason,
                )
                if filled_cn_reason:
                    filled_cn_reason = contextual_reason
                if filled_en_reason:
                    filled_en_reason = contextual_reason
            pre_transform_cn = base_cn
            pre_transform_en = base_en
            transformed_cn, transformed_en = transform_base_notes(base_cn, base_en, rule_type, product, name, os_scope, cve)
            if transformed_cn != read_cell_value(cell_map.get(7), shared).strip() or transformed_en != read_cell_value(cell_map.get(8), shared).strip():
                set_inline_string(ensure_cell(row, 7, row_idx), transformed_cn)
                set_inline_string(ensure_cell(row, 8, row_idx), transformed_en)
                transform_reasons: List[str] = []
                if is_cloud_system_or_software(product, name, rule_type):
                    transform_reasons.append("云上系统/软件口径调整")
                if build_vendor_contact_notes(product, rule_type, name)[0] and "联络" in transformed_cn:
                    transform_reasons.append("厂商/设备联系厂商口径调整")
                if product in {
                    "深信服运维安全管理系统",
                    "深信服下一代防火墙",
                    "安恒明御WEB应用防火墙",
                    "BMC FootPrints ITSM",
                    "F5 BIG-IP",
                    "HPE OneView",
                    "Cisco ASA 和 Firepower",
                    "Cisco vManage",
                    "Cisco Smart Licensing Utility",
                    "Cisco IOS 和 IOS XE 集群管理协议（CMP）",
                    "Pulse Secure SSL VPN",
                    "Ivanti Connect Secure VPN",
                    "Ivanti Endpoint Manager Mobile",
                    "Fortinet FortiMail",
                    "D-Link",
                    "D-Link Central WiFiManager 软件控制器",
                    "Buffalo 路由器",
                    "飞鱼星路由器",
                    "腾达 FH1201 路由器",
                    "网神SecGate 3600防火墙",
                    "WIFISKY-7层流控路由器",
                    "瑞斯康达 多业务智能网关",
                }:
                    transform_reasons.append("硬件/网络设备口径调整")
                sandbox_omitted_destructive = (
                    "受保护的沙盘" in rule_type
                    and sandbox_should_omit_destructive_head(name)
                    and (
                        "破坏性" in pre_transform_cn
                        or "危险性" in pre_transform_cn
                        or "destructive" in pre_transform_en
                        or "dangerous" in pre_transform_en
                    )
                )
                if sandbox_omitted_destructive:
                    transform_reasons.append("沙盘非破坏性行为句式清理")
                elif had_dangerous_old_wording:
                    transform_reasons.append("危险性旧口径纠正")
                if "受保护的沙盘" not in rule_type and (
                    "破坏性" in pre_transform_cn or "destructive" in pre_transform_en
                ):
                    transform_reasons.append("非沙盘破坏性句式清理")
                if not transform_reasons:
                    transform_reasons.append("mitigation口径标准化")
                if transformed_cn != original_cn:
                    for reason in transform_reasons:
                        add_update(auto_updates, sheet_name, row_idx, 7, "cn_notes", reason, original_cn, transformed_cn)
                if transformed_en != original_en:
                    for reason in transform_reasons:
                        add_update(auto_updates, sheet_name, row_idx, 8, "en_notes", reason, original_en, transformed_en)
                changed = True
            base_cn, base_en = transformed_cn, transformed_en

            if not cve or cve == "#N/A" or cve not in cve_data:
                if filled_cn_reason:
                    add_update(auto_updates, sheet_name, row_idx, 7, "cn_notes", filled_cn_reason, original_cn, base_cn)
                if filled_en_reason:
                    add_update(auto_updates, sheet_name, row_idx, 8, "en_notes", filled_en_reason, original_en, base_en)
                continue

            manual = manual_rows.get(uuid, {}).get("values", [])
            manual_cn = manual[6].strip() if len(manual) > 6 else ""
            manual_en = manual[7].strip() if len(manual) > 7 else ""

            base_cn = keep_remediation_only(base_cn)
            base_en = keep_remediation_only(base_en)
            raw_refs = list(cve_data[cve].get("references") or [])
            source_url = str(cve_data[cve].get("source_url") or "").strip()
            if source_url and source_url not in raw_refs:
                raw_refs.append(source_url)
            refs = normalize_reference_priority(raw_refs)
            description_en = str(cve_data[cve].get("description_en") or "").strip()
            appendix_cn = build_cn_appendix(translate_en_to_zh(description_en, product) if description_en else "", refs)
            appendix_en = build_en_appendix(description_en, refs)
            new_cn = append_block(base_cn, appendix_cn)
            new_en = append_block(base_en, appendix_en)

            row_cve_changed = False
            if new_cn != original_cn:
                set_inline_string(ensure_cell(row, 7, row_idx), new_cn)
                add_update(auto_updates, sheet_name, row_idx, 7, "cn_notes", "CVE描述/reference补充", original_cn, new_cn)
                if filled_cn_reason:
                    add_update(auto_updates, sheet_name, row_idx, 7, "cn_notes", filled_cn_reason, original_cn, new_cn)
                row_cve_changed = True
            if new_en != original_en:
                set_inline_string(ensure_cell(row, 8, row_idx), new_en)
                add_update(auto_updates, sheet_name, row_idx, 8, "en_notes", "CVE描述/reference补充", original_en, new_en)
                if filled_en_reason:
                    add_update(auto_updates, sheet_name, row_idx, 8, "en_notes", filled_en_reason, original_en, new_en)
                row_cve_changed = True
            if row_cve_changed:
                changed = True
                summary["cve_rows_updated"] += 1

            cmp = compare_notes(new_cn, new_en, manual_cn, manual_en)
            summary["compare_rows"] += 1
            if not cmp["cn_same"] or not cmp["en_same"]:
                summary["diff_rows"] += 1
                if cmp["manual_network_only"] and cmp["ours_network_or_host"]:
                    summary["improvement_hints"]["network_or_host_vs_network_only"] += 1
                if cmp["manual_removed_waf_rsap"] and cmp["ours_has_waf_rsap"]:
                    summary["improvement_hints"]["waf_rsap_phrase_should_be_removed"] += 1
                if cmp["manual_removed_app_isolation"] and cmp["ours_has_app_isolation"]:
                    summary["improvement_hints"]["application_isolation_should_be_removed"] += 1

                summary["details"].append(
                    {
                        "uuid": uuid,
                        "sheet": sheet_name,
                        "row": row_idx,
                        "cve": cve,
                        "name": name,
                        "cn_same": cmp["cn_same"],
                        "en_same": cmp["en_same"],
                        "manual_network_only": cmp["manual_network_only"],
                        "ours_network_or_host": cmp["ours_network_or_host"],
                        "manual_removed_waf_rsap": cmp["manual_removed_waf_rsap"],
                        "ours_has_waf_rsap": cmp["ours_has_waf_rsap"],
                        "manual_removed_app_isolation": cmp["manual_removed_app_isolation"],
                        "ours_has_app_isolation": cmp["ours_has_app_isolation"],
                        "our_cn_head": new_cn[:220],
                        "manual_cn_head": manual_cn[:220],
                    }
                )

        if changed:
            archive.files[target] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    archive.write(args.output)
    apply_yellow_fills_and_audit_sheet(args.output, auto_updates)
    source_consistency = audit_output_against_source(args.base, args.output, auto_updates)
    reason_counts = Counter()
    for item in auto_updates.values():
        for reason in item["reasons"]:
            reason_counts[str(reason)] += 1
    summary["highlighted_cells"] = len(auto_updates)
    summary["highlight_reasons"] = dict(sorted(reason_counts.items()))
    summary["source_consistency"] = source_consistency
    summary["auto_updates"] = sorted(
        auto_updates.values(),
        key=lambda item: (str(item["sheet"]), int(item["row"]), str(item["column"])),
    )
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    if source_consistency["unexpected_diff_count"]:
        raise RuntimeError(
            f"Output changed {source_consistency['unexpected_diff_count']} cells not recorded as approved updates; "
            f"see source_consistency in {args.report}"
        )
    print(json.dumps({"output": str(args.output), "report": str(args.report), "diff_rows": summary["diff_rows"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
