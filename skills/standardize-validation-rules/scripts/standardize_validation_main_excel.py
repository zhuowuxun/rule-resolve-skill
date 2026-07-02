#!/usr/bin/env python3
"""Standardize validation main workbooks with an XML-level XLSX editor.

Applies the main validation rules to workbooks such as `t_1.xlsx`:
- `Actions`: standardize `cn_name`, `cn_desc`, `cn_notes`
- `Sequences`: standardize `cn_name`, `cn_desc`

This script intentionally preserves workbook structure and formatting by
editing worksheet XML directly.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import zipfile
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import unquote, urlparse
import xml.etree.ElementTree as ET


NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "p": "http://schemas.openxmlformats.org/package/2006/relationships",
}

REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
SHEET_MAIN_NS = NS["a"]

URL_RE = re.compile(r"https?://[^\s<>\]]+")
MARKDOWN_LINK_RE = re.compile(r"\[.*?\]\((https?://[^)]+)\)")
CVE_RE = re.compile(r"\bCVE[-\s]*(\d{4})[-\s]*(\d{4,})\b", re.IGNORECASE)
VARIANT_RE = re.compile(r"(?i)(?:variant|变种|方法)\s*[-# ]?\s*(\d+)")
CAMPAIGN_CODE_RE = re.compile(r"\bCAMP\.\d{2,4}\.\d{2,4}\b", re.IGNORECASE)
OS_TOKEN_RE = re.compile(
    r"(Windows Server \d{4}|Windows \d{1,2}|Windows|Ubuntu(?: \d+\.\d+)?|CentOS(?: \d+)?|Debian(?: \d+)?|Linux|macOS|Mac OS)"
)

WEB_PREFIX = "Web安全验证 - "
WEB_NOTE_DEFAULT = "塞讯验证建议在外部/不受信、内部/受信的安全区域中选择源验证机器人，在目标/DMZ区域中选择目标验证机器人。"
VALIDATION_TITLE_PREFIXES = (
    "恶意文件传输 - ",
    "主机命令行 - ",
    "命令与控制 - ",
    "受保护的沙盘 - ",
    "受保护剧场 - ",
    "受保护的剧场 - ",
    "Web安全验证 - ",
    "Web 安全验证 - ",
    "web安全验证 - ",
    "Web应用程序漏洞 - ",
    "应用程序漏洞 - ",
    "AI应用程序漏洞 - ",
    "工控安全 - ",
    "OT安全 - ",
    "钓鱼邮件 - ",
)


def format_cve_match(match: re.Match[str]) -> str:
    return f"CVE-{match.group(1)}-{match.group(2)}".upper()


def extract_cve(text: str) -> str:
    match = CVE_RE.search(text or "")
    return format_cve_match(match) if match else ""

HARDWARE_KEYWORDS = (
    "管理系统",
    "堡垒机",
    "OneView",
    "ITSM",
    "服务器",
    "ERP",
    "校园",
    "调度",
    "审计",
    "指挥",
)

AI_APPLICATION_PRODUCTS = (
    "langflow",
    "librechat",
    "mindsdb",
    "open webui",
    "openclaw",
    "mlflow",
    "nocobase",
    "litellm",
    "opencode",
    "short-video-maker",
)

AI_APPLICATION_DESC_MARKERS = (
    "AI 驱动",
    "AI驱动",
    "AI 编程智能体",
    "AI编程智能体",
    "AI Agent",
    "AI agent",
    "LLM",
    "大模型",
    "MCP 协议",
    "Model Context Protocol",
)

INDUSTRIAL_CONTROL_PRODUCTS = (
    "lean mes",
    "深科特",
)

INDUSTRIAL_CONTROL_DESC_MARKERS = (
    "制造执行系统",
    "生产过程管理",
    "SCADA",
    "数据采集与监控",
    "工业协议",
    "工业设备",
    "协议转换网关",
    "通信网关",
    "智能仓储",
    "智能排程",
    "设备夹具",
    "精益制造",
)

APPLICATION_VULN_PRODUCTS = (
    "citrix netscaler",
    "d-link nas",
    "fortinet forticlientems",
    "fortinet fortisandbox",
    "infoblox netmri",
    "jumpserver",
    "深信服运维安全管理系统",
)

EXTENSION_OS_MAP = {
    "APP": "macOS",
    "DMG": "macOS",
    "MACHO": "macOS",
    "PKG": "macOS",
    "SCPT": "macOS",
}

CITY_NAMES = [
    "北京", "上海", "广州", "深圳", "杭州", "南京", "苏州", "无锡", "常州",
    "宁波", "温州", "嘉兴", "绍兴", "金华", "台州", "天津", "重庆", "成都",
    "武汉", "西安", "郑州", "长沙", "合肥", "济南", "青岛", "福州", "厦门",
    "泉州", "东莞", "佛山", "中山", "珠海", "昆明", "南宁", "南昌", "石家庄",
    "沈阳", "大连", "长春", "哈尔滨", "海口", "贵阳", "乌鲁木齐", "呼和浩特",
    "兰州", "太原", "唐山", "烟台", "潍坊", "临沂", "徐州", "南通", "盐城",
]

PROVINCE_NAMES = [
    "河北", "山西", "辽宁", "吉林", "黑龙江", "江苏", "浙江", "安徽", "福建", "江西",
    "山东", "河南", "湖北", "湖南", "广东", "海南", "四川", "贵州", "云南", "陕西",
    "甘肃", "青海", "台湾", "内蒙古", "广西", "西藏", "宁夏", "新疆",
]

NAME_ALIAS_MAP = {
    "神圣游戏": "SACREDGAME",
    "太空锤": "SPACEHAMMER",
}

ACTION_CONTEXTUAL_ALIAS_RULES = (
    (re.compile(r"(?<![A-Za-z0-9])Qilin(?![A-Za-z0-9])", re.IGNORECASE), "麒麟勒索软件", "Qilin 勒索软件"),
)

RANSOMWARE_REFERENCE_ENV = "VALIDATION_RANSOMWARE_REFERENCE_WORKBOOK"


def qname(local: str) -> str:
    return f"{{{SHEET_MAIN_NS}}}{local}"


def col_letters_to_index(letters: str) -> int:
    result = 0
    for char in letters:
        result = result * 26 + (ord(char) - 64)
    return result


def col_index_to_letters(index: int) -> str:
    letters: List[str] = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def split_ref(ref: str) -> Tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", ref)
    if not match:
        raise ValueError(f"Unsupported cell ref: {ref}")
    return col_letters_to_index(match.group(1)), int(match.group(2))


def normalize_text(value: str) -> str:
    return (value or "").replace("\r\n", "\n").replace("\r", "\n")


def normalize_org_relationships(text: str) -> str:
    value = text
    value = re.sub(
        r"(?P<child>APT-[A-Z]*U?\d+)\s*(?P<verb>被怀疑是|被认为是|被视为|被描述为|是)\s*(?:[^。；，]*?)?(?P<parent>APT-[A-Z]*U?\d+)\s*的(?:一个|一支|一)?(?:子组织|子集群|分支|下属组织|附属组织)",
        r"\g<child> \g<verb> \g<parent> 的子组织",
        value,
    )
    value = re.sub(
        r"(?P<child>APT-[A-Z]*U?\d+)\s*隶属于\s*(?:[^。；，]*?)?(?P<parent>APT-[A-Z]*U?\d+)",
        r"\g<child> 隶属于 \g<parent>，为其子组织",
        value,
    )
    return value


def regex_sub_outside_urls(pattern: re.Pattern[str], repl, text: str) -> str:
    """Apply a regex replacement without mutating protected URL strings."""
    pieces: List[str] = []
    last = 0
    for match in URL_RE.finditer(text):
        pieces.append(pattern.sub(repl, text[last:match.start()]))
        pieces.append(match.group(0))
        last = match.end()
    pieces.append(pattern.sub(repl, text[last:]))
    return "".join(pieces)


class XlsxArchive:
    def __init__(self, path: Path) -> None:
        self.path = path
        with zipfile.ZipFile(path) as archive:
            self.files: Dict[str, bytes] = {name: archive.read(name) for name in archive.namelist()}

    def write(self, output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, data in self.files.items():
                archive.writestr(name, data)


def load_shared_strings(files: Dict[str, bytes]) -> List[str]:
    if "xl/sharedStrings.xml" not in files:
        return []
    root = ET.fromstring(files["xl/sharedStrings.xml"])
    values: List[str] = []
    for item in root.findall("a:si", NS):
        values.append("".join(text.text or "" for text in item.iterfind(".//a:t", NS)))
    return values


def _sheet_target_for_name(files: Dict[str, bytes], sheet_name: str) -> Optional[str]:
    for name, target in workbook_sheet_targets(files):
        if name == sheet_name:
            return target
    return None


def _rows_from_sheet(files: Dict[str, bytes], target: str) -> List[Dict[str, str]]:
    shared_strings = load_shared_strings(files)
    root = ET.fromstring(files[target])
    rows: List[Dict[str, str]] = []
    for row in root.findall(".//a:sheetData/a:row", NS):
        values: Dict[str, str] = {}
        for cell in row.findall("a:c", NS):
            ref = cell.attrib.get("r", "")
            col = "".join(ch for ch in ref if ch.isalpha())
            values[col] = read_cell_value(cell, shared_strings).strip()
        rows.append(values)
    return rows


def workbook_sheet_targets(files: Dict[str, bytes]) -> List[Tuple[str, str]]:
    workbook_root = ET.fromstring(files["xl/workbook.xml"])
    rel_root = ET.fromstring(files["xl/_rels/workbook.xml.rels"])
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rel_root.findall("p:Relationship", NS)}
    targets: List[Tuple[str, str]] = []
    for sheet in workbook_root.find("a:sheets", NS):
        name = sheet.attrib["name"]
        rel_id = sheet.attrib[f"{{{REL_NS}}}id"]
        targets.append((name, "xl/" + rel_map[rel_id]))
    return targets


def read_cell_value(cell: Optional[ET.Element], shared_strings: List[str]) -> str:
    if cell is None:
        return ""
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.iterfind(".//a:t", NS))
    value_node = cell.find("a:v", NS)
    if value_node is None:
        return ""
    raw_value = value_node.text or ""
    if cell_type == "s" and raw_value.isdigit():
        idx = int(raw_value)
        if idx < len(shared_strings):
            return shared_strings[idx]
    return raw_value


def row_cells_by_col(row: ET.Element) -> Dict[int, ET.Element]:
    mapping: Dict[int, ET.Element] = {}
    for cell in row.findall("a:c", NS):
        ref = cell.attrib.get("r", "")
        if not ref:
            continue
        col_idx, _ = split_ref(ref)
        mapping[col_idx] = cell
    return mapping


def ensure_cell(row: ET.Element, col_idx: int, row_idx: int) -> ET.Element:
    existing = row_cells_by_col(row)
    if col_idx in existing:
        return existing[col_idx]

    new_cell = ET.Element(qname("c"))
    new_cell.attrib["r"] = f"{col_index_to_letters(col_idx)}{row_idx}"

    inserted = False
    cells = row.findall("a:c", NS)
    for idx, cell in enumerate(cells):
        current_col, _ = split_ref(cell.attrib["r"])
        if current_col > col_idx:
            row.insert(idx, new_cell)
            inserted = True
            break
    if not inserted:
        row.append(new_cell)
    return new_cell


def set_inline_string(cell: ET.Element, value: str) -> None:
    value = normalize_text(value)
    for child in list(cell):
        cell.remove(child)
    cell.attrib["t"] = "inlineStr"
    is_node = ET.SubElement(cell, qname("is"))
    t_node = ET.SubElement(is_node, qname("t"))
    if value.startswith(" ") or value.endswith(" ") or "\n" in value:
        t_node.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"
    t_node.text = value


def normalize_variant(text: str) -> str:
    def repl(match: re.Match[str]) -> str:
        return f"变种 #{match.group(1)}"

    return VARIANT_RE.sub(repl, text)


def normalize_common_text(text: str) -> str:
    value = normalize_text(text)
    value = value.replace("_x000D_", "")
    value = value.replace("、", "，")
    value = value.replace("linuxOS", "Linux")
    value = value.replace("LinuxOS", "Linux")
    value = re.sub(r"(?<![A-Za-z0-9-])APT[\s-]*U[\s-]*(\d+)(?![A-Za-z0-9-])", r"APT-U\1", value)
    value = regex_sub_outside_urls(CVE_RE, format_cve_match, value)
    value = normalize_org_relationships(value)
    # Repair common translation mistakes where product names were over-normalized.
    value = value.replace("AWS EC&C", "AWS EC2")
    value = value.replace("Amazon Web Services (AWS) EC&C", "Amazon Web Services (AWS) EC2")
    value = value.replace("EC&C 实例元数据服务", "EC2 实例元数据服务")
    value = value.replace("AWS EC&C 实例元数据服务", "AWS EC2 实例元数据服务")
    value = value.replace("TP Link", "TP-Link")
    value = value.replace("Add MpPreference", "Add-MpPreference")
    value = value.replace("Msi 安装程序", "msi 安装程序")
    value = value.replace("AdaptixC&C", "AdaptixC2")
    value = value.replace("Bluenoroff", "BlueNoroff")
    value = value.replace("Axios供应链", "Axios 供应链")
    value = value.replace("电报", "Telegram")
    value = value.replace("CJIAJIA", "C++").replace("C+JIAJIA", "C++")
    value = re.sub(r"C&C\s*或\s*C&C", "C&C", value)
    value = re.sub(r"C&C\s*和\s*C&C", "C&C", value)
    value = value.replace("威胁行为体", "威胁组织")
    value = value.replace("行为体", "威胁组织")
    value = value.replace("恶意软件集群", "威胁组织")
    value = value.replace("威胁集群", "威胁组织")
    value = value.replace("子集群", "威胁组织")
    value = value.replace("威胁威胁组织", "威胁组织")
    value = value.replace("攻击技巧", "攻击手法")
    value = value.replace("系统变种", "系统版本")
    value = value.replace("活动集群", "威胁组织")
    value = value.replace("敌对集群", "威胁组织")
    value = value.replace("集群", "威胁组织")
    value = value.replace("网络间谍组织", "威胁组织")
    value = value.replace("网络间谍活动集群", "威胁组织")
    value = value.replace("网络钓鱼电子邮件", "钓鱼邮件")
    value = value.replace("为企事业单位", "为企业")
    value = value.replace("面向企事业单位", "面向企业")
    value = value.replace("服务于企事业单位", "服务于企业")
    value = value.replace("该操作", "该验证动作")
    value = value.replace("此操作", "此验证动作")
    value = value.replace("登记", "签入")
    value = value.replace("丢弃", "投放")
    value = re.sub(r"\b[Dd]ropper\b", "释放器", value)
    value = re.sub(r"\b[Dd]roppers\b", "释放器", value)
    value = re.sub(r"\b[Mm]alwaredropper\b", "恶意软件释放器", value)
    value = re.sub(r"\.Exe\b", ".exe", value)
    value = re.sub(r"\.Dll\b", ".dll", value)
    value = re.sub(r"\.Bat\b", ".bat", value)
    value = re.sub(r"\.Ps1\b", ".ps1", value)
    value = re.sub(r"\.Vbs\b", ".vbs", value)
    value = re.sub(r"\.Hta\b", ".hta", value)
    value = re.sub(r"\.Tmp\b", ".tmp", value)
    value = re.sub(r"\.Xls(?=\b|[\u4e00-\u9fff])", ".xls", value)
    value = re.sub(r"\.Doc(?=\b|[\u4e00-\u9fff])", ".doc", value)
    value = re.sub(r"\.Ppt(?=\b|[\u4e00-\u9fff])", ".ppt", value)
    value = re.sub(r"C\+\+语言", "C++ 语言", value)
    value = re.sub(r"([A-Za-z0-9+#)])(?=[\u4e00-\u9fff])", r"\1 ", value)
    value = re.sub(r"(?<=[\u4e00-\u9fff])([A-Za-z0-9(])", r" \1", value)
    value = normalize_variant(value)
    value = re.sub(r"(?<!\d)(20\d{2})\s+(\d{2})\s+(\d{2})(?!\d)", r"\1-\2-\3", value)
    value = re.sub(r"披露时间\s*[:：]\s*(20\d{2})[-\s/]*(\d{2})[-\s/]*(\d{2})", r"披露时间：\1-\2-\3", value)
    value = value.replace("。，", "。")
    value = re.sub(r"\b([A-Z][A-Z0-9_.-]{2,})。\s+\1。", r"\1。", value)
    value = re.sub(r"([。！？])\s+(?=[A-Za-z0-9\u4e00-\u9fff])", r"\1", value)
    value = re.sub(r"-\s*\[\*\]\s*-", "", value)
    value = re.sub(r"[ \t]+\n", "\n", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    value = re.sub(r" {2,}", " ", value)
    return value.strip()


def normalize_contextual_alias_spacing(text: str) -> str:
    value = normalize_common_text(text)
    value = re.sub(r"(?<=[\u4e00-\u9fff])\s*(Qilin)\b", r" \1", value)
    value = re.sub(r"\b(Qilin)\s*(?=[\u4e00-\u9fff])", r"\1 ", value)
    value = re.sub(r"又名\s*([A-Za-z0-9_.-]+)", r"又名 \1", value)
    return value


def infer_contextual_aliases_from_action(name: str, desc: str) -> Dict[str, str]:
    clean_name = normalize_common_text(name)
    clean_desc = normalize_common_text(desc)
    aliases: Dict[str, str] = {}
    for marker_re, source, target in ACTION_CONTEXTUAL_ALIAS_RULES:
        if source in clean_name and marker_re.search(clean_desc):
            aliases[source] = target
    return aliases


def apply_contextual_aliases(text: str, aliases: Dict[str, str]) -> str:
    value = normalize_common_text(text)
    for source, target in aliases.items():
        source_base = source.removesuffix("勒索软件").strip()
        target_base = target.removesuffix(" 勒索软件").removesuffix("勒索软件").strip()
        if source not in value and not (source_base and source_base in value):
            continue

        if source_base and target_base:
            parenthetical_re = re.compile(
                rf"{re.escape(source_base)}\s*[（(]\s*{re.escape(target_base)}"
                rf"(?:\s*[，,]\s*又名\s*([^）)]+))?\s*[）)]"
            )

            def parenthetical_repl(match: re.Match[str]) -> str:
                aka = (match.group(1) or "").strip()
                if aka:
                    return f"{target_base}（又名 {aka}）"
                return target_base

            value = parenthetical_re.sub(parenthetical_repl, value)
        value = value.replace(source, target)
    return normalize_contextual_alias_spacing(value)


def normalize_geo_company_text(text: str) -> str:
    normalized = normalize_common_text(text)
    company_suffix = r"(?:有限公司|科技有限公司|信息技术有限公司|软件有限公司|股份有限公司|有限责任公司)"
    for place in PROVINCE_NAMES + CITY_NAMES:
        normalized = re.sub(rf"(^|[。；;!?！？]\s*){place}(?=[^，。；;、]{{0,30}}{company_suffix})", r"\1", normalized)
        normalized = re.sub(rf"(?<=是)\s*{place}(?=[^，。；;、]{{0,30}}{company_suffix})", "", normalized)
        normalized = re.sub(rf"是{place}(?=[^，。；;、]{{0,30}}{company_suffix})", "是", normalized)
        normalized = re.sub(rf"(?<=[（(、，,\s]){place}(?=[^，。；;、]{{0,20}}{company_suffix})", "", normalized)
        normalized = re.sub(rf"(?<=是){place}(?=[\u4e00-\u9fffA-Za-z0-9]{{2,16}}(?:自主研发|推出|开发|一同推出))", "", normalized)
        normalized = re.sub(rf"(?<=与){place}(?=[\u4e00-\u9fffA-Za-z0-9]{{2,16}}(?:自主研发|推出|开发|一同推出))", "", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _normalize_lookup_name(text: str) -> str:
    value = normalize_common_text(text).lower()
    value = value.replace(" 勒索软件", "")
    value = re.sub(r"\s+", "", value)
    return value


@lru_cache(maxsize=1)
def load_known_ransomware_families() -> set[str]:
    families: set[str] = set()
    configured_path = os.environ.get(RANSOMWARE_REFERENCE_ENV, "").strip()
    if not configured_path:
        return families
    path = Path(configured_path).expanduser()
    if not path.exists():
        return families
    try:
        archive = XlsxArchive(path)
        target = _sheet_target_for_name(archive.files, "malware_rule_rows") or _sheet_target_for_name(archive.files, "malware_tag_flat")
        if not target:
            return families
        rows = _rows_from_sheet(archive.files, target)
        for row in rows[1:]:
            if normalize_common_text(row.get("F", "")) != "勒索软件" and normalize_common_text(row.get("G", "")) != "勒索软件":
                continue
            for key in ("D", "E", "B"):
                value = normalize_common_text(row.get(key, ""))
                if not value:
                    continue
                if key == "B":
                    match = re.search(r"-\s*([^，,]+?)\s*勒索软件", value)
                    if match:
                        value = match.group(1).strip()
                    else:
                        continue
                families.add(_normalize_lookup_name(value))
    except Exception:
        return families
    return {item for item in families if item}


def collect_inline_ransomware_families(action_names: List[str]) -> set[str]:
    families: set[str] = set()
    for name in action_names:
        clean_name = normalize_common_text(name)
        if "勒索软件" not in clean_name:
            continue
        match = re.search(r"恶意文件传输\s*-\s*([^，,]+)\s+勒索软件", clean_name)
        if match:
            families.add(_normalize_lookup_name(match.group(1)))
            continue
        parts = [part.strip() for part in re.split(r"[，,]", clean_name.replace("恶意文件传输 - ", "")) if part.strip()]
        for part in parts:
            if "勒索软件" in part:
                families.add(_normalize_lookup_name(part.replace("勒索软件", "").strip()))
    return families


def infer_ransomware_family_from_desc(desc: str) -> set[str]:
    families: set[str] = set()
    text = normalize_common_text(desc)
    patterns = [
        r"\[\*\*([A-Za-z][A-Za-z0-9._-]{1,39})\*\*\]\s*-\s*\1\s*是.*?勒索软件",
        r"\b([A-Za-z][A-Za-z0-9._-]{1,39})\b\s*是.*?勒索软件",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text):
            families.add(_normalize_lookup_name(match.group(1)))
    return {item for item in families if item}


def apply_ransomware_suffix(name: str, known_families: set[str], desc: str = "") -> str:
    if not name.startswith("恶意文件传输 - "):
        return name
    if "勒索软件" in name:
        return name
    effective_families = set(known_families) | infer_ransomware_family_from_desc(desc)
    parts = [part.strip() for part in re.split(r"[，,]", name.replace("恶意文件传输 - ", "")) if part.strip()]
    if not parts:
        return name
    first_family_idx: Optional[int] = None
    for idx, part in enumerate(parts):
        normalized = _normalize_lookup_name(part)
        if not normalized or normalized in {
            "下载",
            "windows",
            "linux",
            "macos",
        }:
            continue
        if part.startswith("APT-") or part.startswith("变种 #") or "(" in part:
            continue
        if normalized in effective_families:
            first_family_idx = idx
            break
    if first_family_idx is None:
        return name
    parts[first_family_idx] = f"{parts[first_family_idx]} 勒索软件"
    return "恶意文件传输 - " + "，".join(parts)


def normalize_cn_action_terms(text: str) -> str:
    value = normalize_common_text(text)
    for src, dst in NAME_ALIAS_MAP.items():
        value = value.replace(src, dst)
    replacements = [
        ("Drops The Executable", "投放可执行文件"),
        ("Drop The Executable", "投放可执行文件"),
        ("Drops the executable", "投放可执行文件"),
        ("Drop the executable", "投放可执行文件"),
        ("Drops Backdoor", "投放后门"),
        ("Execution", "执行"),
        ("Drop Backdoor", "投放后门"),
        ("Drops ", "投放"),
        ("放置", "投放"),
    ]
    for src, dst in replacements:
        value = value.replace(src, dst)
    return value


def normalize_notes(text: str) -> str:
    value = normalize_common_text(text)
    value = value.replace("此验证动作需要受保护的沙盘才能正确执行。", "此验证动作需要在受保护的沙盘中才能正确执行。")
    value = value.replace("此验证动作需要受保护的沙盘才能正确执行", "此验证动作需要在受保护的沙盘中才能正确执行")
    value = value.replace("此验证动作需要在受保护的沙盘环境中才能正确执行。", "此验证动作需要在受保护的沙盘中才能正确执行。")
    value = value.replace("此验证动作需要在受保护的沙盘环境中才能正确执行", "此验证动作需要在受保护的沙盘中才能正确执行")
    value = value.replace("源 Actor", "源验证机器人")
    value = value.replace("目标 Actor", "目标验证机器人")
    value = value.replace("源 actor", "源验证机器人")
    value = value.replace("目标 actor", "目标验证机器人")
    value = value.replace("源验证机器人 和外部", "源验证机器人和外部")
    value = value.replace("中的源验证机器人 和", "中的源验证机器人和")
    value = re.sub(r"Windows 10，Windows 11，Windows Server 2016，Windows Server 2019，Windows Server 2022", "Windows 10、Windows 11、Windows Server 2016、Windows Server 2019、Windows Server 2022", value)
    value = value.replace("系统，非管理员，管理员", "系统、非管理员、管理员")
    value = value.replace("验证机器人 上", "验证机器人上")
    value = value.replace("验证机器人上以", "验证机器人上，以")
    return value.strip()


def extract_os_suffix(text: str) -> str:
    matches = OS_TOKEN_RE.findall(normalize_common_text(text))
    ordered: List[str] = []
    for item in matches:
        normalized = item.replace("Mac OS", "macOS")
        if normalized.startswith("Windows"):
            family = "Windows"
        elif normalized in {"Ubuntu", "Linux"} or normalized.startswith("Ubuntu") or normalized.startswith("CentOS") or normalized.startswith("Debian"):
            family = "Linux"
        elif normalized == "macOS":
            family = "macOS"
        else:
            family = normalized
        if family not in ordered:
            ordered.append(family)
    if not ordered:
        return ""
    return "(" + "/".join(ordered) + ")"


def infer_os_suffix_from_extension(name: str) -> str:
    systems: List[str] = []
    for ext in re.findall(r"\.([A-Za-z0-9]{2,8})\s*文件", normalize_common_text(name)):
        system = EXTENSION_OS_MAP.get(ext.upper())
        if system and system not in systems:
            systems.append(system)
    if not systems:
        return ""
    return "(" + "/".join(systems) + ")"


def append_os_suffix(name: str, notes: str) -> str:
    existing_match = re.search(r"\((?:Windows|Linux|macOS)(?:/(?:Windows|Linux|macOS))*\)", name)
    suffix = extract_os_suffix(notes) or (existing_match.group(0) if existing_match else "") or infer_os_suffix_from_extension(name)
    clean_name = re.sub(r"\s*\((?:Windows|Linux|macOS)(?:/(?:Windows|Linux|macOS))*\)", "", name).strip()
    if not suffix:
        return clean_name
    parts = [part.strip() for part in clean_name.split("，")]
    for idx, part in enumerate(parts):
        if "释放器" in part and idx > 0:
            # Generic `释放器` belongs to the preceding malware/tool, while a typed
            # releaser such as `Stage Script释放器` is itself the executable object.
            attach_idx = idx if re.search(r"[A-Za-z0-9].*释放器", part) else idx - 1
            while attach_idx > 0 and parts[attach_idx].startswith("APT-"):
                attach_idx -= 1
            if suffix not in parts[attach_idx]:
                parts[attach_idx] = f"{parts[attach_idx]} {suffix}"
            return move_os_suffix_before_action("，".join(parts))
    action_keywords = (
        "执行",
        "下载",
        "签入",
        "签到",
        "注册",
        "数据泄露",
        "数据聚合",
        "泄露",
        "渗透",
        "信标",
        "C&C",
        "DNS",
        "任务流量",
        "任务分配",
        "投放",
        "释放",
        "连接",
        "安装",
        "卸载",
        "创建",
        "侧加载",
        "计划任务",
        "持久化",
        "设置",
        "使用",
        "隐藏",
    )
    for idx in range(len(parts) - 1, -1, -1):
        part = parts[idx]
        if part.startswith("变种 #"):
            continue
        if any(keyword in part for keyword in action_keywords):
            attach_idx = idx - 1 if idx - 1 >= 0 else idx
            if attach_idx > 0 and "释放器" in parts[attach_idx]:
                attach_idx -= 1
            while attach_idx > 0 and parts[attach_idx].startswith("APT-"):
                attach_idx -= 1
            if suffix not in parts[attach_idx]:
                parts[attach_idx] = f"{parts[attach_idx]} {suffix}"
            return move_os_suffix_before_action("，".join(parts))
    variant_match = re.search(r"(，变种 #\d+)$", clean_name)
    if variant_match:
        head = clean_name[: variant_match.start()].rstrip()
        tail = variant_match.group(1)
        return move_os_suffix_before_action(f"{head} {suffix}{tail}")
    return move_os_suffix_before_action(f"{clean_name} {suffix}")


def move_os_suffix_before_action(name: str) -> str:
    """Keep OS suffix near the malware/tool name instead of after the action verb."""
    suffix_re = re.compile(r"\s*(\((?:Windows|Linux|macOS)(?:/(?:Windows|Linux|macOS))*\))$")
    parts = [part.strip() for part in name.split("，")]
    if len(parts) < 2:
        return name
    action_markers = (
        "执行",
        "下载",
        "签入",
        "注册",
        "通信",
        "连接",
        "安装",
        "卸载",
        "创建",
        "投放",
        "删除",
        "清除",
        "收集",
        "聚合",
        "读取",
        "写入",
        "泄露",
        "渗透",
        "信标",
        "C&C",
        "DNS",
        "持久化",
        "计划任务",
        "设置",
        "使用",
        "隐藏",
    )
    for idx, part in enumerate(parts):
        match = suffix_re.search(part)
        if not match:
            continue
        stem = suffix_re.sub("", part).strip()
        if ("APT-" in stem or "威胁组织" in stem or "威胁集团" in stem) and idx + 1 < len(parts):
            suffix = match.group(1)
            for next_idx in range(idx + 1, len(parts)):
                if re.search(r"\b[A-Za-z0-9_.-]+\.(?:exe|dll|sys|ps1|vbs|js|jar|sh|elf|bin)\b", parts[next_idx], re.IGNORECASE):
                    parts[next_idx] = re.sub(
                        r"(\b[A-Za-z0-9_.-]+\.(?:exe|dll|sys|ps1|vbs|js|jar|sh|elf|bin)\b)",
                        rf"\1 {suffix}",
                        parts[next_idx],
                        count=1,
                        flags=re.IGNORECASE,
                    )
                    parts[idx] = stem
                    return "，".join(part for part in parts if part)
            for next_idx in range(idx + 1, len(parts)):
                if parts[next_idx].startswith("变种 #"):
                    continue
                if suffix_re.search(parts[next_idx]):
                    parts[idx] = stem
                    return "，".join(part for part in parts if part)
                parts[next_idx] = f"{parts[next_idx]} {suffix}"
                parts[idx] = stem
                return "，".join(part for part in parts if part)
        if not stem or not any(marker in stem for marker in action_markers):
            continue
        attach_idx = idx - 1
        while attach_idx > 0 and (
            parts[attach_idx].startswith("APT-")
            or parts[attach_idx].startswith("变种 #")
            or parts[attach_idx] in {"释放器", "恶意软件释放器"}
        ):
            attach_idx -= 1
        if attach_idx < 0:
            continue
        if suffix_re.search(parts[attach_idx]):
            parts[idx] = stem
            continue
        parts[attach_idx] = f"{parts[attach_idx]} {match.group(1)}"
        parts[idx] = stem
    return "，".join(part for part in parts if part)


def clean_url(url: str) -> str:
    return url.rstrip(".,;)\u3002")


def clean_campaign_codes_from_title(title: str) -> str:
    if not CAMPAIGN_CODE_RE.search(title):
        return title
    value = normalize_common_text(title)
    if " - " in value:
        prefix, rest = value.split(" - ", 1)
        parts = [part.strip() for part in rest.split("，") if part.strip() and not CAMPAIGN_CODE_RE.fullmatch(part.strip())]
        return f"{prefix} - " + "，".join(parts)
    parts = [part.strip() for part in value.split("，") if part.strip() and not CAMPAIGN_CODE_RE.fullmatch(part.strip())]
    return "，".join(parts)


def extract_uri_paths(text: str) -> List[str]:
    """Extract stable URI paths from descriptions or related Email bodies."""
    value = normalize_common_text(text)
    markdown_urls = {clean_url(match.group(1)) for match in MARKDOWN_LINK_RE.finditer(value)}
    paths: List[str] = []
    for url in URL_RE.findall(value):
        clean = clean_url(url)
        if clean in markdown_urls:
            continue
        parsed = urlparse(clean)
        path = unquote(parsed.path or "").strip()
        if path and path != "/" and path not in paths:
            paths.append(path)
    value_without_urls = URL_RE.sub(" ", value)
    for match in re.finditer(r"(?<![:/A-Za-z0-9])/(?:[A-Za-z0-9._~%+-]+/)*[A-Za-z0-9._~%+-]+/?", value_without_urls):
        path = unquote(match.group(0).rstrip("`'\"”）),，。；;"))
        if path and path != "/" and path not in paths:
            paths.append(path)
    return paths


def insert_title_part_before_action(title: str, part: str) -> str:
    if not part or part in title:
        return title
    prefix, rest = title.split(" - ", 1) if " - " in title else ("", title)
    parts = [item.strip() for item in rest.split("，") if item.strip()]
    action_markers = (
        "注册",
        "签入",
        "通信",
        "连接",
        "信标",
        "任务流量",
        "任务请求",
        "下载",
        "执行",
    )
    insert_at = len(parts)
    for idx, item in enumerate(parts):
        if item.startswith("变种 #"):
            insert_at = idx
            break
        if any(marker in item for marker in action_markers):
            insert_at = idx
            break
    parts.insert(insert_at, part)
    body = "，".join(parts)
    return f"{prefix} - {body}" if prefix else body


def strip_placeholder_link_chunks(text: str) -> str:
    value = text or ""
    # Delete any disposable chunk shaped like `- [任意值] - ...url...`
    value = re.sub(r"-\s*\[[^\]]*?\]\s*-\s*\[[^\]]*?\]\(https?://[^)]+\)\s*-?", " ", value)
    value = re.sub(r"-\s*\[[^\]]*?\]\s*-\s*（https?://[^）]+）\s*-?", " ", value)
    value = re.sub(r"-\s*\[[^\]]*?\]\s*-\s*\(https?://[^)]+\)\s*-?", " ", value)
    value = re.sub(r"-\s*\[[^\]]*?\]\s*-\s*https?://\S+\s*-?", " ", value)
    return value


def split_references(text: str) -> Tuple[str, List[str]]:
    sanitized = strip_placeholder_link_chunks(text)
    protected_dot = "__PROTECTED_DEFANGED_DOT__"
    sanitized = sanitized.replace("[.]", protected_dot)
    urls: List[str] = []
    for url in MARKDOWN_LINK_RE.findall(sanitized):
        cleaned = clean_url(url)
        if cleaned and cleaned not in urls:
            urls.append(cleaned)
    for url in URL_RE.findall(sanitized):
        cleaned = clean_url(url)
        if cleaned and cleaned not in urls:
            urls.append(cleaned)
    body = MARKDOWN_LINK_RE.sub("", sanitized)
    body = URL_RE.sub("", body)
    body = re.sub(r"\[\*\*(" + CAMPAIGN_CODE_RE.pattern + r")\*\*\]\s*-\s*", "攻击活动。", body, flags=re.IGNORECASE)
    body = re.sub(r"\[\*\*(" + CAMPAIGN_CODE_RE.pattern + r")\*\*\]", "攻击活动", body, flags=re.IGNORECASE)
    body = CAMPAIGN_CODE_RE.sub("攻击活动", body)
    body = re.sub(r"\[\*\*.*?\*\*\]", " ", body)
    body = re.sub(r"\[[^\]]+\]", " ", body)
    body = re.sub(r"\s+-\s+\[\*\*.*?\*\*\]\s+-\s*", " ", body)
    body = re.sub(r"\s+-\s+\[.*?\]\s+-\s*", " ", body)
    body = re.sub(r"\s+-\s+", " ", body)
    body = body.replace(protected_dot, "[.]")
    return normalize_common_text(body), urls[:5]


def build_reference_block(urls: Iterable[str]) -> str:
    unique_urls: List[str] = []
    for url in urls:
        cleaned = clean_url(url)
        if cleaned and cleaned not in unique_urls:
            unique_urls.append(cleaned)
    if not unique_urls:
        return ""
    return "\n\n请参考：\n" + "\n".join(unique_urls)


def ensure_terminal_punctuation(text: str) -> str:
    value = text.rstrip()
    if value and value[-1] not in "。！？.!?）)】]":
        value += "。"
    return value


def split_sentences(text: str) -> List[str]:
    value = normalize_common_text(text)
    if not value:
        return []
    parts = re.split(r"(?<=[。！？])", value)
    sentences = [part.strip() for part in parts if part and part.strip()]
    return sentences


def cleanup_unwanted_attribution(text: str) -> str:
    if not text:
        return ""
    blocked_markers = (
        "与中国有关联",
        "中国有关联",
        "中国相关",
        "中国支持",
        "活动最早可追溯到",
        "并有证据表明其利用了各种边缘设备的漏洞",
        "利用了各种边缘设备的漏洞",
        "分发集群",
        "他们散布",
        "他们分发",
        "犯罪团伙",
        "归属于 FireEye 跟踪的未分类威胁组织的指标或活动",
        "归属于 FireEye 跟踪的未分类威胁组织",
    )
    sentences = split_sentences(text)
    kept = []
    blocked_patterns = (
        r".*APT-U\d+.*(?:分发集群|犯罪团伙).*",
        r".*FAKEUPDATES.*(?:下载器|释放器).*",
        r".*[A-Z][A-Za-z0-9._-]{2,39}\s*是一个用.*(?:下载器|释放器|后门|木马|勒索软件).*",
        r".*(?:他们散布|他们分发).*(?:FAKEUPDATES|恶意软件).*",
        r".*FireEye.*未分类威胁组织.*",
    )
    for sentence in sentences:
        if any(marker in sentence for marker in blocked_markers):
            continue
        if any(re.match(pattern, sentence) for pattern in blocked_patterns):
            continue
        kept.append(sentence)
    cleaned = " ".join(s.strip() for s in kept if s.strip())
    cleaned = re.sub(r"(^|。)\s*-?\s*攻击活动。.*$", r"\1", cleaned).strip()
    cleaned = re.sub(r"(?<=\s)-\s+", " ", cleaned)
    cleaned = re.sub(r"(?<!\d)(20\d{2})\s+(\d{2})\s+(\d{2})(?!\d)", r"\1-\2-\3", cleaned)
    cleaned = re.sub(r"披露时间\s*[:：]\s*(20\d{2})[-\s/]*(\d{2})[-\s/]*(\d{2})", r"披露时间：\1-\2-\3", cleaned)
    cleaned = cleaned.replace("*", "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned and text.strip():
        # Do not let attribution cleanup erase valid validation-action evidence.
        cleaned = normalize_common_text(text)
    return cleaned


def normalize_web_spacing(text: str) -> str:
    value = text
    value = re.sub(r"针对([A-Za-z0-9])", r"针对 \1", value)
    value = re.sub(r"([A-Za-z0-9/._-])，", r"\1，", value)
    value = re.sub(r"针对 ([A-Za-z0-9][^，。]*?)(存在)", r"针对 \1 \2", value)
    return value


def parse_web_name(name: str) -> Tuple[str, str, str, str]:
    raw = normalize_common_text(name)
    cve_match = CVE_RE.search(raw)
    cve = format_cve_match(cve_match) if cve_match else ""
    raw = CVE_RE.sub("", raw).strip()
    raw = re.sub(r"[（(]\s*[）)]", "", raw).strip()
    raw = re.sub(r"(?i)^web\s*安全验证\s*-\s*", "", raw).strip()
    raw = re.sub(r"^Web应用程序漏洞\s*-\s*", "", raw).strip()
    raw = re.sub(r"^应用程序漏洞\s*-\s*", "", raw).strip()
    raw = re.sub(r"^AI应用程序漏洞\s*-\s*", "", raw).strip()
    raw = re.sub(r"^工控安全\s*-\s*", "", raw).strip()
    raw = re.sub(r"^OT安全\s*-\s*", "", raw).strip()
    parts = [part.strip() for part in raw.split(" - ") if part.strip()]
    target = parts[0] if parts else raw
    vuln_type = parts[1] if len(parts) > 1 else ""
    product = target
    entry = ""
    if " /" in target:
        product, entry = target.rsplit(" /", 1)
        entry = "/" + entry.strip()
    else:
        tokens = target.split()
        if len(tokens) >= 2 and re.fullmatch(r"[A-Za-z0-9_./%-]+", tokens[-1]):
            product = " ".join(tokens[:-1]).strip()
            entry = tokens[-1].strip()
    if entry and not vuln_type and "-" in entry and "漏洞" in entry:
        split_entry, split_vuln = re.split(r"\s*-\s*", entry, maxsplit=1)
        entry = split_entry.strip()
        vuln_type = split_vuln.strip()
    return product.strip(), entry.strip(), vuln_type.strip(), cve


def normalize_web_entry_path(path: str) -> str:
    value = normalize_common_text(path).strip().strip("，,。.;；:：)）]】")
    if not value:
        return ""
    value = re.sub(r"^/+", "/", value)
    return value


def extract_web_entry_candidates(text: str) -> List[str]:
    body, _ = split_references(text)
    candidates: List[str] = []
    for match in re.finditer(r"(?<!:)/+[A-Za-z0-9._~/%-]+", body):
        candidate = normalize_web_entry_path(match.group(0))
        if not candidate:
            continue
        # Avoid obvious prose/date fragments; references are already split out.
        if re.search(r"\.(?:html?|md)$", candidate, flags=re.IGNORECASE):
            continue
        if candidate not in candidates:
            candidates.append(candidate)
    return candidates


def prefer_desc_web_entry(entry: str, desc: str) -> str:
    current = normalize_web_entry_path(entry)
    candidates = extract_web_entry_candidates(desc)
    if not current:
        return candidates[0] if candidates else ""
    current_key = re.sub(r"[^a-z0-9]+", "", current.lower())
    current_tail = current.rstrip("/").split("/")[-1].lower()
    for candidate in candidates:
        candidate_key = re.sub(r"[^a-z0-9]+", "", candidate.lower())
        candidate_tail = candidate.rstrip("/").split("/")[-1].lower()
        if candidate == current:
            return current
        if len(candidate) > len(current) and (
            current_key and current_key in candidate_key
            or current_tail and current_tail == candidate_tail
        ):
            return candidate
    return current


def is_hardware_like(product: str, desc: str) -> bool:
    return any(keyword in product for keyword in HARDWARE_KEYWORDS)


def is_web_entry_path(entry: str) -> bool:
    return bool(entry and re.match(r"^/[A-Za-z0-9._~/%-]+", entry))


def is_industrial_control_target(product: str, desc: str) -> bool:
    product_key = normalize_common_text(product).lower()
    desc_text = normalize_common_text(desc)
    return any(item in product_key for item in INDUSTRIAL_CONTROL_PRODUCTS) or any(
        marker in desc_text for marker in INDUSTRIAL_CONTROL_DESC_MARKERS
    )


def select_web_vuln_prefix(product: str, desc: str, entry: str = "") -> str:
    product_key = normalize_common_text(product).lower()
    desc_text = normalize_common_text(desc)
    if any(item in product_key for item in AI_APPLICATION_PRODUCTS):
        return "AI应用程序漏洞"
    if any(marker in desc_text for marker in AI_APPLICATION_DESC_MARKERS):
        return "AI应用程序漏洞"
    if is_industrial_control_target(product, desc):
        return "工控安全"
    if any(item in product_key for item in APPLICATION_VULN_PRODUCTS):
        return "应用程序漏洞"
    if is_web_entry_path(entry):
        return "Web应用程序漏洞"
    if is_hardware_like(product, desc):
        return "应用程序漏洞"
    return "Web应用程序漏洞"


def normalize_web_name(name: str, desc: str) -> str:
    product, entry, vuln_type, cve = parse_web_name(name)
    entry = prefer_desc_web_entry(entry, desc)
    prefix = select_web_vuln_prefix(product, desc, entry)
    parts = [product]
    if cve:
        parts.append(cve)
    if entry:
        parts.append(entry)
    if vuln_type:
        parts.append(vuln_type)
    return f"{prefix} - " + "，".join(part for part in parts if part)


def has_validation_title_prefix(name: str) -> bool:
    return normalize_common_text(name).startswith(VALIDATION_TITLE_PREFIXES)


def standardize_raw_vulnerability_name(name: str, desc: str, notes: str) -> str:
    raw = normalize_variant(normalize_cn_action_terms(name)).strip(" ，,。")
    if "漏洞" not in raw:
        return ""
    cve = extract_cve(f"{raw} {desc}")
    raw = CVE_RE.sub("", raw).strip(" ，,。")
    raw = re.sub(r"[（(]\s*[）)]", "", raw).strip(" ，,。")

    product = ""
    vuln = raw
    specific_patterns = [
        (r"^(Linux\s*PackageKit)\s*(权限提升漏洞)$", r"\1", r"\2"),
        (r"^(Linux\s*内核)(.+?权限提升漏洞)$", r"\1", r"\2"),
        (r"^(Windows\s*Defender)\s+(.+?权限提升漏洞)$", r"\1", r"\2"),
        (r"^(Google\s+Chrome\s+Blink\s+CSS\s+引擎)(释放后使用漏洞)$", r"\1", r"\2"),
    ]
    for pattern, product_template, vuln_template in specific_patterns:
        match = re.match(pattern, raw, flags=re.IGNORECASE)
        if match:
            product = match.expand(product_template).strip()
            vuln = match.expand(vuln_template).strip()
            break

    if not product:
        match = re.match(r"^(.+?)\s+([^，,。]*漏洞)$", raw)
        if match:
            product = match.group(1).strip()
            vuln = match.group(2).strip()
        else:
            match = re.match(r"^(.+?)(权限提升漏洞|释放后使用漏洞|远程代码执行漏洞|远程命令执行漏洞|SQL注入漏洞|SSRF漏洞|任意文件读取漏洞|任意文件上传漏洞)$", raw)
            if match:
                product = match.group(1).strip()
                vuln = match.group(2).strip()

    if not product:
        return ""

    product_key = normalize_common_text(product).lower()
    if any(item in product_key for item in AI_APPLICATION_PRODUCTS) or any(
        marker in normalize_common_text(desc) for marker in AI_APPLICATION_DESC_MARKERS
    ):
        prefix = "AI应用程序漏洞"
    elif is_industrial_control_target(product, desc):
        prefix = "工控安全"
    else:
        prefix = "应用程序漏洞"
    parts = [product]
    if cve:
        parts.append(cve)
    parts.append(vuln)
    return f"{prefix} - " + "，".join(part for part in parts if part)


def standardize_raw_vulnerability_transfer_name(name: str, desc: str, notes: str) -> str:
    clean_desc = normalize_common_text(desc)
    clean_notes = normalize_common_text(notes)
    if "漏洞" not in normalize_common_text(name):
        return ""
    if not re.search(r"源(?:验证机器人|AI攻防机器人).+目标(?:验证机器人|AI攻防机器人)", clean_notes):
        return ""
    if not any(marker in clean_desc for marker in ("可执行的漏洞利用程序", "可执行文件，该文件利用", "漏洞利用程序")):
        return ""
    vuln_title = standardize_raw_vulnerability_name(name, desc, notes)
    if not vuln_title or " - " not in vuln_title:
        return ""
    parts = [part.strip() for part in vuln_title.split(" - ", 1)[1].split("，") if part.strip()]
    if not parts:
        return ""
    vuln_idx = len(parts) - 1
    if not parts[vuln_idx].endswith("漏洞利用程序"):
        parts[vuln_idx] = f"{parts[vuln_idx]}利用程序"
    parts.append("下载")
    return "恶意文件传输 - " + "，".join(parts)


def extract_disclosure(sentences: List[str]) -> Tuple[List[str], str]:
    disclosure = ""
    remaining: List[str] = []
    for sentence in sentences:
        if "披露时间" in sentence and not disclosure:
            disclosure = re.sub(r"披露时间\s*[:：]\s*(\d{4})[-\s/]*(\d{2})[-\s/]*(\d{2})", r"披露时间：\1-\2-\3", sentence.strip())
            disclosure = ensure_terminal_punctuation(disclosure)
        else:
            remaining.append(sentence.strip())
    return remaining, disclosure


def is_software_description_sentence(sentence: str) -> bool:
    text = normalize_common_text(sentence)
    if not text:
        return False
    attack_keywords = (
        "漏洞",
        "攻击者",
        "利用尝试",
        "未授权",
        "未经身份验证",
        "远程",
        "注入",
        "读取",
        "上传",
        "泄露",
        "绕过",
        "参数",
        "请求",
        "接口",
        "端点",
        "触发",
        "构造",
        "执行",
    )
    if any(keyword in text for keyword in attack_keywords):
        return False
    description_markers = (
        "是一款",
        "是一个",
        "是一套",
        "是用于",
        "是面向",
        "推出的",
        "用于",
        "负责",
        "支持",
        "开源",
        "一款",
        "一个",
        "它填补了",
        "其核心",
        "可帮助",
        "帮助企业",
        "多供应商网络",
        "用于构建",
        "用于交付",
        "信息化服务商",
        "提供了",
        "为企业提供",
        "为城市",
    )
    if any(marker in text for marker in description_markers):
        return True
    if text.startswith(("它", "其", "该系统", "该平台", "该产品", "该框架", "该插件", "该软件")):
        return True
    return False


def dedupe_web_attack_sentences(
    sentences: List[str], product: str, entry: str, vuln_type: str
) -> List[str]:
    target = " ".join(part for part in (product, entry) if part).strip()
    product_compact = product.replace(" ", "")
    entry_tail = entry.rsplit("/", 1)[-1] if entry else ""
    deduped: List[str] = []
    seen: set[str] = set()
    for sentence in sentences:
        clean = normalize_common_text(sentence).strip()
        if not clean:
            continue
        if clean in seen:
            continue
        if vuln_type and vuln_type in clean and ("存在" in clean or "接口" in clean):
            mentions_target = bool(
                (target and target in clean)
                or (product and product in clean)
                or (product_compact and product_compact in clean.replace(" ", ""))
                or (entry and entry in clean)
                or (entry_tail and entry_tail in clean)
            )
            if mentions_target:
                remainder = re.sub(rf"^.*?{re.escape(vuln_type)}[，,]\s*", "", clean).strip()
                if remainder and remainder != clean:
                    clean = remainder
                else:
                    continue
        elif "漏洞" in clean and ("存在" in clean or "接口" in clean or "端点" in clean):
            mentions_target = bool(
                (target and target in clean)
                or (product and product in clean)
                or (product_compact and product_compact in clean.replace(" ", ""))
                or (entry and entry in clean)
                or (entry_tail and entry_tail in clean)
            )
            if mentions_target:
                remainder = re.sub(r"^.*?漏洞[，,]\s*", "", clean).strip()
                if remainder and remainder != clean:
                    clean = remainder
        if target and target in clean and vuln_type and vuln_type in clean and "利用尝试" in clean:
            continue
        seen.add(clean)
        deduped.append(clean)
    return deduped


def standardize_web_desc(name: str, desc: str) -> str:
    body, urls = split_references(desc)
    clean_body = cleanup_unwanted_attribution(normalize_geo_company_text(body))
    if not clean_body:
        return build_reference_block(urls).lstrip()

    sentences = split_sentences(clean_body)
    software_sentences: List[str] = []
    attack_sentences: List[str] = []
    switched = False
    for sentence in sentences:
        if not switched and is_software_description_sentence(sentence):
            software_sentences.append(sentence)
            continue
        switched = True
        attack_sentences.append(sentence)

    attack_sentences, disclosure = extract_disclosure(attack_sentences)
    product, entry, vuln_type, _ = parse_web_name(name)
    entry = prefer_desc_web_entry(entry, desc)
    attack_sentences = dedupe_web_attack_sentences(attack_sentences, product, entry, vuln_type)

    target = product
    if entry:
        target = f"{product} {entry}"
    action_sentence = f"此验证动作还原了针对 {target} 存在的{vuln_type}的利用尝试。"
    software_desc = " ".join(
        sentence if sentence.endswith(("。", "！", "？")) else f"{sentence}。"
        for sentence in software_sentences
    ).strip()

    pieces = [normalize_web_spacing(action_sentence)]
    if attack_sentences:
        pieces.extend(attack_sentences)
    if disclosure:
        pieces.append(disclosure)
    if software_desc:
        pieces.append(software_desc)

    text = " ".join(piece.strip() for piece in pieces if piece.strip())
    text = text.replace(
        "可以自动化，可视性和持续洞察帮助企业智能地管理其多供应商网络。",
        "可以通过自动化、可视性和持续洞察，帮助企业智能地管理多供应商网络。",
    )
    text = text.replace(
        "Infoblox NETMRI 是美国 Infoblox 公司的一个网络管理产品。",
        "Infoblox NETMRI 是美国 Infoblox 公司的一款网络管理产品。",
    )
    text = re.sub(r" +", " ", text)
    text = normalize_common_text(text.strip())
    text = re.sub(r"([。！？])\s+(?=[A-Za-z0-9\u4e00-\u9fff])", r"\1", text)
    text = text.replace(
        "生产过程管理，智能仓储管理，智能排程管理，品质管理， 供应商信息管理，设备夹具管理，SCADA",
        "生产过程管理、智能仓储管理、智能排程管理、品质管理、供应商信息管理、设备夹具管理、SCADA",
    )
    text = text.replace(
        "SCADA 数据采集与监控，BI 与大数据",
        "SCADA 数据采集与监控、BI 与大数据",
    )
    text = text.replace(
        "生产过程管理，智能仓储管理，智能排程管理，品质管理，供应商信息管理，设备夹具管理，SCADA",
        "生产过程管理、智能仓储管理、智能排程管理、品质管理、供应商信息管理、设备夹具管理、SCADA",
    )
    text = text.replace(
        "功能全面，平台化设计，智能化，全程电子化，移动办公等特点",
        "功能全面、平台化设计、智能化、全程电子化、移动办公等特点",
    )
    text = text.replace(
        "TikTok，Instagram Reels，YouTube Shorts",
        "TikTok、Instagram Reels、YouTube Shorts",
    )
    text = text.replace(
        "可以通过自动化，可视性和持续洞察，帮助企业智能地管理多供应商网络。",
        "可以通过自动化能力、可视化能力和持续洞察，帮助企业智能管理多供应商网络。",
    )
    text = text.replace(
        "帮助企业用好自己的客户资源，管好商机跟进过程，引导好业务员跟单行为，促进团队销售能力的提升；",
        "",
    )
    return ensure_terminal_punctuation(text) + build_reference_block(urls)


def extract_file_identity_text(desc: str) -> str:
    text = normalize_common_text(desc)
    patterns = (
        r"(?:该文件|这个文件|此文件|这是|该样本|文件)\s*(?:是|为|属于|充当|用作)\s*(?:一个|一种|一款|一段)?[^。；]{0,120}",
        r"这是\s*(?:一个|一种|一款|一段)?[^。；]{0,120}",
        r"(?:该文件|这个文件|此文件|该样本)\s*利用[^。；]{0,120}",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(0)
    return ""


def infer_file_type(name: str, desc: str) -> str:
    full_text = f"{name} {desc}"
    identity_text = extract_file_identity_text(desc)
    text = f"{name} {identity_text}" if identity_text else full_text
    if "Web Shell" in text or "JSP Web Shell" in text:
        return "Web Shell文件"
    if "恶意且经过混淆处理的 .NET 可执行文件" in text or ".NET 可执行文件" in text:
        return ".NET 可执行文件"
    if "恶意 Python 脚本" in text:
        return "恶意 Python 脚本"
    if "Python 脚本" in text:
        return "Python 脚本文件"
    if "32 位 Windows 动态链接库" in text or "恶意 32 位动态链接库" in text or "32 位 Windows DLL" in text:
        return "32 位 Windows .DLL文件"
    if "恶意 Windows 动态链接库" in text:
        return "恶意 Windows .DLL文件"
    if "恶意动态链接库" in text:
        return "恶意动态链接库文件"
    if "恶意软件组件" in text:
        return "恶意软件组件文件"
    if "木马化的软件组件" in text:
        return "木马化的软件组件文件"
    if "恶意配置脚本" in text:
        return "恶意配置脚本文件"
    if "恶意混淆脚本" in text:
        return "恶意混淆脚本文件"
    if "混淆脚本" in text:
        return "混淆脚本文件"
    if "恶意批处理脚本" in text:
        return "恶意批处理脚本文件"
    if "恶意配置文件" in text:
        return "恶意配置文件"
    if "JavaScript 下载脚本" in text or "javascript 下载脚本" in text:
        return "恶意 JavaScript 下载脚本"
    if "JavaScript 下载器" in text or "javascript 下载器" in text:
        return "恶意 JavaScript 下载器"
    if "JavaScript 脚本" in text or "javascript 脚本" in text or "JavaScript 文件" in text:
        return "恶意 JavaScript 脚本"
    if "64 位 Windows 可执行文件" in text or "64 位可执行文件" in text:
        return "恶意 64 位可执行文件"
    if "32 位 Windows 可执行文件" in text or "32 位可执行文件" in text:
        return "恶意 32 位可执行文件"
    if "远程访问可执行文件" in text:
        return "恶意远程访问可执行文件"
    if "恶意可执行文件" in text:
        return "恶意可执行文件"
    if "恶意移动应用程序" in text:
        return "恶意移动应用程序"
    if "Windows Installer 程序包" in text or "Windows Installer程序包" in text:
        return "恶意 Windows Installer 程序包"
    if "安装程序包" in text:
        return "恶意安装程序包"
    if "安装程序" in text:
        return "恶意安装程序"
    if "远程访问木马" in text:
        return "恶意远程访问木马文件"
    if "远程访问工具" in text:
        return "恶意远程访问工具"
    if "后门木马" in text:
        return "恶意后门木马文件"
    if "后门程序" in text or "恶意后门" in text or "后门恶意软件" in text:
        return "恶意后门文件"
    if "木马下载器" in text:
        return "恶意木马下载器"
    if "恶意快捷方式文件" in text or ".LNK 文件" in text or ".lnk 文件" in text:
        return "恶意快捷方式文件"
    if "恶意下载文件" in text:
        return "恶意下载文件"
    if "恶意下载脚本" in text:
        return "恶意下载脚本"
    if "恶意木马脚本" in text or "木马脚本" in text:
        return "恶意木马脚本文件"
    if "恶意脚本" in text:
        return "恶意脚本文件"
    if "PowerShell 脚本" in text:
        return "恶意 PowerShell 脚本"
    if "VBScript" in text:
        return "恶意 VBScript 脚本"
    if "恶意 .NET 可执行文件" in text:
        return "恶意 .NET 可执行文件"
    if ".NET 木马" in text:
        return "恶意 .NET 木马文件"
    if ".NET 程序" in text:
        return "恶意 .NET 程序"
    if "64 位 DLL" in text or "64 位动态链接库" in text:
        return "恶意 64 位 DLL 文件"
    if "压缩 ZIP 存档" in text or "压缩存档文件" in text or "压缩存档" in text:
        return "压缩存档文件"
    if "服务器端脚本" in text:
        return "Web Shell文件"
    return ""


def has_specific_transfer_name(parts: List[str]) -> bool:
    generic_parts = {
        "下载",
        "Web Shell文件",
        "恶意配置脚本文件",
        "恶意混淆脚本文件",
        "恶意批处理脚本文件",
        "恶意配置文件",
        "恶意 JavaScript 脚本",
        "恶意 64 位可执行文件",
        "恶意可执行文件",
        "恶意移动应用程序",
        "恶意远程访问工具",
        "恶意远程访问木马文件",
        "恶意远程访问可执行文件",
        "恶意后门木马文件",
        "恶意后门文件",
        "恶意木马下载器",
        "恶意快捷方式文件",
        "恶意下载文件",
        "恶意下载脚本",
        "恶意木马脚本文件",
        "恶意 JavaScript 下载脚本",
        "恶意 JavaScript 下载器",
        "恶意脚本文件",
        "恶意 PowerShell 脚本",
        "恶意 VBScript 脚本",
        ".NET 可执行文件",
        "恶意 .NET 程序",
        "Python 脚本文件",
        "32 位 Windows .DLL文件",
        "恶意 Windows .DLL文件",
        "恶意动态链接库文件",
        "恶意软件组件文件",
        "木马化的软件组件文件",
        "混淆脚本文件",
        "恶意 64 位 DLL 文件",
        "恶意 32 位可执行文件",
        "压缩存档文件",
        "恶意 Windows Installer 程序包",
        "恶意安装程序包",
        "恶意安装程序",
    }
    for part in parts:
        if not part or part.startswith("APT-") or part.startswith("变种 #"):
            continue
        if part in generic_parts:
            continue
        if re.search(r"[A-Za-z]", part):
            return True
    return False


def should_add_inferred_file_type(parts: List[str], inferred: str) -> bool:
    if not inferred or inferred in parts:
        return False
    if any(re.search(r"\b[A-Za-z0-9_-]+\.[A-Za-z0-9]{2,8}\b", part) for part in parts):
        return False
    if not has_specific_transfer_name(parts):
        return True
    # Keep precise malware/container types from the description even when a family name is present.
    return inferred in {
        "恶意 .NET 木马文件",
        "恶意 .NET 可执行文件",
        "恶意 .NET 程序",
        "恶意配置脚本文件",
        "恶意混淆脚本文件",
        "恶意批处理脚本文件",
        "恶意配置文件",
        "恶意下载脚本",
        "恶意脚本文件",
        "恶意木马脚本文件",
        "恶意 JavaScript 脚本",
        "恶意 JavaScript 下载脚本",
        "恶意 JavaScript 下载器",
        "恶意远程访问工具",
        "恶意远程访问木马文件",
        "恶意远程访问可执行文件",
        "恶意后门木马文件",
        "恶意后门文件",
        "恶意木马下载器",
        "恶意 64 位可执行文件",
        "恶意 32 位可执行文件",
        "恶意可执行文件",
        "恶意安装程序",
        "恶意安装程序包",
        "恶意 Windows Installer 程序包",
        "压缩存档文件",
    }


def normalize_actor_name(name: str) -> str:
    value = normalize_common_text(name).strip(" ，,。")
    compact = re.sub(r"\s+", "", value)
    aliases = {
        "Bluenoroff": "BlueNoroff",
        "BLUENOROFF": "BlueNoroff",
        "Muddywater": "MuddyWater",
        "MUDDYWATER": "MuddyWater",
    }
    return aliases.get(compact, value)


def format_file_extension_type(ext: str) -> str:
    clean = ext.strip().strip(".").upper()
    if not clean:
        return ""
    return f".{clean} 文件"


def normalize_malicious_transfer_item(part: str) -> str:
    clean = normalize_common_text(part).strip()
    clean = clean.replace("恶意软件释放器", "释放器")
    clean = clean.replace("后门恶意软件", "后门")
    clean = re.sub(r"(.+?后门)恶意软件$", r"\1", clean)
    return clean


def expand_malicious_transfer_desc_part(part: str) -> str:
    clean = normalize_common_text(part).strip()
    if clean == "释放器":
        return "恶意软件释放器"
    if clean == "后门" or re.search(r"后门$", clean):
        return f"{clean}恶意软件"
    return clean


def normalize_file_extension_part(part: str) -> str:
    clean = normalize_common_text(part).strip()
    match = re.fullmatch(r"\.?\s*([A-Za-z0-9]{2,8})\s*文件", clean, re.IGNORECASE)
    if not match:
        return normalize_malicious_transfer_item(clean)
    return format_file_extension_type(match.group(1))


def split_download_body(body: str) -> Tuple[str, str]:
    clean = normalize_common_text(body)
    clean = clean.replace("。.", ".").replace("。", "，").strip(" ，,。")
    match = re.search(r"[，,\s]*\.?\s*([A-Za-z0-9]{2,8})\s*文件\s*$", clean, re.IGNORECASE)
    if not match:
        return normalize_malicious_transfer_item(clean), ""
    item = clean[: match.start()].strip(" ，,。.")
    return normalize_malicious_transfer_item(item), format_file_extension_type(match.group(1))


def titleize_raw_malicious_download(name: str) -> str:
    raw = normalize_variant(normalize_cn_action_terms(name))
    variant_match = re.search(r"变种\s*#\d+", raw)
    variant = variant_match.group(0) if variant_match else ""
    raw = re.sub(r"\s*变种\s*#\d+\s*[。.]?\s*$", "", raw).strip(" ，,。")

    match = re.match(r"^(?P<item>.+?)[，,]\s*由\s*(?P<actor>.+?)\s*威胁组织使用[，,。.]\s*(?P<body>.+?)下载$", raw)
    if match:
        actor = normalize_actor_name(match.group("actor"))
        item = normalize_malicious_transfer_item(match.group("item"))
        _, file_type = split_download_body(match.group("body"))
    else:
        match = re.match(r"^(?P<item>.+?)\s*由\s*(?P<actor>.+?)\s*威胁组织使用[，,。.]\s*(?P<body>.+?)下载$", raw)
        if match:
            actor = normalize_actor_name(match.group("actor"))
            item = normalize_malicious_transfer_item(match.group("item"))
            _, file_type = split_download_body(match.group("body"))
        else:
            match = re.match(r"^(?P<item>.+?)\s*是\s*(?P<actor>.+?)\s*威胁组织使用的(?P<body>.+?)下载$", raw)
            if match:
                actor = normalize_actor_name(match.group("actor"))
                body, file_type = split_download_body(match.group("body"))
                item = f"{match.group('item').strip()} {body}".strip()
            else:
                match = re.match(r"^(?P<actor>.+?)威胁组织使用的(?P<body>.+?)下载$", raw)
                if match:
                    actor = normalize_actor_name(match.group("actor"))
                    item, file_type = split_download_body(match.group("body"))
                else:
                    match = re.match(r"^(?P<campaign>.+?供应链攻击活动)中使用的(?P<body>.+?)下载$", raw)
                    if match:
                        actor = normalize_common_text(match.group("campaign")).strip()
                        item, file_type = split_download_body(match.group("body"))
                    else:
                        match = re.match(r"^(?P<body>.+?文件)下载$", raw)
                        if not match:
                            return ""
                        actor = ""
                        item, file_type = split_download_body(match.group("body"))

    parts = [actor] if actor else []
    if item:
        parts.append(normalize_malicious_transfer_item(item))
    if file_type:
        parts.append(file_type)
    parts.append("下载")
    if variant:
        parts.append(variant)
    return "恶意文件传输 - " + "，".join(part for part in parts if part)


def titleize_malicious_transfer(name: str, desc: str) -> str:
    raw = normalize_cn_action_terms(name)
    raw = normalize_variant(raw)
    raw = raw.replace("恶意文件传输 - ", "")
    parts = [normalize_file_extension_part(part) for part in re.split(r"[，,]", raw) if part.strip()]
    if parts and not parts[0].startswith("APT"):
        parts.insert(0, parts.pop(0))
    inferred = infer_file_type(name, desc)
    if should_add_inferred_file_type(parts, inferred):
        insert_at = parts.index("下载") if "下载" in parts else min(len(parts), 2)
        parts.insert(insert_at, inferred)
    if "下载" not in parts:
        if parts:
            parts.insert(min(len(parts), 2), "下载")
        else:
            parts.append("下载")
    return "恶意文件传输 - " + "，".join(parts)


def _is_transfer_desc_object(part: str) -> bool:
    value = normalize_common_text(part).strip()
    if not value:
        return False
    if value == "下载" or value.startswith("变种 #") or value.startswith("APT-"):
        return False
    file_type_markers = (
        "文件",
        "脚本",
        "程序",
        "工具",
        "木马",
        "后门",
        "下载器",
        "释放器",
        "组件",
        "存档",
        "安装程序",
        "程序包",
        "可执行",
        "DLL",
    )
    if any(marker in value for marker in file_type_markers):
        return False
    return True


def preferred_transfer_association_phrase(title: str) -> str:
    subject = title.replace("恶意文件传输 - ", "")
    parts = [part.strip() for part in subject.split("，") if part.strip()]
    actor = next((part for part in parts if re.fullmatch(r"APT-U\d+", part)), "")
    objects = [part for part in parts if _is_transfer_desc_object(part)]
    if not actor:
        return ""
    if not objects:
        return f"与 {actor} 关联的文件"
    formatted = [f"{objects[0]} ({actor})", *objects[1:]]
    if len(formatted) == 1:
        object_text = formatted[0]
    elif len(formatted) == 2:
        object_text = f"{formatted[0]} 和 {formatted[1]}"
    else:
        object_text = f"{'，'.join(formatted[:-1])} 和 {formatted[-1]}"
    return f"与 {object_text} 关联的文件"


def normalize_transfer_association_opening(text: str, title: str) -> str:
    preferred = preferred_transfer_association_phrase(title)
    if not preferred:
        return text
    value = normalize_common_text(text)
    value = re.sub(r"（(APT-U\d+)）", r"(\1)", value)
    return re.sub(
        r"^(此验证动作还原了主机尝试下载)与\s*.*?\s*关联的文件。",
        rf"\1{preferred}。",
        value,
        count=1,
    )


def standardize_malicious_transfer_desc(title: str, desc: str) -> str:
    text = standardize_generic_desc(desc)
    text = normalize_transfer_association_opening(text, title)
    subject = title.replace("恶意文件传输 - ", "")
    parts = [part.strip() for part in subject.split("，") if part.strip()]
    detail_parts = [part for part in parts if part not in {"下载"} and not part.startswith("变种 #")]
    associated_name = ""
    if detail_parts and (
        detail_parts[0].startswith("APT-")
        or re.fullmatch(r"[A-Z][A-Z0-9_.-]{2,}", detail_parts[0])
    ):
        associated_name = detail_parts[0]
    if detail_parts and any(token in detail_parts[0] for token in ("勒索软件", "恶意软件", "木马", "后门", "加载器")):
        target = "，".join(detail_parts)
    else:
        target = "，".join(detail_parts[1:]) if len(detail_parts) > 1 else "相关文件"
    target = "，".join(expand_malicious_transfer_desc_part(part) for part in target.split("，") if part)
    if associated_name and target and target != "相关文件":
        download_target = f"与 {associated_name} 关联的{target}"
    else:
        download_target = f" {target}" if re.match(r"[A-Za-z0-9]", target) else target
    opening = f"此验证动作还原了主机尝试下载{download_target}。"
    if text.startswith("此验证动作还原了"):
        return normalize_common_text(text)
    return normalize_common_text(f"{opening} {text}".strip())


def extract_releaser_name(desc: str) -> str:
    text = normalize_cn_action_terms(desc)
    patterns = [
        r"还原了\s*([A-Z][A-Z0-9_.-]{2,})\s*释放器的执行",
        r"还原了\s*([A-Z][A-Z0-9_.-]{2,})\s+JavaScript 下载器的初始执行",
        r"\[\*\*([A-Z][A-Z0-9_.-]{2,})\*\*\]\s*-\s*\1\s*是[^。]*?释放器",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            candidate = match.group(1).upper()
            if CAMPAIGN_CODE_RE.fullmatch(candidate):
                continue
            return candidate
    return ""


def _is_stable_named_object(candidate: str) -> bool:
    value = normalize_common_text(candidate).strip(" “”。,，")
    if not value:
        return False
    if CAMPAIGN_CODE_RE.fullmatch(value) or value.startswith("APT-"):
        return False
    if value.lower() in {"windows", "linux", "macos", "dns", "c2", "c&c"}:
        return False
    return bool(re.search(r"[A-Za-z0-9]", value))


def extract_primary_malware_name(desc: str) -> str:
    text = normalize_cn_action_terms(desc)
    patterns = [
        r"\[\*\*([A-Z][A-Z0-9_.-]{2,})\*\*\]\s*-\s*\1\s+是",
        r"(?:运行的|执行)\s+([A-Z][A-Z0-9_.-]{2,})\s+恶意软件",
        r"([A-Z][A-Z0-9_.-]{2,})\s+恶意软件执行",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            candidate = match.group(1)
            if _is_stable_named_object(candidate):
                return candidate.upper() if candidate.isupper() else candidate
    return ""


def extract_named_sandbox_object(desc: str) -> str:
    text = normalize_cn_action_terms(desc)
    patterns = [
        r"名为[“\"]([^”\"]{2,80})[”\"]的(?:任务|服务|Windows 服务|计划任务)",
        r"(?:任务名|服务名)为[“\"]([^”\"]{2,80})[”\"]",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text):
            candidate = normalize_common_text(match.group(1)).strip()
            if _is_stable_named_object(candidate):
                return candidate
    return ""


def titleize_sandbox(name: str, desc: str) -> str:
    raw = normalize_cn_action_terms(name)
    raw = raw.replace("受保护的剧场", "受保护的沙盘").replace("受保护剧场", "受保护的沙盘")
    raw = raw.replace(" Execution", "，执行")
    raw = raw.replace("C&C域信标", "C&C 域名信标")
    raw = raw.replace("C&C 域信标", "C&C 域名信标")
    raw = re.sub(r"\bDns\b", "DNS", raw, flags=re.IGNORECASE)
    raw = raw.replace("舞台脚本释放器", "Stage Script释放器")
    raw = raw.replace("释放器 执行", "释放器，执行")
    raw = raw.replace("Execute 释放器", "释放器，执行")
    raw = raw.replace("执行 释放器", "释放器，执行")
    raw = raw.replace("执行释放器", "释放器，执行")
    raw = raw.replace(" Javascript ", " JavaScript ")
    raw = raw.replace("Powershell", "PowerShell")
    raw = normalize_variant(raw)
    prefix, rest = raw.split(" - ", 1) if " - " in raw else ("受保护的沙盘", raw)
    raw_parts = [part.strip() for part in re.split(r"[，,]", rest) if part.strip()]
    parts: List[str] = []
    for part in raw_parts:
        split_match = re.match(
            r"^([A-Za-z0-9_.-]+\.(?:exe|dll|sys|ps1|vbs|js|jar|sh|elf|bin))\s+(.+)$",
            part,
            flags=re.IGNORECASE,
        )
        if split_match and any(marker in split_match.group(2) for marker in ("C&C", "信标", "渗透", "泄露", "下载", "执行")):
            parts.extend([split_match.group(1), split_match.group(2).strip()])
        else:
            parts.append(part)
    releaser_name = extract_releaser_name(desc)
    primary_malware_name = extract_primary_malware_name(desc)
    named_object = extract_named_sandbox_object(desc)

    if releaser_name:
        if len(parts) > 1 and re.fullmatch(r"[\u4e00-\u9fff]+", parts[1]):
            parts[1] = releaser_name
        elif releaser_name not in parts:
            parts.insert(1 if parts and parts[0].startswith("APT-") else 0, releaser_name)

    if primary_malware_name:
        for idx, part in enumerate(parts):
            if re.fullmatch(r"[\u4e00-\u9fff]+", part) and not any(
                marker in part
                for marker in (
                    "执行",
                    "下载",
                    "投放",
                    "持久化",
                    "释放",
                    "通信",
                    "连接",
                    "数据",
                    "聚合",
                    "泄露",
                    "渗透",
                    "信标",
                )
            ):
                parts[idx] = primary_malware_name
                break

    if named_object and named_object not in parts:
        insert_idx = 1 if parts and parts[0].startswith("APT-") else 0
        while insert_idx < len(parts) and not any(
            marker in parts[insert_idx]
            for marker in ("执行", "下载", "投放", "持久化", "释放", "通信", "连接", "信标", "渗透", "泄露")
        ):
            insert_idx += 1
        parts.insert(insert_idx, named_object)

    new_parts: List[str] = []
    for part in parts:
        clean_part = part.replace(" Dll", "").replace(" DLL", "").strip()
        if "释放器" in clean_part:
            before, _, after = clean_part.partition("释放器")
            before = before.replace("执行", "").strip()
            after = after.replace("执行", "").strip()
            if before:
                separator = "" if re.search(r"[\u4e00-\u9fff]$", before) else " "
                releaser_type = f"{before}{separator}释放器"
            else:
                releaser_type = "释放器"
            if releaser_type not in new_parts:
                new_parts.append(releaser_type)
            if "执行" not in new_parts:
                new_parts.append("执行")
            if after:
                new_parts.append(after)
            continue
        new_parts.append(clean_part)

    deduped: List[str] = []
    for part in new_parts:
        if not part:
            continue
        if part in deduped:
            continue
        deduped.append(part)
    result = f"{prefix} - " + "，".join(deduped)
    result = result.replace("，执行，执行", "，执行，")
    result = result.replace("，执行投放", "，执行，投放")
    result = result.replace("Stage Script 释放器", "Stage Script释放器")
    return result


def titleize_c2(name: str) -> str:
    raw = normalize_cn_action_terms(name)
    raw = raw.replace("电报通信", "Telegram 通信")
    raw = raw.replace("，电报，", "，Telegram，")
    raw = raw.replace("，任务，", "，任务流量，")
    raw = raw.replace("，任务，", "，任务流量，")
    raw = raw.replace("，签到，", "，签入，")
    raw = raw.replace("，外泄，", "，泄漏，")
    raw = raw.replace("，利用，", "，利用请求，")
    return raw


def titleize_c2_with_context(name: str, desc: str) -> str:
    raw = titleize_c2(name)
    if "FRONTLOAD" in desc:
        raw = raw.replace("前置式", "FRONTLOAD")
    for path in extract_uri_paths(desc)[:1]:
        raw = insert_title_part_before_action(raw, path)
    return raw


def titleize_phishing_email(name: str) -> str:
    return normalize_variant(normalize_cn_action_terms(name))


def titleize_host_cmd(name: str, desc: str = "") -> str:
    raw = normalize_variant(normalize_cn_action_terms(name))
    raw = re.sub(
        r"^(主机命令行\s*-\s*使用\s+Wevtutil\s+工具)\s+变种\s*#(\d+)\s+(清除系统事件日志)(?:\s*\((Windows|Linux|macOS)\))?$",
        lambda m: f"{m.group(1)}{m.group(3)}" + (f" ({m.group(4)})" if m.group(4) else "") + f"，变种 #{m.group(2)}",
        raw,
        flags=re.IGNORECASE,
    )
    if "4096 字节随机写入" in raw and "Handala 威胁组织" not in raw and "Handala" in desc:
        raw = raw.replace("主机命令行 - ", "主机命令行 - Handala 威胁组织，", 1)
    patterns = [
        (
            r"^主机命令行\s*-\s*使用[“\"]([^”\"]+)[”\"]命令显示可配置服务的列表(?:\s*\([^)]*\))?$",
            "主机命令行 - {cmd}，可配置服务列表显示",
        ),
        (
            r"^主机命令行\s*-\s*使用[“\"]([^”\"]+)[”\"]命令显示应用程序和服务列表(?:\s*\([^)]*\))?$",
            "主机命令行 - {cmd}，应用程序和服务列表显示",
        ),
    ]
    for pattern, template in patterns:
        match = re.match(pattern, raw)
        if match:
            return template.format(cmd=match.group(1).strip())
    return raw


def standardize_host_cmd_desc(name: str, desc: str) -> str:
    clean_name = normalize_common_text(name)
    clean_desc = normalize_common_text(desc)
    lower_name = clean_name.lower()
    if "net config" in lower_name:
        return "此验证动作还原了在 Windows 主机上执行 Net config 命令以显示正在运行的可配置服务列表的行为。该命令也可用于显示和更改服务器服务或工作站服务的设置。"
    if "tasklist /svc" in lower_name:
        return "此验证动作还原了在 Windows 主机上执行 tasklist /svc 命令以显示本地或远程计算机上正在运行的任务、对应服务及进程 ID (PID) 的行为。"
    return standardize_generic_desc(clean_desc)


def standardize_generic_desc(desc: str) -> str:
    body, urls = split_references(desc)
    text = cleanup_unwanted_attribution(normalize_geo_company_text(body))
    text = re.sub(r"^在此验证动作中，", "此验证动作还原了", text)
    text = re.sub(r"^在此攻击中，", "此验证动作还原了", text)
    text = re.sub(r"^在此过程中，", "此验证动作还原了", text)
    text = re.sub(r"^在此行为中，", "此验证动作还原了", text)
    text = re.sub(r"^在执行中，", "此验证动作还原了", text)
    text = re.sub(r"^在此执行中，", "此验证动作还原了", text)
    text = re.sub(r"^此次攻击包括", "此验证动作还原了", text)
    text = re.sub(r"^此攻击包括", "此验证动作还原了", text)
    text = re.sub(r"^此验证动作展示了", "此验证动作还原了", text)
    text = re.sub(r"^此验证动作模拟了", "此验证动作还原了", text)
    text = re.sub(r"^此验证动作表明主机正在尝试", "此验证动作还原了主机尝试", text)
    text = re.sub(r"^此验证动作表明主机试图", "此验证动作还原了主机尝试", text)
    text = re.sub(r"^此验证动作表明攻击者正在尝试", "此验证动作还原了攻击者尝试", text)
    text = re.sub(r"^此验证动作表明攻击者试图", "此验证动作还原了攻击者尝试", text)
    text = re.sub(r"^此验证动作表明", "此验证动作还原了", text)
    text = re.sub(r"^此验证动作通过", "此验证动作还原了通过", text)
    text = re.sub(r"^此验证动作在", "此验证动作还原了在", text)
    text = re.sub(r"^此验证动作使用", "此验证动作还原了使用", text)
    text = re.sub(r"^此验证动作尝试", "此验证动作还原了尝试", text)
    text = re.sub(r"^此验证动作会", "此验证动作还原了", text)
    text = re.sub(r"^此验证动作还原了下载", "此验证动作还原了主机尝试下载", text)
    text = re.sub(r"^该验证动作通过", "此验证动作还原了通过", text)
    text = re.sub(r"^该验证动作在", "此验证动作还原了在", text)
    text = re.sub(r"^该验证动作使用", "此验证动作还原了使用", text)
    text = re.sub(r"^该验证动作尝试", "此验证动作还原了尝试", text)
    text = re.sub(r"^该验证动作会", "此验证动作还原了", text)
    text = re.sub(r"^该恶意软件脚本使用", "此验证动作还原了恶意软件脚本使用", text)
    text = re.sub(r"^该恶意软件脚本", "此验证动作还原了恶意软件脚本", text)
    text = re.sub(r"^该脚本通过", "此验证动作还原了该脚本通过", text)
    text = re.sub(r"^该脚本", "此验证动作还原了该脚本", text)
    text = re.sub(r"^此脚本", "此验证动作还原了该脚本", text)
    text = re.sub(r"^这款脚本", "此验证动作还原了该脚本", text)
    text = re.sub(r"^此验证动作还原了攻击者会", "此验证动作还原了攻击者", text)
    text = re.sub(r"^此验证动作还原了主机会", "此验证动作还原了主机", text)
    text = re.sub(r"^此验证动作还原如何", "此验证动作还原了如何", text)
    text = text.replace("。- ", "。")
    text = text.replace("；- ", "；")
    text = text.replace("：- ", "：")
    text = text.replace("。 - ", "。")
    text = text.replace("； - ", "；")
    text = text.replace("： - ", "：")
    text = normalize_common_text(text.strip())
    text = re.sub(r"([。！？])\s+(?=[A-Za-z0-9\u4e00-\u9fff])", r"\1", text)
    text = text.replace("Shellcode 注入工具 传播", "Shellcode 注入工具，传播")
    return ensure_terminal_punctuation(text) + build_reference_block(urls)


def normalize_actor_spacing(text: str) -> str:
    value = text
    value = re.sub(r"(APT-[A-Z]*U?\d+)(在攻击活动)", r"\1 在攻击活动", value)
    value = re.sub(r"APT-APT-", "APT-", value)
    return value


def derive_sequence_subject(name: str) -> str:
    raw = normalize_cn_action_terms(name)
    raw = normalize_variant(raw)
    raw = raw.replace("恶意软件场景 - ", "").replace("恶意活动场景 - ", "")
    raw = raw.replace("APT-APT-", "APT-")
    variant = ""
    variant_match = re.search(r"(?:变种\s*#(\d+)|#(\d+)|-\s*(\d+))\s*$", raw)
    if variant_match:
        variant_number = next(group for group in variant_match.groups() if group)
        variant = f" #{variant_number}"
        raw = raw[: variant_match.start()].strip(" ，,-")
    raw = raw.replace("威胁组织发起的", " ")
    raw = raw.replace("恶意软件下载威胁活动", "恶意软件下载攻击活动")
    raw = raw.replace("恶意软件下载威胁", "恶意软件下载攻击活动")
    raw = raw.replace("软件下载威胁", "软件下载攻击活动")
    raw = raw.replace("下载威胁", "下载攻击活动")
    raw = raw.replace("威胁组织活动恶意软件下载", "威胁组织恶意软件下载")
    raw = raw.replace("供应链攻击活动恶意软件下载攻击活动", "供应链攻击活动恶意软件下载")
    raw = raw.replace("供应链攻击活动恶意软件下载威胁", "供应链攻击活动恶意软件下载")
    raw = re.sub(r"(APT-U\d+)\s*攻击活动", r"\1 攻击活动", raw)
    raw = re.sub(r"\s*攻击活动\s*$", "攻击活动", raw)
    raw = re.sub(r"(APT-U\d+)攻击活动", r"\1 攻击活动", raw)
    raw = re.sub(r"(威胁组织|勒索软件|病毒|恶意软件)活动\s*攻击活动$", r"\1攻击活动", raw)
    raw = re.sub(r"(威胁组织|勒索软件|病毒|恶意软件)活动$", r"\1攻击活动", raw)
    raw = re.sub(r"攻击活动\s*攻击活动$", "攻击活动", raw)
    raw = raw.replace("恶意软件 攻击活动", "恶意软件攻击活动")
    raw = raw.replace("勒索软件 攻击活动", "勒索软件攻击活动")
    if variant and not re.search(r"\s#\d+$", raw):
        raw = f"{raw}{variant}"
    return normalize_common_text(raw.strip())


def standardize_sequence_name(name: str) -> str:
    subject = derive_sequence_subject(name)
    if not subject:
        subject = "验证对象"
    if "攻击活动" in subject:
        return f"恶意活动场景 - {subject}"
    return f"恶意活动场景 - {subject} 攻击活动"


def sequence_desc_subject(subject: str) -> str:
    clean = normalize_common_text(subject)
    clean = re.sub(r"，?变种\s*#\d+\s*$", "", clean)
    clean = re.sub(r"\s*攻击活动\s*$", "", clean).strip()
    clean = clean.replace("攻击活动恶意软件下载", "攻击活动中的恶意软件下载")
    return clean or "验证对象"


def join_sequence_subject_clause(subject: str, clause: str) -> str:
    if not subject:
        return clause
    if re.search(r"[A-Za-z0-9)]$", subject):
        return f"{subject} {clause}"
    return f"{subject}{clause}"


def join_sequence_prefix(prefix: str, text: str) -> str:
    if not text:
        return prefix
    if re.match(r"[A-Za-z0-9]", text):
        return f"{prefix} {text}"
    return f"{prefix}{text}"


def is_sequence_download_subject(subject: str) -> bool:
    return any(token in subject for token in ("恶意文件传输", "恶意样本传输", "恶意软件下载", "软件下载", "文件下载"))


def is_sequence_sample_download_subject(subject: str) -> bool:
    return any(token in subject for token in ("恶意软件下载", "恶意软件", "勒索软件下载", "加载器下载", "后门下载", "释放器下载"))


def sequence_sample_download_intro(desc_subject: str) -> str:
    subject = normalize_common_text(desc_subject)
    if "供应链攻击活动" in subject and "恶意软件下载" in subject:
        campaign = re.sub(r"(?:中的)?恶意软件下载.*$", "", subject).strip()
        target = f"{join_sequence_prefix('与', campaign)}中使用过的各种变种的下载"
        return f"{join_sequence_prefix('此验证场景包括了', target)}。"
    if "勒索软件" in subject:
        target_name = subject
        target_name = re.sub(r"下载.*$", "", target_name).strip()
        target_name = re.sub(r"攻击活动$", "", target_name).strip()
        target = join_sequence_subject_clause(join_sequence_prefix("与", target_name), "相关的各种变种的下载")
        return f"{join_sequence_prefix('此验证场景包括了', target)}。"
    target_name = subject
    with_prefix = any(token in target_name for token in ("恶意软件下载", "下载"))
    target_name = re.sub(r"恶意软件下载.*$", "", target_name).strip()
    target_name = re.sub(r"下载.*$", "", target_name).strip()
    if not target_name:
        target_name = re.sub(r"恶意软件.*$", "", subject).strip()
    target_name = re.sub(r"攻击活动$", "", target_name).strip()
    relation_subject = join_sequence_prefix("与", target_name) if with_prefix else target_name
    target = join_sequence_subject_clause(relation_subject, "相关的各种变种的下载")
    return f"{join_sequence_prefix('此验证场景包括了', target)}。"


def sequence_desc_intro(desc_subject: str) -> str:
    if is_sequence_sample_download_subject(desc_subject):
        return sequence_sample_download_intro(desc_subject)
    if is_sequence_download_subject(desc_subject):
        clause = f"{join_sequence_prefix('与', desc_subject)}相关的攻击手法"
    else:
        clause = join_sequence_subject_clause(desc_subject, "在攻击活动中使用过的相关攻击手法")
    return f"{join_sequence_prefix('此验证场景包括了', clause)}。"


def should_drop_sequence_original_body(desc_subject: str, clean: str) -> bool:
    return bool(is_sequence_sample_download_subject(desc_subject) and re.match(r"^(?:这种威胁|该场景)包括下载", clean))


def normalize_sequence_desc_text(text: str) -> str:
    text = re.sub(r"\s+", " ", text).strip()

    def replace_download_intro(match: re.Match[str]) -> str:
        subject = normalize_common_text(match.group(1).strip(" ，,"))
        return sequence_desc_intro(subject)

    def replace_single_variant(match: re.Match[str]) -> str:
        target = normalize_common_text(match.group(1).strip())
        download_target = join_sequence_prefix("下载", target)
        if target.endswith("攻击"):
            return f"该场景包括{download_target}活动相关变种的行为。"
        if target.endswith("活动"):
            return f"该场景包括{download_target}相关变种的行为。"
        return f"该场景包括{download_target}变种的行为。"

    text = re.sub(
        r"此验证场景包括了\s*([^。]{1,80}?下载)\s*在攻击活动中使用过的相关攻击手法。",
        replace_download_intro,
        text,
    )
    text = re.sub(
        r"此验证场景包括了\s*([^。]{1,80}?下载)相关攻击手法。",
        replace_download_intro,
        text,
    )
    text = re.sub(r"此验证场景包括了\s+([\u4e00-\u9fff])", r"此验证场景包括了\1", text)
    text = text.replace("这种威胁包括", "该场景包括")
    text = re.sub(r"该场景包括下载([^。]+?)的变种。", replace_single_variant, text)
    text = re.sub(r"该场景包括下载([^。]+?)的各种变种。", r"该场景包括下载\1各类变种的行为。", text)
    return text


def recover_sequence_extra_sentences(body: str, clean: str) -> str:
    if len(split_sentences(clean)) > 1:
        return clean
    raw = normalize_actor_spacing(normalize_geo_company_text(normalize_cn_action_terms(body)))
    extras: List[str] = []
    for sentence in split_sentences(raw)[1:]:
        if sentence in clean:
            continue
        if any(marker in sentence for marker in ("[**", "CAMP.", "FireEye", "与中国有关联", "中国相关", "中国支持", "分发集群")):
            continue
        recovered = sentence.replace("犯罪团伙", "威胁组织")
        recovered = recovered.replace("他们散布", "其分发").replace("他们分发", "其分发")
        extras.append(recovered)
    if not extras:
        return clean
    return clean + "".join(extras)


def standardize_pipeline_name(name: str) -> str:
    return normalize_common_text(name)


def standardize_pipeline_desc(desc: str) -> str:
    text = normalize_geo_company_text(normalize_cn_action_terms(desc))
    text = re.sub(r"参考链接\s*:", "参考链接：", text, flags=re.IGNORECASE)
    text = re.sub(r"([。！？])\s*", r"\1", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def standardize_sequence_desc(name: str, desc: str) -> str:
    subject = derive_sequence_subject(name)
    desc_subject = sequence_desc_subject(subject)
    body, urls = split_references(desc)
    clean = normalize_actor_spacing(cleanup_unwanted_attribution(normalize_geo_company_text(normalize_cn_action_terms(body))))
    clean = recover_sequence_extra_sentences(body, clean)
    if clean.startswith("此验证场景包括了"):
        text = clean
    elif clean:
        if should_drop_sequence_original_body(desc_subject, clean):
            text = sequence_desc_intro(desc_subject)
        else:
            text = f"{sequence_desc_intro(desc_subject)}{clean}"
    else:
        text = sequence_desc_intro(desc_subject)
    text = normalize_sequence_desc_text(text)
    text = re.sub(r"(APT-[A-Z]*U?\d+)(在)", r"\1 在", text)
    text = re.sub(r"(FIN\d+)(在)", r"\1 在", text)
    text = re.sub(r"(恶意软件|勒索软件|木马|后门)\s+在攻击活动中", r"\1在攻击活动中", text)
    text = re.sub(r"\bAPT\s*U\s*(\d+)\b", r"APT-U\1", text)
    text = re.sub(r"包括了\s*APT[\s-]*U\s*(\d+)", r"包括了 APT-U\1", text)
    text = text.replace("包括了 ", "包括了 ")
    text = normalize_common_text(text.strip())
    return ensure_terminal_punctuation(text) + build_reference_block(urls)


def infer_email_subject(body: str) -> str:
    text = normalize_common_text(body)
    lower = text.lower()
    if "filemail.com" in lower:
        return "最新油价提醒"
    if "docbridge" in lower or "openclaw" in lower:
        return "OpenClaw使用攻略"
    if "githubusercontent.com" in lower or "cv%20new.zip" in lower or "cv new.zip" in lower:
        return "团队 Token 购买"
    if "github.com" in lower:
        return "LLM可用清单"
    if "附件" in text and "报告" in text:
        return "最新行业分析报告"
    if "附件" in text:
        return "请查收附件资料"
    if "报告" in text:
        return "最新研究报告"
    if ".zip" in lower or ".rar" in lower or ".7z" in lower:
        return "最新资料下载"
    return "测试邮件"


def standardize_email_row(subject: str, body: str) -> Tuple[str, str]:
    clean_body = normalize_common_text(body)
    return subject, clean_body


def standardize_actions_row(name: str, desc: str, notes: str, context_text: str = "") -> Tuple[str, str, str]:
    clean_name = normalize_common_text(name)
    clean_desc = normalize_common_text(desc)
    clean_context = normalize_common_text(context_text)
    clean_notes = normalize_notes(notes)
    row_aliases = infer_contextual_aliases_from_action(clean_name, clean_desc)
    if row_aliases:
        clean_name = apply_contextual_aliases(clean_name, row_aliases)
        clean_desc = apply_contextual_aliases(clean_desc, row_aliases)

    if re.match(r"(?i)^web\s*安全验证\s*-\s*", clean_name):
        return (
            normalize_web_name(clean_name, clean_desc),
            standardize_web_desc(clean_name, clean_desc),
            clean_notes or WEB_NOTE_DEFAULT,
        )
    if clean_name.startswith("恶意文件传输 - "):
        title = titleize_malicious_transfer(clean_name, clean_desc)
        return append_os_suffix(title, clean_notes), standardize_malicious_transfer_desc(title, clean_desc), clean_notes
    raw_download_title = titleize_raw_malicious_download(clean_name)
    if raw_download_title:
        return append_os_suffix(raw_download_title, clean_notes), standardize_malicious_transfer_desc(raw_download_title, clean_desc), clean_notes
    raw_vuln_transfer_title = standardize_raw_vulnerability_transfer_name(clean_name, clean_desc, clean_notes)
    if raw_vuln_transfer_title:
        return (
            append_os_suffix(raw_vuln_transfer_title, clean_notes),
            standardize_malicious_transfer_desc(raw_vuln_transfer_title, clean_desc),
            clean_notes,
        )
    raw_vuln_title = standardize_raw_vulnerability_name(clean_name, clean_desc, clean_notes)
    if raw_vuln_title:
        return append_os_suffix(raw_vuln_title, clean_notes), standardize_generic_desc(clean_desc), clean_notes
    if clean_name.startswith("受保护的沙盘") or clean_name.startswith("受保护剧场") or clean_name.startswith("受保护的剧场"):
        title = titleize_sandbox(clean_name, clean_desc)
        return append_os_suffix(title, clean_notes), standardize_generic_desc(clean_desc), clean_notes
    if clean_name.startswith("命令与控制 - "):
        title = titleize_c2_with_context(clean_name, clean_desc)
        return append_os_suffix(title, clean_notes), standardize_generic_desc(clean_desc), clean_notes
    if clean_name.startswith("主机命令行 - "):
        title = titleize_host_cmd(clean_name, clean_desc)
        return append_os_suffix(title, clean_notes), standardize_host_cmd_desc(clean_name, clean_desc), clean_notes
    if clean_name.startswith("钓鱼邮件 - "):
        title = titleize_phishing_email(clean_name)
        return title, standardize_generic_desc(clean_desc), clean_notes
    return clean_name, standardize_generic_desc(clean_desc), clean_notes


def standardize_sequences_row(name: str, desc: str, aliases: Optional[Dict[str, str]] = None) -> Tuple[str, str]:
    if aliases:
        name = apply_contextual_aliases(name, aliases)
        desc = apply_contextual_aliases(desc, aliases)
    return standardize_sequence_name(name), standardize_sequence_desc(name, desc)


def apply_added_prefix_highlights(output_path: Path, added_prefix_cells: List[Dict[str, object]]) -> None:
    if not added_prefix_cells:
        return
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to mark added validation prefixes") from exc

    wb = load_workbook(output_path)
    fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for item in added_prefix_cells:
        sheet_name = str(item["sheet"])
        row_idx = int(item["row"])
        col_idx = int(item["column"])
        if sheet_name in wb.sheetnames:
            wb[sheet_name].cell(row=row_idx, column=col_idx).fill = fill
    wb.save(output_path)


def rewrite_xlsx_for_excel_compatibility(output_path: Path) -> None:
    """Force a normal openpyxl save so Excel can open the final workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for final Excel compatibility rewrite") from exc

    wb = load_workbook(output_path)
    # Touch all worksheets so XML parsing failures surface before delivery.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                _ = cell.value
    wb.save(output_path)


def header_index_from_sheet(root: ET.Element, shared_strings: List[str]) -> Dict[str, int]:
    header_index: Dict[str, int] = {}
    header_row = root.find(".//a:sheetData/a:row[@r='1']", NS)
    if header_row is None:
        return header_index
    for cell in header_row.findall("a:c", NS):
        ref = cell.attrib.get("r", "")
        if not ref:
            continue
        col_idx, _ = split_ref(ref)
        header = read_cell_value(cell, shared_strings).strip()
        if header:
            header_index[header] = col_idx
    return header_index


def preferred_col(header_index: Dict[str, int], *headers: str, fallback: int) -> int:
    for header in headers:
        if header in header_index:
            return header_index[header]
    return fallback


def standardize_workbook(input_path: Path, output_path: Path, report_path: Path) -> Dict[str, object]:
    archive = XlsxArchive(input_path)
    shared_strings = load_shared_strings(archive.files)
    sheet_targets = workbook_sheet_targets(archive.files)
    action_target = _sheet_target_for_name(archive.files, "Actions")
    email_target = _sheet_target_for_name(archive.files, "Email")
    action_names: List[str] = []
    action_contextual_aliases: Dict[str, str] = {}
    email_context_by_vid: Dict[str, str] = {}
    if action_target:
        for row in _rows_from_sheet(archive.files, action_target)[1:]:
            name = normalize_common_text(row.get("C", ""))
            desc = normalize_common_text(row.get("E", ""))
            if name:
                action_names.append(name)
            action_contextual_aliases.update(infer_contextual_aliases_from_action(name, desc))
    if email_target:
        for row in _rows_from_sheet(archive.files, email_target)[1:]:
            vid = normalize_common_text(row.get("B", ""))
            body = normalize_common_text(row.get("E", ""))
            if vid and body:
                email_context_by_vid[vid] = body
    known_ransomware_families = load_known_ransomware_families() | collect_inline_ransomware_families(action_names)

    summary: Dict[str, object] = {
        "input": str(input_path),
        "output": str(output_path),
        "sheets": {},
        "added_prefix_cells": [],
    }

    for sheet_name, target in sheet_targets:
        root = ET.fromstring(archive.files[target])
        header_index = header_index_from_sheet(root, shared_strings)
        changed = False
        sheet_summary = {
            "rows_processed": 0,
            "rows_updated": 0,
            "examples": [],
        }

        for row in root.findall(".//a:sheetData/a:row", NS):
            row_idx = int(row.attrib.get("r", "0"))
            if row_idx <= 1:
                continue

            cell_map = row_cells_by_col(row)
            name_col = preferred_col(header_index, "cn_name", fallback=3)
            desc_col = preferred_col(header_index, "cn_desc", fallback=5)
            note_col = preferred_col(header_index, "cn_notes", "notes", fallback=7)
            name = read_cell_value(cell_map.get(name_col), shared_strings).strip()
            desc = read_cell_value(cell_map.get(desc_col), shared_strings).strip()
            notes = read_cell_value(cell_map.get(note_col), shared_strings).strip()
            if not name:
                continue

            sheet_summary["rows_processed"] += 1

            if sheet_name == "Actions":
                vid = read_cell_value(cell_map.get(2), shared_strings).strip()
                new_name, new_desc, new_notes = standardize_actions_row(
                    name,
                    desc,
                    notes,
                    email_context_by_vid.get(vid, ""),
                )
                new_name = apply_ransomware_suffix(new_name, known_ransomware_families, new_desc)
                new_name = clean_campaign_codes_from_title(new_name)
                original_missing_prefix = not has_validation_title_prefix(name)
                added_prefix = ""
                if original_missing_prefix and has_validation_title_prefix(new_name):
                    added_prefix = new_name.split(" - ", 1)[0] + " - "
                updates = {
                    name_col: new_name,
                    desc_col: new_desc,
                    note_col: new_notes,
                }
            elif sheet_name == "Sequences":
                new_name, new_desc = standardize_sequences_row(name, desc, action_contextual_aliases)
                new_name = clean_campaign_codes_from_title(new_name)
                updates = {
                    name_col: new_name,
                    desc_col: new_desc,
                }
            elif sheet_name == "Email":
                subject_col = preferred_col(header_index, "cn_subject", fallback=3)
                body_col = preferred_col(header_index, "cn_body", fallback=5)
                subject = read_cell_value(cell_map.get(subject_col), shared_strings).strip()
                body = read_cell_value(cell_map.get(body_col), shared_strings).strip()
                new_subject, new_body = standardize_email_row(subject, body)
                updates = {
                    subject_col: new_subject,
                    body_col: new_body,
                }
            elif sheet_name == "Pipelines":
                new_name = standardize_pipeline_name(name)
                new_desc = standardize_pipeline_desc(desc)
                updates = {
                    name_col: new_name,
                    desc_col: new_desc,
                }
            else:
                continue

            row_changed = False
            for col_idx, new_value in updates.items():
                current = read_cell_value(cell_map.get(col_idx), shared_strings).strip()
                if normalize_text(current) == normalize_text(new_value):
                    continue
                cell = ensure_cell(row, col_idx, row_idx)
                set_inline_string(cell, new_value)
                row_changed = True

            if row_changed:
                changed = True
                sheet_summary["rows_updated"] += 1
                if sheet_name == "Actions" and added_prefix:
                    added_prefix_cells: List[dict] = summary["added_prefix_cells"]  # type: ignore[assignment]
                    added_prefix_cells.append(
                        {
                            "sheet": sheet_name,
                            "row": row_idx,
                            "column": name_col,
                            "vid": read_cell_value(cell_map.get(2), shared_strings).strip(),
                            "added_prefix": added_prefix,
                            "original_name": name,
                            "updated_name": updates[name_col],
                        }
                    )
                examples: List[dict] = sheet_summary["examples"]  # type: ignore[assignment]
                if len(examples) < 10:
                    examples.append(
                        {
                            "row": row_idx,
                            "vid": read_cell_value(cell_map.get(2), shared_strings).strip(),
                            "original_name": name,
                            "updated_name": updates.get(name_col, updates.get(3, "")),
                        }
                    )

        if changed:
            archive.files[target] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        summary["sheets"][sheet_name] = sheet_summary

    archive.write(output_path)
    apply_added_prefix_highlights(output_path, summary["added_prefix_cells"])  # type: ignore[arg-type]
    rewrite_xlsx_for_excel_compatibility(output_path)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Standardize validation main workbook by XML editing.")
    parser.add_argument("--input", required=True, type=Path, help="Path to the validation main workbook.")
    parser.add_argument("--output", required=True, type=Path, help="Output xlsx path.")
    parser.add_argument("--report", required=True, type=Path, help="Output json report path.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    summary = standardize_workbook(args.input, args.output, args.report)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
