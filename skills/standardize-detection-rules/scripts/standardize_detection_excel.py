#!/usr/bin/env python3
"""Standardize detection Excel name.1, desc, and notes fields."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


WEB_PREFIX = "Web应用程序漏洞 - "
APP_PREFIX = "应用程序漏洞 - "
AI_APP_PREFIX = "AI应用程序漏洞 - "
PREFIX = WEB_PREFIX
PATH_RE = re.compile(r"(?<![A-Za-z0-9_:/])/[A-Za-z0-9._~:?#\[\]@!$&'()*+,;=%/-]+")
ACTION_ENTRY_RE = re.compile(r"\b[A-Za-z_][A-Za-z0-9_]*\s+AJAX\s+Action\b", re.IGNORECASE)
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
HISTORICAL_VENDOR_URLS = {
    "用友 NC": "https://www.yonyou.com/",
    "用友NC": "https://www.yonyou.com/",
}
HARDWARE_PRODUCT_KEYWORDS = (
    "上网行为管理",
    "防火墙",
    "安全网关",
    "网关",
    "路由器",
    "交换机",
    "VPN",
    "负载均衡",
    "堡垒机",
    "网闸",
    "入侵防御",
    "入侵检测",
    "无线控制器",
    "PAN-OS",
    "GlobalProtect",
    "Palo Alto",
    # Industry/operation platforms with web endpoints should still use the generic application prefix.
    "Acrel EEMS",
    "电力运维平台",
)
AI_APPLICATION_PRODUCTS = (
    "9Router",
    "Blinko",
    "Crawl4AI",
    "Crawl4ai",
    "Gradio",
    "Langflow",
    "LMDeploy",
    "MiroFish",
    "Scramble",
    "Open WebUI",
    "MLflow",
    "NocoBase",
    "LiteLLM",
)
CITY_NAMES = [
    "北京", "上海", "广州", "深圳", "杭州", "南京", "苏州", "无锡", "常州",
    "宁波", "温州", "嘉兴", "绍兴", "金华", "台州", "天津", "重庆", "成都",
    "武汉", "西安", "郑州", "长沙", "合肥", "济南", "青岛", "福州", "厦门",
    "泉州", "东莞", "佛山", "中山", "珠海", "昆明", "南宁", "南昌", "石家庄",
    "沈阳", "大连", "长春", "哈尔滨", "海口", "贵阳", "乌鲁木齐", "呼和浩特",
    "兰州", "太原", "唐山", "烟台", "潍坊", "临沂", "徐州", "南通", "盐城"
]


def clean_text(value: Any) -> str:
    raw = str(value or "").replace("\u3000", " ").strip()
    text = "".join(
        unicodedata.normalize("NFKC", char)
        if ("KANGXI RADICAL" in unicodedata.name(char, "") or "CJK RADICAL" in unicodedata.name(char, ""))
        else char
        for char in raw
    )
    if not text or text.lower() == "nan":
        return ""
    return re.sub(r"\s+", " ", text).strip()


def normalize_product_key(text: str) -> str:
    text = clean_text(text).lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[\s,，/._-]+", "", text)
    return text


def normalize_product_name(product: str) -> str:
    normalized = clean_text(product)
    for city in CITY_NAMES:
        if normalized.startswith(city) and len(normalized) > len(city) + 1:
            normalized = normalized[len(city) :].strip()
            break
    return normalized


def split_sentences(text: str) -> List[str]:
    normalized = clean_text(text)
    normalized = normalized.replace("。", "。\n").replace("！", "！\n").replace("？", "？\n")
    return [item.strip() for item in normalized.splitlines() if item.strip()]


def split_disclosure_time(text: str) -> Tuple[str, str]:
    cleaned = clean_text(text)
    matched = re.search(r"披露时间[:：]\s*(\d{4}-\d{2}(?:-\d{2})?)\s*[.。]?\s*$", cleaned)
    if not matched:
        return cleaned, ""
    main = cleaned[: matched.start()].strip(" ，,。")
    disclosure = f"披露时间：{matched.group(1)}"
    return main, disclosure


def is_software_description_sentence(sentence: str) -> bool:
    text = clean_text(sentence)
    if not text:
        return False
    if text.startswith(("系统不仅具有", "该系统不仅具有", "产品不仅具有")) or "还具有强大的" in text:
        return True
    if text.startswith(("它不仅仅是", "不仅仅是")):
        return True
    known_product_intro_prefixes = (
        "Microsoft Exchange 是微软公司开发",
        "Microsoft Exchange是微软公司开发",
        "OpenStack Swift是开源对象存储系统",
        "OpenStack Swift 是开源对象存储系统",
    )
    if text.startswith(known_product_intro_prefixes):
        return True
    strong_attack_markers = (
        "漏洞",
        "攻击者",
        "未授权",
        "未经",
        "未认证",
        "未对",
        "用户受控",
        "payload",
        "Payload",
        "nonce",
        "SQL",
        "SSRF",
        "XSS",
        "路径遍历",
        "路径穿越",
        "任意文件",
        "任意代码",
        "任意命令",
        "权限提升",
        "绕过",
        "注入",
        "读取服务器",
        "上传恶意",
        "解压",
        "符号链接",
        "沙箱逃逸",
        "直接拼接",
        "直接将其用于",
    )
    if any(marker in text for marker in strong_attack_markers):
        return False
    if "是国内领先" in text or ("服务商" in text and "攻击者" not in text):
        return True
    clear_description_markers = (
        "是一款",
        "是一个",
        "是一套",
        "是用于",
        "是面向",
        "用于",
        "提供",
        "支持",
        "核心目标是",
    )
    if any(marker in text for marker in clear_description_markers) and not any(
        marker in text for marker in ("漏洞", "攻击者", "未授权", "未经身份验证", "接口存在", "端点存在")
    ):
        return True
    attack_keywords = (
        "漏洞",
        "攻击者",
        "未授权",
        "未经身份验证",
        "注入",
        "读取",
        "上传",
        "泄露",
        "绕过",
        "参数",
        "请求",
        "接口存在",
        "端点",
        "利用该漏洞",
        "构造恶意",
        "触发",
        "导致",
    )
    if any(keyword in text for keyword in attack_keywords):
        return False
    if text.startswith((
        "其核心价值在于",
        "它的核心价值在于",
        "核心价值在于",
        "其核心使命是",
        "它的核心使命是",
        "核心使命是",
        "核心目标是",
        "它主要面向",
        "主要面向",
        "它尤其适合",
        "尤其适合",
        "它能够",
        "该系统具备",
        "该系统提供",
        "该系统面向",
        "该产品提供",
        "它旨在",
        "该平台旨在",
        "该产品旨在",
        "该系统旨在",
        "旨在",
        "它使",
    )):
        return True
    if re.search(r"是[^。！？]{0,40}(?:一款|一个|一种|一套)", text) or "自主研发" in text:
        return True
    description_markers = (
        "是一款",
        "是一个",
        "是一套",
        "是用于",
        "是面向",
        "推出的",
        "用于",
        "负责",
        "支持",
        "核心作用是",
        "核心价值在于",
        "核心使命是",
        "核心目标是",
        "旨在",
        "不仅具有",
        "还具有",
        "专业产品",
    )
    return any(marker in text for marker in description_markers)


def parse_rule_name(name: str) -> Tuple[str, str, str, str]:
    text = clean_text(name).replace("（", "(").replace("）", ")")

    for known_prefix in (WEB_PREFIX, APP_PREFIX, AI_APP_PREFIX):
        if not text.startswith(known_prefix):
            continue
        remainder = text[len(known_prefix) :].strip()
        parts = [part.strip() for part in remainder.split("，") if part.strip()]
        product = parts[0] if parts else remainder
        cve = ""
        endpoint = ""
        vuln = ""
        for part in parts[1:]:
            if re.fullmatch(r"CVE-\d{4}-\d+", part, flags=re.IGNORECASE):
                cve = part.upper()
            elif part.startswith("/") or ACTION_ENTRY_RE.fullmatch(part) or re.fullmatch(r"[A-Za-z0-9._?=&:/$-]+", part):
                endpoint = part
            else:
                vuln = part
        product = normalize_product_name(product)
        return product, endpoint, vuln, cve

    cve = ""
    cve_match = re.search(r"\((CVE-\d{4}-\d+)\)\s*$", text, flags=re.IGNORECASE)
    if cve_match:
        cve = cve_match.group(1).upper()
        text = text[: cve_match.start()].strip()
    else:
        loose_cve = re.search(r"\b(CVE-\d{4}-\d+)\b", text, flags=re.IGNORECASE)
        if loose_cve:
            cve = loose_cve.group(1).upper()
            text = (text[: loose_cve.start()] + text[loose_cve.end() :]).strip("() ")

    text = re.sub(
        r"(?<=[A-Za-z0-9_])-\s*(?=(?:任意|本地|远程|信息|SQL|SSRF|RCE|认证|文件|代码|命令|路径).*(?:漏洞|缺陷|绕过|注入|读取|上传|执行|包含|泄露))",
        " - ",
        text,
        count=1,
    )

    if " - " in text:
        left, vuln = text.rsplit(" - ", 1)
    else:
        left, vuln = text, ""

    left = left.strip()
    vuln = vuln.strip("() ，,")
    product = left
    endpoint = ""

    slash_match = re.search(r"\s+(/.+)$", left)
    if slash_match:
        product = left[: slash_match.start()].strip()
        endpoint = slash_match.group(1).strip()
    else:
        tokens = left.split()
        if len(tokens) >= 2:
            tail = tokens[-1]
            if re.fullmatch(r"[A-Za-z0-9._?=&:/$-]+", tail):
                product = " ".join(tokens[:-1]).strip()
                endpoint = tail

    return normalize_product_name(product or left), endpoint, vuln, cve


def extract_paths(text: str) -> List[str]:
    paths: List[str] = []
    for match in PATH_RE.finditer(clean_text(text)):
        path = match.group(0).strip("`'\"()[]{}<>，,。；;：:、")
        if path and path not in paths:
            paths.append(path)
    return paths


def extract_entry_points(text: str) -> List[str]:
    entries = extract_paths(text)
    for match in ACTION_ENTRY_RE.finditer(clean_text(text)):
        entry = clean_text(match.group(0))
        if entry and entry not in entries:
            entries.append(entry)
    return entries


def normalize_path_key(path: str) -> str:
    cleaned = clean_text(path).strip("`'\"()[]{}<>，,。；;：:、").lower()
    cleaned = re.sub(r"[?#].*$", "", cleaned)
    return cleaned.rstrip("/")


def endpoint_path_related(endpoint: str, desc_path: str) -> bool:
    endpoint_key = normalize_path_key(endpoint)
    desc_key = normalize_path_key(desc_path)
    if not endpoint_key or not desc_key:
        return False
    endpoint_token = endpoint_key.lstrip("/")
    desc_token = desc_key.lstrip("/")
    if endpoint_key == desc_key:
        return True
    endpoint_prefix = endpoint_key.rstrip("/")
    if endpoint_prefix and desc_key.startswith(f"{endpoint_prefix}/"):
        return True
    if endpoint_key.startswith("/") and (
        desc_key.endswith(endpoint_key) or desc_key.endswith(f"/{endpoint_token}")
    ):
        return True
    if desc_key.startswith("/") and (
        endpoint_key.endswith(desc_key) or endpoint_token.endswith(desc_token)
    ):
        return True
    endpoint_tail = endpoint_token.rsplit("/", 1)[-1]
    desc_tail = desc_token.rsplit("/", 1)[-1]
    if endpoint_tail and len(endpoint_tail) >= 4 and endpoint_tail == desc_tail:
        return True
    if not endpoint_key.startswith("/") and endpoint_tail and endpoint_tail in desc_token:
        return True
    return False


def resolve_endpoint_from_desc(endpoint: str, desc: str) -> Tuple[str, bool, List[str]]:
    """Use the full path from desc when it clearly expands the title endpoint.

    Returns `(resolved_endpoint, mismatch_warning, desc_paths)`.
    """
    endpoint = clean_text(endpoint)
    desc_entries = extract_entry_points(desc)
    if not desc_entries:
        return endpoint, False, []

    if endpoint:
        related = [entry for entry in desc_entries if endpoint_path_related(endpoint, entry)]
        if related:
            return max(related, key=len), False, desc_entries
        # Only flag path mismatch when the title already contains a path-like endpoint.
        # Non-path components such as `JavaScript` or `PythonREPLComponent` should not
        # be compared against implementation paths in the description.
        return endpoint, endpoint.startswith("/"), desc_entries

    return desc_entries[0], False, desc_entries


def refine_vulnerability_from_desc(vuln: str, desc: str) -> str:
    """Promote precise vulnerability families from the source description."""
    normalized_vuln = clean_text(vuln)
    normalized_desc = clean_text(desc)
    if not normalized_vuln or not normalized_desc:
        return normalized_vuln
    if re.fullmatch(r"SQL\s*注入漏洞", normalized_vuln, flags=re.IGNORECASE) and re.search(
        r"ClickHouse\s*SQL\s*注入漏洞", normalized_desc, flags=re.IGNORECASE
    ):
        return "ClickHouse SQL注入漏洞"
    return normalized_vuln


def is_hardware_like_product(product: str) -> bool:
    return any(keyword in product for keyword in HARDWARE_PRODUCT_KEYWORDS)


def is_ai_application_product(product: str) -> bool:
    normalized = normalize_product_key(product)
    return any(normalize_product_key(keyword) in normalized for keyword in AI_APPLICATION_PRODUCTS)


def build_standardized_name(name: str, desc: str = "") -> Tuple[str, bool]:
    product, endpoint, vuln, cve = parse_rule_name(name)
    endpoint, path_mismatch, _ = resolve_endpoint_from_desc(endpoint, desc)
    vuln = refine_vulnerability_from_desc(vuln, desc)
    parts = [product]
    if cve:
        parts.append(cve)
    if endpoint:
        parts.append(endpoint)
    if vuln:
        parts.append(vuln)
    if is_ai_application_product(product):
        prefix = AI_APP_PREFIX
    elif is_hardware_like_product(product):
        prefix = APP_PREFIX
    else:
        prefix = WEB_PREFIX
    return prefix + "，".join(parts), path_mismatch


def iter_name_desc_objects(value: Any) -> Iterable[Dict[str, Any]]:
    if isinstance(value, dict):
        if "name" in value and "desc" in value:
            yield value
        for item in value.values():
            yield from iter_name_desc_objects(item)
    elif isinstance(value, list):
        for item in value:
            yield from iter_name_desc_objects(item)


def load_historical_software_descriptions(path: Path | None) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    if not path or not path.exists():
        return mapping
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return mapping

    for item in iter_name_desc_objects(data):
        name = clean_text(item.get("name"))
        desc = clean_text(item.get("desc"))
        if not name or not desc:
            continue
        product, _, _, _ = parse_rule_name(name)
        key = normalize_product_key(product)
        if not key:
            continue
        main_text, _ = split_disclosure_time(desc)
        sentences = split_sentences(main_text)
        if not sentences:
            continue
        software_sentences: List[str] = []
        for sentence in reversed(sentences):
            if is_software_description_sentence(sentence):
                software_sentences.insert(0, sentence)
            else:
                break
        if not software_sentences:
            continue
        software_desc = "".join(
            sentence if sentence.endswith(("。", "！", "？")) else f"{sentence}。"
            for sentence in software_sentences
        )
        mapping[key] = software_desc
    return mapping


def normalize_attack_text(text: str) -> str:
    normalized = clean_text(text)
    normalized = normalized.replace("（", "(").replace("）", ")")
    normalized = normalized.replace("，", ",")
    normalized = normalized.replace("。", ". ")
    normalized = re.sub(r"\s+", " ", normalized).strip()
    normalized = normalized.replace(". ", "。")
    normalized = normalized.replace(",", "，")
    normalized = normalized.replace("上海鸿翼软件", "鸿翼软件")
    normalized = normalized.replace("北京方向标信息", "方向标信息")

    protected_tokens: Dict[str, str] = {}

    def protect(match: re.Match[str]) -> str:
        marker = f"__PROTECTED_TOKEN_{len(protected_tokens)}__"
        protected_tokens[marker] = match.group(0)
        return marker

    # Paths and identifiers are source evidence; do not title-case fragments inside them.
    normalized = re.sub(r"(?<![A-Za-z0-9_])/[A-Za-z0-9._~:/?#\[\]@!$&'()*+,;=%-]+", protect, normalized)
    normalized = re.sub(r"\b[A-Za-z0-9_-]+\.(?:php|aspx|ashx|jsp|ini|json|yaml|yml|xml|txt)\b", protect, normalized, flags=re.IGNORECASE)

    normalized = re.sub(r"sql注入", "SQL注入", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"SQL\s+注入", "SQL注入", normalized)
    normalized = re.sub(r"(?<![A-Za-z])ssrf(?![A-Za-z])", "SSRF", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"(?<![A-Za-z])xss(?![A-Za-z])", "XSS", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"(?<![A-Za-z])rce(?![A-Za-z])", "RCE", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"(?<![A-Za-z])api(?![A-Za-z])", "API", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"(?<![A-Za-z])http(?![A-Za-z])", "HTTP", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"(?<![A-Za-z])url(?![A-Za-z])", "URL", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"(?<![A-Za-z])json(?![A-Za-z])", "JSON", normalized, flags=re.IGNORECASE)
    for marker, token in protected_tokens.items():
        normalized = normalized.replace(marker, token)

    normalized = normalized.strip(" ，,。")
    if normalized and normalized[-1] not in "。！？":
        normalized = f"{normalized}。"
    return normalized


def auth_state_cn(text: str) -> Tuple[bool, bool]:
    """Return `(authenticated, unauthenticated)` for source-aligned auth wording."""
    normalized = clean_text(text)
    unauthenticated = bool(
        re.search(
            r"未授权|未认证|未经认证|未经授权|未经身份(?:认证|验证)|无需(?:身份)?认证|无须(?:身份)?认证|不需要(?:身份)?认证",
            normalized,
        )
    )
    authenticated = bool(
        re.search(
            r"经过身份(?:认证|验证)|经过认证|(?<!未)经身份(?:认证|验证)|(?<!未)经认证|认证用户|已认证用户|已获得登录权限|拥有[^。；，,]{0,24}权限的认证用户|具有[^。；，,]{0,24}权限的?(?:经过身份(?:认证|验证)的?|经过认证的?)?(?:攻击者|用户)|有效凭证|登录权限",
            normalized,
        )
    )
    return authenticated, unauthenticated


def ensure_auth_state_matches_source(source_desc: str, standardized_desc: str, row_idx: int) -> None:
    source_state = auth_state_cn(source_desc)
    standardized_state = auth_state_cn(standardized_desc)
    if source_state != standardized_state:
        raise ValueError(
            "Auth semantics changed while standardizing "
            f"row {row_idx}: source={source_state}, standardized={standardized_state}"
        )


def contains_ascii(text: str) -> bool:
    return bool(re.search(r"[A-Za-z]", text or ""))


def format_target_for_intro(target: str) -> str:
    target = clean_text(target)
    if not target:
        return target
    if contains_ascii(target):
        return f" {target} "
    return target


def format_vulnerability_for_intro(vuln: str) -> str:
    return clean_text(vuln)


def normalize_software_description(text: str) -> str:
    normalized = clean_text(text)
    if not normalized:
        return normalized
    normalized = normalized.replace("为企事业单位", "为企业")
    normalized = normalized.replace("面向企事业单位", "面向企业")
    normalized = normalized.replace("服务于企事业单位", "服务于企业")
    normalized = normalized.replace("面向企业事业单位", "面向企业")
    normalized = normalized.replace("，为企事业单位", "，为企业")
    normalized = normalized.replace(",为企事业单位", "，为企业")
    compact = normalized.replace(" ", "")
    if "9Router" in normalized:
        return "9Router 是一个开源的 AI 路由与代理工具。"
    if "Budibase" in normalized:
        return "Budibase 是一个开源低代码平台。"
    if "Crawl4AI" in normalized or "Crawl4ai" in normalized:
        return "Crawl4AI 是一个面向大型语言模型和 AI 应用的网络爬虫和数据抓取工具。"
    if "Gradio" in normalized:
        return "Gradio 是一个开源的 Python 库，用于构建机器学习、数据科学和 Web 应用演示。"
    if "n8n" in normalized:
        return "n8n 是一款开源的工作流自动化平台。"
    if "Penpot" in normalized:
        return "Penpot 是一个开源设计与原型协作平台。"
    if "XWiki" in normalized:
        return "XWiki 是一个开源的 wiki 平台。"
    if "泛微E-cology" in compact or "泛微Ecology" in compact:
        return "泛微 E-cology 是一款面向中大型组织的协同办公管理平台。"
    if "MicrosoftExchange" in compact:
        return "Microsoft Exchange 是微软公司开发的企业级消息与协作平台。"
    if "方向标邮件网关" in normalized:
        return "方向标邮件网关（FangMail 邮件安全网关）是企业级邮件安全防护产品。"
    if "鸿翼" in normalized and "电子文档管理" in normalized:
        return "鸿翼软件电子文档管理系统为企业提供非结构化数据的全生命周期管控。"
    if "OpenStackSwift" in compact:
        return "OpenStack Swift 是开源对象存储系统。"
    if "同鑫eHR" in compact or "同鑫EHR" in compact:
        return "同鑫 eHR 是一款人力资源管理平台。"
    if "金蝶EAS" in compact:
        return "金蝶 EAS 是面向集团型企业的数字化管理平台。"
    if re.search(r"用友\s*NC", normalized):
        return "用友 NC 是一款企业资源规划产品，提供财务管理、采购管理、销售管理、人力资源管理等功能。"
    compact = normalized.replace(" ", "")
    if "用友时空KSOA" in compact:
        return "用友时空 KSOA 是一款面向流通企业的协同办公与业务管理系统。"
    if "NUUO" in normalized and ("摄像" in normalized or "NVR" in normalized):
        return "NUUO 摄像头是网络视频录像机产品。"
    if "DbGate" in normalized:
        return "DbGate 是一款跨平台数据库管理工具。"
    if "Strapi" in normalized:
        return "Strapi 是一个开源无头内容管理系统 (Headless CMS)。"
    if "dotCMS" in normalized:
        return "dotCMS 是一个面向企业级应用的开源混合型内容管理系统。"
    if "用友政务财务系统" in normalized:
        return "用友政务财务系统是面向政府部门、事业单位、非营利组织的业务管理信息化系统。"
    if "Gitea" in normalized:
        return "Gitea 是一个轻量级、可自行托管的 Git 服务平台。"
    if "OpenEMR" in normalized:
        return "OpenEMR 是一款免费开源的电子健康记录和医疗实践管理应用程序。"
    company_suffix = r"(?:有限公司|科技有限公司|信息技术有限公司|软件有限公司|股份有限公司|有限责任公司)"
    for city in CITY_NAMES:
        normalized = re.sub(rf"^{city}(?=[^，。；;、]{{0,20}}{company_suffix})", "", normalized)
        normalized = re.sub(rf"(?<=是){city}(?=[^，。；;、]{{0,30}}{company_suffix})", "", normalized)
        normalized = re.sub(rf"是{city}(?=[^，。；;、]{{0,30}}{company_suffix})", "是", normalized)
        normalized = re.sub(rf"(?<=[（(、，,\s]){city}(?=[^，。；;、]{{0,20}}{company_suffix})", "", normalized)
    political_markers = (
        "经过充分的客户需求调研",
        "经过充分客户需求调研",
        "并依据国家",
        "依据国家",
        "国家“十三五”",
        "国家十三五",
        "教育信息化建设规范",
    )
    earliest_marker = min(
        (normalized.find(marker) for marker in political_markers if marker in normalized),
        default=-1,
    )
    if earliest_marker >= 0:
        head = normalized[:earliest_marker].rstrip("，,、；; ")
        tail_match = re.search(r"推出的([^。！？]+)", normalized[earliest_marker:])
        if tail_match:
            suffix = tail_match.group(1).strip("，,、；; ")
            normalized = f"{head}推出的{suffix}"
        else:
            normalized = head
    marketing_markers = (
        "功能全面",
        "性能稳定",
        "扩展性强",
        "核心竞争力",
        "护城河",
        "开放、互联、融合、智能",
    )
    earliest_marketing = min(
        (normalized.find(marker) for marker in marketing_markers if marker in normalized),
        default=-1,
    )
    if earliest_marketing >= 0:
        normalized = normalized[:earliest_marketing].rstrip("，,、；; 的")
    normalized = re.sub(r"(?:，|,)?具有$", "", normalized).strip()
    normalized = re.sub(r"(?:，|,)?具有。$", "。", normalized).strip()
    normalized = re.sub(r"\s+", " ", normalized).strip()
    normalized = normalized.replace("（ ", "（").replace("( ", "(")
    normalized = normalized.replace("  ", " ")
    normalized = normalized.strip("，,、；; ")
    if normalized and normalized[-1] not in "。！？":
        normalized = f"{normalized}。"
    return normalized


def remove_redundant_attack_prefix(attack_text: str, product: str, endpoint: str, vuln: str) -> str:
    """Drop a repeated "target + exists vulnerability" prefix already covered by the intro."""
    text = clean_text(attack_text)
    if not text or not vuln:
        return text
    product_text = clean_text(product)
    product_variants = [product_text]
    if product_text.startswith("WordPress "):
        product_variants.append(product_text[len("WordPress ") :])
    product_pattern = "(?:" + "|".join(re.escape(item) for item in product_variants if item) + ")"
    endpoint_pattern = re.escape(clean_text(endpoint))
    vuln_pattern = re.escape(clean_text(vuln))
    if not product_pattern or not vuln_pattern:
        return text
    separators = r"(?:\s+|/|，|,|的|插件\s*)*"
    accessors = r"(?:接口|接口处|端点|方法|函数|配置|API\s*端点|REST\s*API\s*端点|路由)"
    if endpoint_pattern:
        target_pattern = rf"{product_pattern}{separators}{endpoint_pattern}"
    else:
        target_pattern = product_pattern
    pattern = re.compile(
        rf"^{target_pattern}\s*{accessors}?\s*存在{vuln_pattern}，",
        flags=re.IGNORECASE,
    )
    updated = pattern.sub("", text, count=1).strip()
    if updated != text:
        return updated or text

    endpoint_text = clean_text(endpoint)
    if endpoint_text:
        endpoint_loose = re.escape(endpoint_text).replace(r"\ ", r"\s+")
        version_exists_pattern = re.compile(
            rf"^({product_pattern}\s*[^，。；;]{{0,80}}?(?:版本|之前版本|及之前版本))"
            rf"(?:的|通过)\s*{endpoint_loose}\s*"
            rf"(?:REST\s*API\s*)?(?:端点|接口|AJAX\s*Action)?\s*存在"
            rf"[^，。；;]{{0,80}}?(?:漏洞|缺陷)(?:\([A-Za-z0-9+\-\s]+\))?[，。]",
            flags=re.IGNORECASE,
        )
        updated = version_exists_pattern.sub(r"\1中，", text, count=1).strip()
        if updated != text:
            return updated or text

    # Fallback for rows whose source uses a more precise path than the title,
    # e.g. title `/public` but source `/api/upload/public`. Remove only the
    # first comma-delimited clause and keep the attacker condition/impact.
    first_clause, sep, rest = text.partition("，")
    if sep and rest:
        clause_paths = re.findall(r"(?<![A-Za-z0-9_])/[A-Za-z0-9._~:/?#\[\]@!$&'()*+,;=%-]+", first_clause)
        endpoint_text = clean_text(endpoint)
        if endpoint_text and any(path != endpoint_text for path in clause_paths):
            return text
        endpoint_key = re.sub(r"[^a-z0-9]+", "", clean_text(endpoint).lower())
        clause_key = re.sub(r"[^a-z0-9]+", "", first_clause.lower())
        vuln_key = re.sub(r"[\s，,]+", "", clean_text(vuln).lower())
        clause_vuln_key = re.sub(r"[\s，,]+", "", first_clause.lower())
        has_version_context = bool(re.search(r"(?:版本|<=|>=|<|>|及之前|之前版本|以下版本)", first_clause))
        has_detail_context = any(
            marker in first_clause
            for marker in (
                "参数",
                "过滤",
                "拼接",
                "多个模块",
                "处理器",
                "盲注",
                "未对",
                "由于",
                "直接",
            )
        )
        if (
            not has_version_context
            and not has_detail_context
            and product in first_clause
            and "存在" in first_clause
            and vuln in first_clause
            and (not endpoint_key or endpoint_key in clause_key)
        ):
            return rest.strip() or text
        if (
            not has_version_context
            and not has_detail_context
            and product in first_clause
            and "存在" in first_clause
            and vuln in first_clause
            and len(first_clause) <= 80
        ):
            return rest.strip() or text
        if (
            not has_version_context
            and not has_detail_context
            and endpoint_key
            and endpoint_key in clause_key
            and vuln_key
            and vuln_key in clause_vuln_key
            and "存在" in first_clause
        ):
            return rest.strip() or text
        if (
            not has_version_context
            and not has_detail_context
            and endpoint_key
            and endpoint_key in clause_key
            and re.search(r"存在[^，。；;]{0,50}(?:漏洞|缺陷)", first_clause)
        ):
            return rest.strip() or text

    return text


def build_standardized_desc(
    name: str,
    desc: str,
    historical_desc_map: Dict[str, str],
    endpoint_override: str | None = None,
) -> str:
    product, endpoint, vuln, _ = parse_rule_name(name)
    vuln = refine_vulnerability_from_desc(vuln, desc)
    if endpoint_override is not None:
        endpoint = endpoint_override
    main_text, disclosure = split_disclosure_time(desc)
    sentences = split_sentences(main_text)

    software_sentences: List[str] = []
    attack_sentences: List[str] = []
    for sentence in sentences:
        if is_software_description_sentence(sentence):
            software_sentences.append(sentence)
            continue
        attack_sentences.append(sentence)

    software_desc = "".join(
        sentence if sentence.endswith(("。", "！", "？")) else f"{sentence}。"
        for sentence in software_sentences
    )
    software_desc = normalize_software_description(software_desc)
    historical_software_desc = historical_desc_map.get(normalize_product_key(product), "")
    if historical_software_desc:
        software_desc = historical_software_desc
    software_desc = normalize_software_description(software_desc)

    attack_text = normalize_attack_text("".join(attack_sentences))
    attack_text = remove_redundant_attack_prefix(attack_text, product, endpoint, vuln)
    target = " ".join(part for part in (product, endpoint) if part).strip()
    intro_target = format_target_for_intro(target)
    intro = (
        f"此检测规则还原了针对{intro_target}存在的{format_vulnerability_for_intro(vuln)}的利用尝试"
        if vuln
        else f"此检测规则还原了针对{intro_target}的利用尝试"
    )

    parts: List[str] = [intro]
    if attack_text:
        parts.append(attack_text)
    if disclosure:
        parts.append(f"{disclosure}。")
    if software_desc:
        parts.append(software_desc)

    final_text = parts[0]
    for index, part in enumerate(parts[1:], start=1):
        part = clean_text(part)
        if not part:
            continue
        if index == 1:
            final_text = f"{final_text}。{part}"
        else:
            final_text = f"{final_text}{part}"
    return final_text.replace("。。", "。").replace("，，", "，").strip()


def normalize_notes(notes: str, context: str = "", batch_vendor_urls: Dict[str, str] | None = None) -> str:
    text = clean_text(notes)
    fallback_url = ""
    product, _, _, _ = parse_rule_name(context)
    if batch_vendor_urls:
        fallback_url = batch_vendor_urls.get(normalize_product_key(product), "")
    for product, url in HISTORICAL_VENDOR_URLS.items():
        if not fallback_url and product in context:
            fallback_url = url
            break
    if not text:
        return f"塞讯验证建议：\n请关注厂商主页获取更新：\n{fallback_url}" if fallback_url else "塞讯验证建议：\n请关注厂商主页获取更新："
    url_match = re.search(r"(https?://\S+)", text)
    url = url_match.group(1) if url_match else re.sub(r"^请关注厂商主页获取更新[:：]?\s*", "", text).strip()
    if not url and fallback_url:
        url = fallback_url
    return f"塞讯验证建议：\n请关注厂商主页获取更新：\n{url}"


def collect_batch_vendor_urls(ws: Any, header_index: Dict[str, int]) -> Dict[str, str]:
    urls: Dict[str, str] = {}
    if "name.1" not in header_index or "notes" not in header_index:
        return urls
    for row_idx in range(2, ws.max_row + 1):
        name = clean_text(ws.cell(row=row_idx, column=header_index["name.1"]).value)
        notes = clean_text(ws.cell(row=row_idx, column=header_index["notes"]).value)
        url_match = re.search(r"(https?://\S+)", notes)
        if not name or not url_match:
            continue
        product, _, _, _ = parse_rule_name(name)
        key = normalize_product_key(product)
        if key and key not in urls:
            urls[key] = url_match.group(1)
    return urls


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Standardize detection Excel fields.")
    parser.add_argument("--input", required=True, help="Input .xlsx path")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--sheet", default=None, help="Sheet name; defaults to the first sheet")
    parser.add_argument("--historical-json", default=None, help="Optional historical JSON export for software descriptions")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    historical_path = Path(args.historical_json).expanduser().resolve() if args.historical_json else None

    historical_desc_map = load_historical_software_descriptions(historical_path)
    wb = load_workbook(input_path)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    header_row = [cell.value for cell in ws[1]]
    header_index = {str(value): idx + 1 for idx, value in enumerate(header_row)}
    required = {"name.1", "desc", "notes"}
    missing = [key for key in required if key not in header_index]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    batch_vendor_urls = collect_batch_vendor_urls(ws, header_index)

    for row_idx in range(2, ws.max_row + 1):
        original_name = clean_text(ws.cell(row=row_idx, column=header_index["name.1"]).value)
        original_desc = clean_text(ws.cell(row=row_idx, column=header_index["desc"]).value)
        original_notes = clean_text(ws.cell(row=row_idx, column=header_index["notes"]).value)

        standardized_name, path_mismatch = build_standardized_name(original_name, original_desc)
        _, resolved_endpoint, _, _ = parse_rule_name(standardized_name)
        name_cell = ws.cell(row=row_idx, column=header_index["name.1"])
        desc_cell = ws.cell(row=row_idx, column=header_index["desc"])
        name_cell.value = standardized_name
        standardized_desc = build_standardized_desc(
            original_name,
            original_desc,
            historical_desc_map,
            endpoint_override=resolved_endpoint,
        )
        ensure_auth_state_matches_source(original_desc, standardized_desc, row_idx)
        desc_cell.value = standardized_desc
        if path_mismatch:
            name_cell.fill = WARNING_FILL
            desc_cell.fill = WARNING_FILL
        ws.cell(
            row=row_idx,
            column=header_index["notes"],
            value=normalize_notes(original_notes, f"{original_name} {original_desc}", batch_vendor_urls),
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(json.dumps({"output": str(output_path), "historical_desc_count": len(historical_desc_map)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
